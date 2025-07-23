import csv, os
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from dateutil.relativedelta import relativedelta

from django.shortcuts import get_object_or_404
import openpyxl
from django.db.models import (
    F, Prefetch, Q, Sum, Value,
)
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import (
    decorators, parsers, response,
    status, filters, exceptions, viewsets
)
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.decorators import action
from account.models import Income
from apps.mixin.permission import IsSuperUser, ModelPermission
from apps.mixin.views import BaseModelViewSet
from employee.models.employee import Employee, EmployeeUnderTPM
from employee.models.employee_activity import EmployeeProject
from project_management.models import (
    Client,
    ClientReview,
    Country,
    CurrencyType,
    DailyProjectUpdate,
    DailyProjectUpdateAttachment,
    DailyProjectUpdateHistory,
    InvoiceType,
    PaymentMethod,
    Project,
    ProjectContent,
    ProjectHour,
    ProjectResource,
    Teams,
)

from .filters import DailyProjectUpdateFilter, ProjectHourFilter
from .serializers import (
    AddEmployeesSerializer,
    AddProjectsSerializer,
    BulkDailyUpdateSerializer,
    BulkUpdateSerializer,
    ClientBaseModelSerializer,
    ClientModelSerializer,
    ClientReviewModelSerializer,
    CountryModelSerializer,
    CurrencyTypeModelSerializer,
    DailyProjectUpdateCreateSerializer,
    DailyProjectUpdateListSerializer,
    DailyProjectUpdateSerializer,
    EmployeeUpdateSerializer,
    IncomeSerializer,
    InvoiceTypeModelSerializer,
    PaymentMethodModelSerializer,
    ProjectResourceAddDeleteSerializer,
    ProjectResourceModelSerializer,
    ProjectSerializer,
    ProjectUpdateFilterSerializer,
    ProjectUpdateSerializer,
    StatusUpdateSerializer,
    TeamSerializer,
    TeamUpdateSerializer,
    WeeklyProjectHoursSerializer,
    WeeklyProjectUpdate,
)
from rest_framework.response import Response
from datetime import datetime, timedelta
from django.db.models import Sum, Prefetch
# from .models import Project, DailyProjectUpdate, DailyProjectUpdateAttachment, ProjectHour
# from .serializers import WeeklyProjectHoursRequestSerializer, WeeklyProjectHoursResponseSerializer
from rest_framework.permissions import IsAuthenticated

class ProjectViewSet(BaseModelViewSet):
    http_method_names = ["get", "post"]
    queryset = Project.objects.filter(
        active=True,
    ).prefetch_related(
        "projectcontent_set",
    )
    serializer_class = ProjectSerializer
    # permission_classes = [IsAuthenticated]
    # @property
    # def paginator(self):
    #     if self.action == 'list':
    #         return None
    #     return super().paginator



class DailyProjectUpdateViewSet(BaseModelViewSet):
    queryset = (
        DailyProjectUpdate.objects.select_related("employee", "manager", "project")
        .prefetch_related(
            Prefetch(
                "history",
                queryset=DailyProjectUpdateHistory.objects.order_by("-id"),
            ),
            "dailyprojectupdateattachment_set",
            "employee__leave_set",
            "project__client",
        )
        .all().order_by("-created_at")
    )

    serializer_class = DailyProjectUpdateSerializer
    filterset_class = DailyProjectUpdateFilter
    search_fields = (
        "employee__full_name",
        "project__title",
        "manager__full_name",
    )
    serializers = {
        "status_update": StatusUpdateSerializer,
        "create": DailyProjectUpdateCreateSerializer,
        # "list": DailyProjectUpdateListSerializer
    }

    def get_serializer_class(self):
        if self.action in ["export_update", "export_excel", "export_excel_merged"]:
            return BulkDailyUpdateSerializer
        return super().get_serializer_class()

    @swagger_auto_schema(
        responses={201: DailyProjectUpdateSerializer}
    )
    def create(self, request, *args, **kwargs):
        self.parser_classes = [
            parsers.MultiPartParser,
            parsers.FormParser,
        ]
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        # Re-serialize using the output serializer
        read_serializer = self.serializer_class(
            instance=serializer.instance,
            context=self.get_serializer_context()
        )
        headers = self.get_success_headers(read_serializer.data)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @decorators.action(detail=False, methods=["PATCH"], url_path="status-update")
    def status_update(self, request, *args, **kwargs):
        """
        this view method accept list of daily project id and update status
        """
        serializer = self.get_serializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        queryset = self.filter_queryset(
            self.get_queryset()
            .filter(id__in=data["update_ids"])
            .select_related("employee", "project")
            .prefetch_related("history", "dailyprojectupdateattachment_set")
        )
        updates = list(queryset)
        for update in updates:
            update.status = data["status"]
        DailyProjectUpdate.objects.bulk_update(updates, ["status"])
        return Response(DailyProjectUpdateSerializer(queryset, many=True).data)

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "project_hour_id",
                openapi.IN_QUERY,
                description="Project ID",
                type=openapi.TYPE_INTEGER,
            ),
        ]
    )
    @decorators.action(
        detail=False,
        methods=["GET"],
        url_path="(?P<project_id>[^/.]+)/daily-update/(?P<date>[^/.]+)",
    )
    def get_daily_update_by_weekly(self, request, project_id, date, *args, **kwargs):
        """
        this view accept project id and date then response all daily project update base on the project hour manager id

        """
        project_hour_id = request.GET.get("project_hour_id")
        manager_id = request.user.employee.id
        if project_hour_id:
            manager_id = (
                ProjectHour.objects.select_related("manager")
                .get(id=project_hour_id)
                .manager_id
            )
        q_obj = Q(
            project=project_id,
            manager=manager_id,
            status="approved",
            created_at__date__lte=datetime.strptime(date, "%Y-%m-%d"),
            created_at__date__gte=datetime.strptime(date, "%Y-%m-%d")
            - timedelta(days=6),
        )
        daily_update = (
            self.get_queryset()
            .filter(
                q_obj,
                employee__active=True,
                employee__project_eligibility=True,
            )
            .annotate(
                full_name=F("employee__full_name"), update_text=F("updates_json__0__0")
            )
            .exclude(hours=0.0)
            .values(
                "id",
                "created_at",
                "employee_id",
                "full_name",
                "hours",
                "update",
                "updates_json",
                "update_text",
            )
        )

        totalHours = daily_update.aggregate(total=Sum("hours"))["total"] or 0

        data = {
            "manager_id": manager_id,
            "weekly_hour": list(daily_update),
            "total_project_hours": totalHours,
        }

        return response.Response(data)

    @decorators.action(detail=False, methods=["POST"], url_path="export-text")
    def export_update(self, request):
        """
        export selected update into txt file
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        queryset = (
            self.get_queryset()
            .filter(id__in=serializer.validated_data.get("update_ids", []))
            .select_related("employee", "project")
            .values(
                "employee__full_name", "updates_json", "created_at", "project__title"
            )
        )
        queryset = self.filter_queryset(queryset)

        if not queryset:
            return Response(
                {"error": "No records found for export."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        total_hours = queryset.aggregate(total=Sum("hours"))["total"] or 0

        updates_by_employee = {}
        for update in queryset:
            if not update["updates_json"]:
                continue

            emp_name = update["employee__full_name"]
            if emp_name not in updates_by_employee:
                updates_by_employee[emp_name] = []

            updates_by_employee[emp_name].extend(update["updates_json"])

        content = [
            "Today's Update\n",
            "-----------------\n",
            f"{queryset[0]['created_at'].strftime('%d-%m-%Y')}\n\n",
            f"Total Hours: {round(total_hours, 3)}H\n\n",
        ]

        for emp_name, updates in updates_by_employee.items():
            content.extend(
                [
                    f"{emp_name}\n\n",
                    "\n".join(f"{u[0]} - {u[1]}H." for u in updates),
                    "\n\nAssociated Links:\n",
                    "\n".join(f"{i+1}. {u[2]}" for i, u in enumerate(updates)),
                    "\n-------------------------------------------------------------\n\n",
                ]
            )

        response = FileResponse(
            BytesIO("".join(content).encode("utf-8")), content_type="text/plain"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{queryset[0]["project__title"].replace(" ", "_")}_'
            f'{queryset[0]["created_at"].strftime("%d-%m-%Y")}.txt"'
        )
        return response

    @decorators.action(detail=False, methods=["POST"], url_path="export-excel")
    def export_excel(self, request):
        """
        export selected update into excel file
        """
        serialize = self.get_serializer(data=request.data)
        serialize.is_valid(raise_exception=True)
        product_ids = serialize.validated_data.get("update_ids", [])

        queryset = self.get_queryset().filter(id__in=product_ids)
        project_name = queryset[0].project.title.replace(" ", "_")
        start_date = request.GET.get("created_at__date__gte", "not_selected")
        end_date = request.GET.get("created_at__date__lte", "not_selected")
        date_range = (
            f"{start_date}_to_{end_date}"
            if start_date != "not_selected" and end_date != "not_selected"
            else "selective"
        )

        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            f'attachment; filename="{project_name}__{date_range}__exported.xlsx"'
        )

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions["A"].width = 13
        sheet.column_dimensions["B"].width = 98
        sheet.column_dimensions["C"].width = 13
        sheet.column_dimensions["D"].width = 13
        sheet.column_dimensions["E"].width = 20

        # Add headers to sheet
        sheet.append(["  Date  ", "  Updates  ", "  Task Hours  ", "  Day Hours  "])

        start_ref = 1
        total_hours = 0
        for index, obj in enumerate(queryset, start=2):
            if (
                obj.updates_json is None
                or obj.updates_json.__len__() == 0
                or obj.status != "approved"
            ):
                continue
            total_hours += obj.hours
            for index_update, update in enumerate(obj.updates_json):
                sheet.append(
                    [
                        obj.created_at.strftime("%d-%m-%Y"),
                        update[0],
                        update[1],
                        obj.hours,
                    ]
                )

            start_merge = 1 + start_ref
            end_merge = start_merge + index_update
            start_ref = end_merge
            date_cells = f"A{start_merge}:A{end_merge}"
            day_hour_cells = f"D{start_merge}:D{end_merge}"
            sheet.merge_cells(date_cells)
            sheet.merge_cells(day_hour_cells)

        sheet.append(["", "", "Total: ", f"{total_hours} Hours"])

        # Add styles
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(cell):
                cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                cell.fill = PatternFill(
                    start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", indent=0
                )

        for cell in sheet.iter_rows(min_row=2):
            for index, cell in enumerate(cell):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if index == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="top")

        for cell in sheet.iter_rows(min_row=sheet.max_row):
            for index, cell in enumerate(cell):
                if index == 3 or index == 2:
                    cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                    cell.fill = PatternFill(
                        start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", indent=0
                    )

        # Save the workbook to the response
        wb.save(response)

        return response

    @decorators.action(detail=False, methods=["POST"], url_path="export-merged-excel")
    def export_excel_merged(self, request):
        serialize = self.get_serializer(data=request.data)
        serialize.is_valid(raise_exception=True)
        product_ids = serialize.validated_data.get("update_ids", [])
        queryset = self.get_queryset().filter(id__in=product_ids)
        merged_set = []

        for obj in queryset:
            if merged_set.__len__() > 0:
                if merged_set[-1].get("created_at").date() == obj.created_at.date():
                    tmp_obj = {
                        "created_at": obj.created_at,
                        "updates_json": merged_set[-1].get("updates_json")
                        + obj.updates_json,
                        "hours": obj.hours + merged_set[-1].get("hours"),
                        "status": obj.status,
                    }
                    merged_set[-1] = tmp_obj
                else:
                    tmp_obj = {
                        "created_at": obj.created_at,
                        "updates_json": obj.updates_json,
                        "hours": obj.hours,
                        "status": obj.status,
                    }
                    merged_set.append(tmp_obj)
            else:
                tmp_obj = {
                    "created_at": obj.created_at,
                    "updates_json": obj.updates_json,
                    "hours": obj.hours,
                    "status": obj.status,
                }
                merged_set.append(tmp_obj)

        project_name = queryset[0].project.title.replace(" ", "_")
        start_date = request.GET.get("created_at__date__gte", "not_selected")
        end_date = request.GET.get("created_at__date__lte", "not_selected")
        date_range = (
            f"{start_date}_to_{end_date}"
            if start_date != "not_selected" and end_date != "not_selected"
            else "selective"
        )

        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            f'attachment; filename="{project_name}__{date_range}__exported.xlsx"'
        )

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions["A"].width = 13
        sheet.column_dimensions["B"].width = 98
        sheet.column_dimensions["C"].width = 13
        sheet.column_dimensions["D"].width = 13
        sheet.column_dimensions["E"].width = 20
        sheet.append(["  Date  ", "  Updates  ", "  Task Hours  ", "  Day Hours  "])

        start_ref = 1
        total_hours = 0
        for index, obj in enumerate(merged_set, start=2):
            if (
                obj.get("updates_json") is None
                or obj.get("updates_json").__len__() == 0
                or obj.get("status") != "approved"
            ):
                continue
            total_hours += obj.get("hours")
            for index_update, update in enumerate(obj.get("updates_json")):
                sheet.append(
                    [
                        obj.get("created_at").strftime("%d-%m-%Y"),
                        update[0],
                        update[1],
                        obj.get("hours"),
                    ]
                )

            start_merge = 1 + start_ref
            end_merge = start_merge + index_update
            start_ref = end_merge
            date_cells = f"A{start_merge}:A{end_merge}"
            day_hour_cells = f"D{start_merge}:D{end_merge}"
            sheet.merge_cells(date_cells)
            sheet.merge_cells(day_hour_cells)

        sheet.append(["", "", "Total: ", f"{total_hours} Hours"])

        # Add styles
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(cell):
                cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                cell.fill = PatternFill(
                    start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", indent=0
                )

        for cell in sheet.iter_rows(min_row=2):
            for index, cell in enumerate(cell):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if index == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="top")

        for cell in sheet.iter_rows(min_row=sheet.max_row):
            for index, cell in enumerate(cell):
                if index == 3 or index == 2:
                    cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                    cell.fill = PatternFill(
                        start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", indent=0
                    )

        wb.save(response)
        return response

    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                "created_at_after",
                openapi.IN_QUERY,
                description="created_at from date",
                type=openapi.FORMAT_DATE,
            ),
            openapi.Parameter(
                "created_at_before",
                openapi.IN_QUERY,
                description="created_at to date",
                type=openapi.FORMAT_DATE,
            ),
        ]
    )
    @decorators.action(
        detail=False,
        methods=["GET"],
        url_path="daily-update-missing-employees",
    )
    def get_employees_without_project_updates(self, request):
        """
        Returns a paginated list of employees who did not submit
        their daily project update for each working day.
        """

        # current_date_time = timezone.now()
        # start_date_time = request.query_params.get("created_at_after", (current_date_time-relativedelta(months=1)))
        # end_date_time = request.query_params.get("created_at_before", current_date_time)
        
        # working_days = self.get_working_days(start_date_time, end_date_time)
        # update_not_provided_by_employes = []
        # employees = Employee.objects.all()
        # for working_day in working_days:
        #     daily_project_update_employees_id = DailyProjectUpdate.objects.filter(
        #         created_at__date=working_day.date(),
        #     ).values_list("employee_id", flat=True)
        #     # print('daily_project_updates', daily_project_update_employees_id)

        #     missing_employees = employees.exclude(
        #         pk__in=daily_project_update_employees_id
        #     ).annotate(date=Value(working_day.date()))
        #     update_not_provided_by_employes.extend(missing_employees)

        query_set = self.get_queryset().filter(
            history__isnull=True,
            employee__operation=False,
        ).exclude(
            employee_id=os.environ.get("SUPERUSER_ID", 30)
        )
        filtered_queryset = self.filter_queryset(query_set)
        paginator = LimitOffsetPagination()
        paginated_list = paginator.paginate_queryset(
            queryset=filtered_queryset,
            request=request,
        )
        return paginator.get_paginated_response(
            data=self.serializer_class(instance=paginated_list, many=True).data
        )

    # def get_working_days(self, start_date_time: datetime, end_date_time: datetime):
    #     working_days = []
    #     while start_date_time <= end_date_time:
    #         if start_date_time.weekday() < 5:  # 5=Saturday, 6=Sunday, Skip Saturday/Sunday
    #             working_days.append(start_date_time)
    #         start_date_time += timedelta(days=1)
    #     return working_days


class WeeklyProjectUpdateViewSet(BaseModelViewSet):
    queryset = ProjectHour.objects.select_related(
        "project", "manager", "tpm"
    ).prefetch_related(
        "employeeprojecthour_set",
        "projecthourhistry_set",
        "employeeprojecthour_set__employee",
    )
    serializer_class = WeeklyProjectUpdate
    filterset_class = ProjectHourFilter

    def get_permissions(self):
        if self.action in ["enable_payable_status", "disable_payable_status"]:
            return [IsSuperUser()]
        return super().get_permissions()

    def get_serializer_class(self, *args, **kwargs):
        if self.action in [
            "approve_project_hours",
            "export_as_csv",
            "enable_payable_status",
            "disable_payable_status",
            "create_income",
        ]:
            return BulkUpdateSerializer
        return super().get_serializer_class(*args, **kwargs)

    def get_queryset(self):
        qs = super().get_queryset()
        # print(dir(qs.first()))
        user = self.request.user
        if user.is_superuser or user.has_perm("project_management.show_all_hours"):
            return qs
        employee = user.employee
        if employee.manager:
            qs = qs.filter(manager=employee)
        elif employee.is_tpm:
            qs = qs.filter(tpm=employee)
        return qs

    @decorators.action(
        detail=False,
        methods=["POST"],
        url_path="approve-hours",
    )
    def approve_project_hours(self, request):
        """
        Approve multiple project hours based on user permissions.
        Expects a list of project hour IDs in the request data.
        """
        user = request.user

        if not hasattr(user, "employee"):
            return Response(
                {"detail": "User has no employee profile"},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_hour_ids = serializer.validated_data.get("ids", [])
        if not project_hour_ids:
            return Response(
                {"detail": "No project hour IDs provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(id__in=project_hour_ids)
        employee = user.employee

        try:
            if employee.is_tpm:
                tpm_projects = (
                    EmployeeUnderTPM.objects.select_related(
                        "employee", "tpm", "project"
                    )
                    .filter(tpm_id=employee.id)
                    .values_list("project_id", flat=True)
                    .distinct()
                )

                update_queryset = queryset.filter(project_id__in=tpm_projects)
                updated_count = update_queryset.update(status="approved")

            elif user.is_superuser:
                updated_count = queryset.update(status="approved")
            else:
                return Response(
                    {"detail": "Insufficient permissions"},
                    status=status.HTTP_403_FORBIDDEN,
                )

            return Response(
                {
                    "detail": f"Successfully approved {updated_count} project hours",
                    "updated_count": updated_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"detail": f"Error approving project hours: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @decorators.action(detail=False, methods=["POST"], url_path="export-csv")
    def export_as_csv(self, request):
        """
        Export selected project hours as CSV.
        Expects a list of project hour IDs in the request data.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_hour_ids = serializer.validated_data.get("ids", [])

        if project_hour_ids:
            queryset = self.get_queryset().filter(id__in=project_hour_ids)
        else:
            queryset = self.get_queryset()

        if not queryset.exists():
            return Response(
                {"detail": "No project hours found for export"},
                status=status.HTTP_404_NOT_FOUND,
            )

        response = HttpResponse(
            content_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="project_hours.csv"'},
        )

        writer = csv.writer(response)

        writer.writerow(["Date", "Project", "Hours", "Payment", "Manager"])

        total = 0
        for project_hour in queryset:
            payment = project_hour.hours * 10
            total += payment
            writer.writerow(
                [
                    project_hour.date,
                    str(project_hour.project),
                    project_hour.hours,
                    payment,
                    str(project_hour.manager) if project_hour.manager else "N/A",
                ]
            )

        writer.writerow(["", "Total", "", total, ""])

        return response

    @decorators.action(
        detail=False,
        methods=["post"],
        url_path="enable-payable",
    )
    def enable_payable_status(self, request):
        """
        Enable payable status for selected project hours.
        Only accessible to superusers.
        Expects a list of project hour IDs in the request data.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_hour_ids = serializer.validated_data.get("ids", [])
        if not project_hour_ids:
            return Response(
                {"detail": "No project hour IDs provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:

            queryset = self.get_queryset().filter(id__in=project_hour_ids)
            updated_count = queryset.update(payable=True)

            if updated_count == 0:
                return Response(
                    {"detail": "No project hours were updated"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            return Response(
                {
                    "detail": f"Successfully enabled payable status for {updated_count} project hours",
                    "updated_count": updated_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"detail": f"Error updating payable status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @decorators.action(detail=False, methods=["post"], url_path="disable-payable")
    def disable_payable_status(self, request):
        """
        Disable payable status for selected project hours.
        Only accessible to superusers.
        Expects a list of project hour IDs in the request data.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_hour_ids = serializer.validated_data.get("ids", [])
        if not project_hour_ids:
            return Response(
                {"detail": "No project hour IDs provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:

            queryset = self.get_queryset().filter(id__in=project_hour_ids)
            updated_count = queryset.update(payable=False)

            if updated_count == 0:
                return Response(
                    {"detail": "No project hours were updated"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            return Response(
                {
                    "detail": f"Successfully disabled payable status for {updated_count} project hours",
                    "updated_count": updated_count,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"detail": f"Error updating payable status: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @decorators.action(
        detail=False,
        methods=["POST"],
        url_path="create-income",
    )
    def create_income(self, request):
        """
        Create Income records from selected project hours.
        Only accessible to superusers.
        Expects a list of project hour IDs in the request data.
        """

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        project_hour_ids = serializer.validated_data.get("ids", [])
        if not project_hour_ids:
            return Response(
                {"detail": "No project hour IDs provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            queryset = self.get_queryset().filter(id__in=project_hour_ids)
            if not queryset.exists():
                return Response(
                    {"detail": "No valid project hours found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            income_objects = []
            convert_rate = Decimal("90.0")

            for project_hour in queryset:
                project = project_hour.project

                hourly_rate = (
                    project.hourly_rate
                    if project.hourly_rate and project.hourly_rate > 0
                    else (
                        project.client.hourly_rate
                        if project.client
                        else Decimal("0.00")
                    )
                )

                hours = project_hour.hours or 0
                payment = Decimal(hours) * Decimal(hourly_rate) * convert_rate

                pdf_url = None
                if project_hour.report_file:
                    pdf_url = (
                        project_hour.report_file.url
                        if "http" in project_hour.report_file.url
                        else request.build_absolute_uri(project_hour.report_file.url)
                    )

                income_obj = Income(
                    project=project,
                    hours=hours,
                    hour_rate=hourly_rate,
                    convert_rate=convert_rate,
                    date=project_hour.date,
                    status="pending",
                    payment=payment,
                    pdf_url=pdf_url,
                )
                income_objects.append(income_obj)

            created_records = Income.objects.bulk_create(income_objects)
            created_count = len(created_records)

            return Response(
                {
                    "detail": f"Successfully created {created_count} income records",
                    "data": IncomeSerializer(created_records, many=True).data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"detail": f"Error creating income records: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )


class ProjectResourceListView(BaseModelViewSet):
    
    http_method_names = [
        "get", "post",
    ]
    serializer_class = ProjectResourceModelSerializer
    serializers = {
        "add_resource_in_a_project": ProjectResourceAddDeleteSerializer,
        "remove_resource_from_a_project": ProjectResourceAddDeleteSerializer,
    }
    filter_backends = [
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    ordering_fields = [
        "id",
        "title",
    ]
    search_fields = [
        "title",
        "web_title",
    ]
    queryset = Project.objects.filter(
        active=True,
    ).prefetch_related(
        Prefetch(
            lookup="employeeproject_set",
            queryset=EmployeeProject.objects.select_related("employee"),
            to_attr="employeeprojects",
        )
    ).all()

    def create(self, request, *args, **kwargs):
        raise exceptions.MethodNotAllowed(method="create")

    @decorators.action(detail=False, methods=["post"])
    def add_resource_in_a_project(self, request):
        """
        add resources in a Project.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        employees = serializer.validated_data.get("employee_ids")
        project = serializer.validated_data.get("project_id")
        for employee in employees:
            employee_project, is_created = EmployeeProject.objects.get_or_create(
                employee=employee,
                defaults={
                    "project": project,
                }
            )
            if is_created is False:
                employee_project.project.add(project)
        return Response(data={"result": f"{len(employees)} Resource added in the '{project.title}' project."})
    
    @decorators.action(detail=False, methods=["post"])
    def remove_resource_from_a_project(self, request):
        """
        remove resources from a project.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        employees = serializer.validated_data.get("employee_ids")
        project = serializer.validated_data.get("project_id")
        for employee in employees:
            employee_project = EmployeeProject.objects.get(
                employee=employee,
            )
            employee_project.project.remove(project)
        return Response(data={"result": f"{len(employees)} Resource remove from '{project.title}' project."})



class ClientViewSet(BaseModelViewSet):
    http_method_names = [
        "get", "post", "put", "patch",
    ]
    queryset = Client.objects.select_related(
        "country", "payment_method", "invoice_type",
        "currency",
    ).prefetch_related(
        "review",
        "clientinvoicedate_set",
    )
    serializer_class = ClientBaseModelSerializer

    serializers = {
        "update": ClientModelSerializer,
        "partial_update": ClientModelSerializer
    }


class CountryViewSet(BaseModelViewSet):
    http_method_names = [
        "get", "post", "put",
    ]
    serializer_class = CountryModelSerializer
    queryset = Country.objects.all()

    search_fields = [
        "name",
    ]


class CurrencyListView(ListAPIView):
    permission_classes = [ModelPermission, IsAuthenticated]
    filter_backends = [
        filters.SearchFilter,
    ]
    search_fields = [
        "currency_name", "currency_code",
    ]
    serializer_class = CurrencyTypeModelSerializer
    queryset = CurrencyType.objects.filter(
        is_active=True,
    )


class PaymentMethodListView(ListAPIView):
    permission_classes = [ModelPermission, IsAuthenticated]
    filter_backends = [
        filters.SearchFilter,
    ]
    search_fields = ["name",]
    serializer_class = PaymentMethodModelSerializer
    queryset = PaymentMethod.objects.all()


class InvoiceTypeListView(ListAPIView):
    permission_classes = [ModelPermission, IsAuthenticated]
    filter_backends = [
        filters.SearchFilter,
    ]
    search_fields = ["name",]
    serializer_class = InvoiceTypeModelSerializer
    queryset = InvoiceType.objects.all()


class ClientReviewViewSet(BaseModelViewSet):
    http_method_names = [
        "get", "post", "put",
    ]
    search_fields = ["name",]
    serializer_class = ClientReviewModelSerializer
    queryset = ClientReview.objects.all()



class TeamViewSet(viewsets.ModelViewSet):
    """
    API ViewSet for managing Teams.
    Supports CRUD operations and actions to add/remove projects and employees.
    """
    queryset = Teams.objects.all().prefetch_related('projects', 'employees')
    serializer_class = TeamSerializer
    http_method_names = ['get', 'post', 'patch', 'delete']
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    serializer_classes = {
        'update': TeamUpdateSerializer,
        'partial_update': TeamUpdateSerializer,
        'add_projects': AddProjectsSerializer,
        'add_employees': AddEmployeesSerializer,
        'remove_projects': AddProjectsSerializer,
        'remove_employees': AddEmployeesSerializer,
    }

    def get_serializer_class(self):
        return self.serializer_classes.get(self.action, self.serializer_class)

    def _handle_relation_action(self, request, pk, relation_field, serializer_class, model, model_name, filter_conditions=None):
        """
        Generic method to handle add/remove actions for many-to-many relations.
        """
        team = self.get_object()
        ids = request.data.get(f'{model_name}_ids', [])
        
        # Validate input using serializer
        serializer = serializer_class(
            data=request.data,
            context={'request': request},
            model=model,
            model_name=model_name,
            filter_conditions=filter_conditions or {}
        )
        serializer.is_valid(raise_exception=True)

        # Get related objects
        related_objects = model.objects.filter(id__in=ids, **(filter_conditions or {}))
        count_before = getattr(team, relation_field).count()

        # Perform add or remove action
        action = self.action  # 'add_projects', 'remove_projects', etc.
        if 'add' in action:
            getattr(team, relation_field).add(*related_objects)
            count_changed = getattr(team, relation_field).count() - count_before
            action_word = 'Added'
        else:
            getattr(team, relation_field).remove(*related_objects)
            count_changed = count_before - getattr(team, relation_field).count()
            action_word = 'Removed'

        return Response(
            {'message': f'{action_word} {count_changed} {model_name}(s) to/from team "{team.team_name}".'},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], url_path='add-projects')
    def add_projects(self, request, pk=None):
        return self._handle_relation_action(
            request, pk, 'projects', AddProjectsSerializer, Project, 'project', {'active': True}
        )

    @action(detail=True, methods=['delete'], url_path='remove-projects')
    def remove_projects(self, request, pk=None):
        return self._handle_relation_action(
            request, pk, 'projects', AddProjectsSerializer, Project, 'project'
        )

    @action(detail=True, methods=['post'], url_path='add-employees')
    def add_employees(self, request, pk=None):
        return self._handle_relation_action(
            request, pk, 'employees', AddEmployeesSerializer, Employee, 'employee',
            {'active': True, 'project_eligibility': True}
        )

    @action(detail=True, methods=['delete'], url_path='remove-employees')
    def remove_employees(self, request, pk=None):
        return self._handle_relation_action(
            request, pk, 'employees', AddEmployeesSerializer, Employee, 'employee'
        )
    

class DailyUpdatesForWeeklyProjectHoursViewSet(BaseModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectUpdateSerializer
    queryset = Project.objects.all().prefetch_related(
        'employeeproject_set__employee',
        'manager',
        'tpm'
    )
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'project_id',
                openapi.IN_QUERY,
                description="ID of the project to filter updates",
                type=openapi.TYPE_INTEGER,
                required=True
            ),
            openapi.Parameter(
                'selected_date',
                openapi.IN_QUERY,
                description="Date up to which updates are retrieved (YYYY-MM-DD)",
                type=openapi.TYPE_STRING,
                format=openapi.FORMAT_DATE,
                required=True
            )
        ],
        responses={
            200: openapi.Response(
                description="Successful response with project updates grouped by employee",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'manager': openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Name of the project manager"
                        ),
                        'total_approved_hour': openapi.Schema(
                            type=openapi.TYPE_STRING,
                            description="Total approved hours for the project in the specified week"
                        ),
                        'employees': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Items(
                                type=openapi.TYPE_OBJECT,
                                properties={
                                    'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                    'full_name': openapi.Schema(type=openapi.TYPE_STRING),
                                    'total_hour': openapi.Schema(type=openapi.TYPE_STRING),
                                    'updates': openapi.Schema(
                                        type=openapi.TYPE_ARRAY,
                                        items=openapi.Items(
                                            type=openapi.TYPE_OBJECT,
                                            properties={
                                                'hours': openapi.Schema(type=openapi.TYPE_NUMBER),
                                                'update': openapi.Schema(type=openapi.TYPE_STRING),
                                                'status': openapi.Schema(type=openapi.TYPE_STRING),
                                                'created_at': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_DATETIME)
                                            }
                                        )
                                    )
                                }
                            )
                        )
                    }
                )
            ),
            400: openapi.Response(
                description="Bad Request: Missing required fields",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'project_id': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Items(type=openapi.TYPE_STRING),
                            description="Error message for missing project_id"
                        ),
                        'selected_date': openapi.Schema(
                            type=openapi.TYPE_ARRAY,
                            items=openapi.Items(type=openapi.TYPE_STRING),
                            description="Error message for missing selected_date"
                        )
                    }
                ),
                examples={
                    'application/json': {
                        "project_id": ["This field is required."],
                        "selected_date": ["This field is required."]
                    }
                }
            ),
            404: openapi.Response(
                description="Not Found: Project not found or is inactive",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'error': openapi.Schema(type=openapi.TYPE_STRING)
                    }
                )
            )
        },
        operation_description="Retrieve daily project updates for a given project ID and date range (from week start to selected date), grouped by employee with manager name and total approved hours."
    )
    def list(self, request):
        """List daily project updates for a project and date range."""
        serializer = ProjectUpdateFilterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        project_id = serializer.validated_data['project_id']
        selected_date = serializer.validated_data['selected_date']
        
        # Calculate the start of the week (Monday)
        week_start = selected_date - timedelta(days=selected_date.weekday())
        
        try:
            Project.objects.get(id=project_id, active=True)
        except Project.DoesNotExist:
            return Response(
                {"error": "Project not found or is inactive"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Prepare data for serialization
        data = {
            'project_id': project_id,
            'week_start': week_start,
            'selected_date': selected_date
        }
        
        # Serialize and return the data
        serializer = ProjectUpdateSerializer(data, context=data)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class WeeklyProjectHoursViewSet(BaseModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = WeeklyProjectHoursSerializer
    queryset = ProjectHour.objects.all().select_related('project', 'manager', 'tpm')
    parser_classes = [JSONParser, MultiPartParser, FormParser]  # Support both JSON and multipart

    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'date': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_DATE, description="Date of the project hours (YYYY-MM-DD)"),
                'project': openapi.Schema(type=openapi.TYPE_INTEGER, description="ID of the active project"),
                'hour_type': openapi.Schema(type=openapi.TYPE_STRING, enum=['project', 'bonus'], description="Type of project hour", default='project'),
                'hours': openapi.Schema(type=openapi.TYPE_NUMBER, description="Total hours for the week"),
                'report_file': openapi.Schema(type=openapi.TYPE_FILE, description="Weekly update PDF (optional)", nullable=True),
                'employees': openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Items(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                            'full_name': openapi.Schema(type=openapi.TYPE_STRING),
                            'total_hour': openapi.Schema(type=openapi.TYPE_STRING),
                            'updates': openapi.Schema(
                                type=openapi.TYPE_ARRAY,
                                items=openapi.Items(
                                    type=openapi.TYPE_OBJECT,
                                    properties={
                                        'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                                        'hours': openapi.Schema(type=openapi.TYPE_NUMBER),
                                        'update': openapi.Schema(type=openapi.TYPE_STRING),
                                        'status': openapi.Schema(type=openapi.TYPE_STRING),
                                        'created_at': openapi.Schema(type=openapi.TYPE_STRING, format=openapi.FORMAT_DATETIME)
                                    }
                                )
                            )
                        }
                    ),
                    description="List of employee updates from daily project updates"
                )
            },
            required=['date', 'project', 'hours', 'employees']
        ),
        responses={
            201: openapi.Response(
                description="Project hours created successfully",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'status': openapi.Schema(type=openapi.TYPE_STRING, description="Status message"),
                        'code': openapi.Schema(type=openapi.TYPE_INTEGER, description="HTTP status code")
                    }
                )
            ),
            400: openapi.Response(
                description="Bad Request: Invalid input data",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'date': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING)),
                        'project': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING)),
                        'hours': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING)),
                        'employees': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Items(type=openapi.TYPE_STRING)),
                    }
                )
            )
        },
        operation_description="Create a new WeeklyProjectHours entry for a project, including employee updates."
    )
    def create(self, request, *args, **kwargs):
        """Handle POST request to create a WeeklyProjectHours entry."""
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return Response(
                {"status": "successfully created", "code": 201},
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(
        responses={
            200: openapi.Response(
                description="List of project hours for the authenticated user",
                schema=openapi.Schema(
                    type=openapi.TYPE_ARRAY,
                    items=openapi.Items(
                        type=openapi.TYPE_OBJECT,
                        properties={
                            'date': openapi.Schema(type=openapi.TYPE_STRING, description="Date of the project hours"),
                            'project': openapi.Schema(type=openapi.TYPE_STRING, description="Project name"),
                            'hours': openapi.Schema(type=openapi.TYPE_NUMBER, description="Total hours"),
                            'lead': openapi.Schema(type=openapi.TYPE_STRING, description="Manager/lead name"),
                            'resources': openapi.Schema(
                                type=openapi.TYPE_OBJECT,
                                additional_properties=openapi.Schema(type=openapi.TYPE_NUMBER),
                                description="Dictionary of employee names and their total hours"
                            )
                        }
                    )
                )
            )
        },
        operation_description="Retrieve WeeklyProjectHours entries for the authenticated user."
    )
    def list(self, request, *args, **kwargs):
        """Retrieve WeeklyProjectHours entries for the authenticated user."""
        try:
            employee = Employee.objects.get(user=request.user, active=True)
        except Employee.DoesNotExist:
            return Response(
                {"error": "No active Employee instance found for the authenticated user."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.queryset.filter(manager=employee)
        results = []
        for obj in queryset:
            serializer = self.get_serializer(obj, context={'request': request})
            results.append({
                'date': obj.date.strftime('%B %d, %Y'),
                'project': str(obj.project) if obj.project else "No Project Assigned",
                'hours': float(obj.hours),
                'lead': obj.manager.full_name if obj.manager else "No Lead Assigned",
                'resources': serializer.get_resources(obj)
            })
        
        return Response(results, status=status.HTTP_200_OK)