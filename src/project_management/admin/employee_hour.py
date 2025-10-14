import datetime
from datetime import timedelta
import random

import openpyxl
from django.contrib import admin, messages
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import TruncDate
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.loader import get_template
from django.utils import timezone
from django.utils.html import format_html
from icecream import ic
from openpyxl.styles import Alignment, Font, PatternFill

from config.admin import RecentEdit
from config.admin.utils import simple_request_filter
from employee.admin.employee._forms import DailyUpdateFilterForm
from employee.models import Employee
from employee.models.employee_activity import EmployeeProject
# from employee.models.employee_rating_models import EmployeeRating
from project_management.admin.project_hour.options import (
    ProjectLeadFilter,
    ProjectManagerFilter,
)
from project_management.forms import AddDDailyProjectUpdateForm
from project_management.models import (
    DailyProjectUpdate,
    DailyProjectUpdateAttachment,
    DailyProjectUpdateHistory,
    EmployeeProjectHour,
    EnableDailyUpdateNow,
    Project,
    ProjectReport,
)
from project_management.utils.send_report import send_report_slack


class ProjectTypeFilter(admin.SimpleListFilter):
    title = "hour type"
    parameter_name = "project_hour__hour_type"

    def lookups(self, request, model_admin):
        return (
            ("project", "Project"),
            ("bonus", "Bonus"),
        )

    def queryset(self, request, queryset):
        if self.value() == "bonus":
            return queryset.filter(
                project_hour__hour_type="bonus",
            )
        elif self.value() == "project":
            return queryset.filter(
                project_hour__hour_type="project",
            )

        return queryset


@admin.register(EmployeeProjectHour)
class EmployeeHourAdmin(RecentEdit, admin.ModelAdmin):
    list_display = (
        "get_date",
        "employee",
        "hours",
        "get_hour_type",
        "project_hour",
    )
    list_filter = (
        ProjectTypeFilter,
        "employee",
        "created_at",
    )
    search_fields = (
        "hours",
        "employee__full_name",
    )
    date_hierarchy = "project_hour__date"
    autocomplete_fields = ("employee", "project_hour")
    change_list_template = "admin/total.html"

    @admin.display(description="Date", ordering="project_hour__date")
    def get_date(self, obj):
        return obj.project_hour.date

    @admin.display(description="Hour Type", ordering="project_hour__hour_type")
    def get_hour_type(self, obj):
        return obj.project_hour.hour_type.title()

    def manager(self, obj):
        return obj.project_hour.manager

    # query for get total hour by query string
    def get_total_hour(self, request):
        qs = self.get_queryset(request).filter(**simple_request_filter(request))
        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.see_all_employee_hour"
        ):
            if request.user.employee.manager or request.user.employee.lead:
                qs = qs.filter(
                    Q(project_hour__manager=request.user.employee.id)
                    | Q(employee=request.user.employee)
                )
            else:
                qs = qs.filter(employee=request.user.employee)
        return qs.aggregate(tot=Sum("hours"))["tot"]

    # override change list view
    # return total hour count
    def changelist_view(self, request, extra_context=None):
        my_context = {
            "total": self.get_total_hour(request),
        }
        return super(EmployeeHourAdmin, self).changelist_view(
            request, extra_context=my_context
        )

    def get_queryset(self, request):
        """Return query_set

        overrides django admin query set
        allow super admin and permitted user only to see all project hour
        manager's and employees will only see theirs
        @type request: object
        """
        query_set = super(EmployeeHourAdmin, self).get_queryset(request)
        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.see_all_employee_hour"
        ):
            if request.user.employee.manager or request.user.employee.lead:
                return query_set.filter(
                    Q(project_hour__manager=request.user.employee.id)
                    | Q(employee=request.user.employee)
                )
            else:
                return query_set.filter(employee=request.user.employee)
        return query_set

    def get_list_filter(self, request):
        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.see_all_employee_hour"
        ):
            return []
        return super(EmployeeHourAdmin, self).get_list_filter(request)


class DailyProjectUpdateDocumentAdmin(admin.TabularInline):
    model = DailyProjectUpdateAttachment
    extra = 0


@admin.register(DailyProjectUpdate)
class DailyProjectUpdateAdmin(admin.ModelAdmin):
    LAST_TIME_OF_GIVING_UPDATE_FOR_DEVS = datetime.time(19, 30)
    LAST_TIME_OF_GIVING_UPPDATE_FOR_LEADS = datetime.time(23, 59)

    # Duration for cache in seconds
    cache_timeout = 60 * 3  # 15 minutes

    today = timezone.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    deadline = start_of_month + timedelta(days=26)

    inlines = [
        DailyProjectUpdateDocumentAdmin,
    ]
    list_display = (
        "get_date",
        "employee",
        "get_project",
        "get_hours",
        # "history",
        "get_update",
        # "get_updates_json",
        "manager",
        "status_col",
    )
    list_filter = (
        "status",
        "project",
        "employee",
        ProjectManagerFilter,
        ProjectLeadFilter,
    )
    search_fields = (
        "employee__full_name",
        "project__title",
        "manager__full_name",
    )
    date_hierarchy = "created_at"
    autocomplete_fields = (
        "employee",
        "project",
    )
    change_list_template = "admin/total_employee_hour.html"
    readonly_fields = [
        "employee",
        "status",
        "created_at",
        "note",
    ]
    actions = [
        "update_status_approve",
        "update_status_pending",
        "export_updated_in_txt",
        "export_selected",
        "export_selected_merged",
        # "generate_random_updates",  # for developers if needed to add dummy data into daily project updates. below you can find the code
    ]
    form = AddDDailyProjectUpdateForm
    # change_form_template = 'admin/project_management/dailyprojectupdate_form.html'
    fieldsets = (
        (
            "Standard Info",
            {
                "fields": (
                    "created_at",
                    "employee",
                    "manager",
                    "project",
                    "hours",
                    # "update",
                    # "updates_json",
                    "status",
                ),
            },
        ),
        ("Updates", {"fields": ("updates_json", "management_updates")}),
        (
            "Extras",
            {
                "fields": ("note",),
            },
        ),
    )
    list_per_page = 50
    list_select_related = ('employee', 'manager', 'project')

    class Media:
        css = {"all": ("css/list.css", "css/daily-update.css")}
        js = ("js/list.js", "js/new_daily_update.js")

    def get_readonly_fields(self, request, obj=None):

        if request.user.is_superuser or request.user.has_perm(
            "project_management.can_submit_previous_daily_project_update"
        ):
            return []
        if request.user.has_perm(
            "project_management.can_approve_or_edit_daily_update_at_any_time"
        ):
            return [
                "created_at",
            ]

        if not obj:
            return self.readonly_fields

        if obj:
            # If interact  as selected manager for that project
            if obj.manager == request.user.employee:
                # If interacts also as the employee and manager of that project
                if obj.employee == request.user.employee:
                    return [
                        "created_at",
                    ]

                # If not the employee
                return [
                    "created_at",
                    "employee",
                    "manager",
                    "project",
                    "update",
                    # "updates_json",
                ]

            # If interact as the project employee and status approved
            if obj.employee == request.user.employee and obj.status == "approved":
                return self.get_fields(request)

            # If interact as the project employee and status not approved
            return self.readonly_fields

    @admin.display(description="History")
    def history(self, obj):
        history_qs = getattr(obj, "history_list", None)
        if history_qs:
            history_list = list(history_qs)
            hours_str = " > ".join(f"{round(h.hours, 2)}" for h in history_list)
            return format_html(hours_str)
        return "No changes"

    @admin.display(description="Date", ordering="created_at")
    def get_date(self, obj):
        return obj.created_at.date()
    
    @admin.display(description="Project", ordering="project")
    def get_project(self, obj):
        if obj.project.client:
            project = f"{obj.project.title} <br /> {obj.project.client.name}"
        else:
            project = f"{obj.project.title} <br /> (No Client)"
        return format_html(project)

    @admin.display(description="Update")
    def get_update(self, obj):
        # return obj.update
        html_template = get_template(
            "admin/project_management/list/col_dailyupdate.html"
        )

        is_github_link_show = True
        html_content = html_template.render(
            {
                "update": (
                    obj.update.replace("{", "_").replace("}", "_")
                    if obj.updates_json is None
                    else obj.str_updates_json.replace("{", "_").replace("}", "_")
                ),
                "update_json": None if obj.updates_json is None else obj.updates_json,
                "is_github_link_show": is_github_link_show,
            }
        )

        try:
            data = format_html(html_content)
        except:
            data = "-"

        return data

    @admin.display(description="Hours", ordering="hours")
    def get_hours(self, obj):
        history_qs = getattr(obj, "history_list", None)
        hours_str = ""
        if history_qs:
            history_list = list(history_qs)
            hours_str = " > ".join(f"{round(h.hours, 2)}" for h in history_list)
        custom_style = ' style="font-size: 16px"'
        if obj.hours < 6:
            custom_style = ' style="color:red; font-weight: bold; font-size: 16px;"'
        html_content = f"<span{custom_style}>{round(obj.hours, 2)} </span><br /> {hours_str}"
        return format_html(html_content)

    # @method_decorator(cache_page(cache_timeout))
    def changelist_view(self, request, extra_context=None):
        # cache_key = 'dailyprojectupdate_changelist'
        # cached_response = cache.get(cache_key)

        # if cached_response:
        #     return cached_response

        filter_form = DailyUpdateFilterForm(
            initial={
                "created_at__date__gte": request.GET.get(
                    "created_at__date__gte",
                    timezone.now().date() - datetime.timedelta(days=7),
                ),
                "created_at__date__lte": request.GET.get(
                    "created_at__date__lte", timezone.now().date()
                ),
            }
        )
        # is_have_pending = LeaveManagement.objects.filter(
        #     manager=request.user.employee, status="pending"
        # ).exists()

        my_context = {
            "total": self.get_total_hour(request),
            "filter_form": filter_form,
            "is_have_pending": False,  # Pass the variable to the template context
        }

        # Add a message to display in the template if there are pending leave requests
        # if is_have_pending:
        #     messages.info(request, "You have pending leave request(s).")

        if self.today > self.deadline:
            if not self.is_rating_completed(request):
                messages.error(
                    request,
                    "You have to complete your 'Employee Rating' first to add daily project update",
                )

        response = super().changelist_view(request, extra_context=my_context)
        # if isinstance(response, TemplateResponse):
        #     response.render()
        #
        # cache.set(cache_key, response, self.cache_timeout)

        return response

    def get_total_hour(self, request):
        base_qs = self.get_queryset(request)
        filtered_qs = base_qs.filter(**simple_request_filter(request))
        result = filtered_qs.aggregate(total_hours=Sum("hours"))
        return result.get("total_hours") or 0

    def get_queryset(self, request):

        # if condition remove if need to show all
        if (
            request.GET.get("created_at__date__gte", None) is None
            and request.GET.get("q", None) is None
        ):
            one_month_ago = timezone.now() - timedelta(days=30)
            query_set = (
                super(DailyProjectUpdateAdmin, self)
                .get_queryset(request)
                .select_related("employee", "manager", "project")
                .filter(created_at__gte=one_month_ago)
            )
        else:
            query_set = (
                super(DailyProjectUpdateAdmin, self)
                .get_queryset(request)
                .select_related("employee", "manager", "project")
            )

        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.see_all_employee_update"
        ):
            if request.user.employee.manager or request.user.employee.lead:
                query_set = query_set.filter(
                    Q(manager=request.user.employee)
                    | Q(employee=request.user.employee),
                )
            else:
                query_set = query_set.filter(employee=request.user.employee)
        query_set = query_set.prefetch_related(
            Prefetch(
                "history",
                queryset=DailyProjectUpdateHistory.objects.order_by("-created_at"),
                to_attr="history_list",
            )
        )
        return query_set.defer('management_updates')

    def get_list_filter(self, request):
        filters = list(super(DailyProjectUpdateAdmin, self).get_list_filter(request))
        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.see_all_employee_update"
        ):
            if "employee" in filters:
                filters.remove("employee")
        return filters

    def has_change_permission(self, request, obj=None):
        if request.user.has_perm(
            "project_management.can_approve_or_edit_daily_update_at_any_time"
        ):
            return True
        # special_permission = EnableDailyUpdateNow.objects.first()
        # if obj:
        #     if ((request.user.employee.lead or request.user.employee.manager or request.user.employee.sqa) and obj.created_at.date() < timezone.now().date()):
        #         if special_permission is not None and special_permission.enableproject == True:
        #             return True
        #         return False
        #     else:
        #         return True
        #
        # is_have_panding =  LeaveManagement.objects.filter(manager=request.user.employee,status='pending').exists()
        # if is_have_panding:
        #     return False
        #
        # if self.today > self.deadline:
        #     if self.is_rating_completed(request) == False:
        #         return False

        # if request.user.has_perm("project_management.")

        permitted = super().has_change_permission(request, obj=obj)
        if obj is not None and obj.pk:
            if (
                not request.user.is_superuser
                and obj
                and obj.employee != request.user.employee
                and obj.manager != request.user.employee
            ):
                # permitted = False
                return False

        special_permission = EnableDailyUpdateNow.objects.filter(
            enableproject=True
        ).first()

        if obj:
            # is lead / manager / sqa
            if (
                request.user.employee.lead
                or request.user.employee.manager
                or request.user.employee.sqa
            ):
                # previous date
                if obj.created_at.date() < timezone.now().date():
                    if (
                        special_permission is not None
                        and special_permission.enableproject == False
                    ):
                        permitted = False
                # today
                else:
                    permitted = True
            # not lead / manager / sqa
            else:
                if special_permission is not None:
                    if special_permission.enableproject == False:
                        if timezone.now().time() > special_permission.last_time:
                            permitted = False

        # is_have_panding = LeaveManagement.objects.filter(
        #     manager=request.user.employee, status="pending"
        # ).exists()
        # if is_have_panding:
        #     return False

        if self.today > self.deadline:
            if self.is_rating_completed(request) == False:
                return False

        # if not (request.user.employee.lead or request.user.employee.manager or request.user.employee.sqa):
        #     if special_permission is not None:
        #         if special_permission.enableproject == True:
        #             return True
        #         if timezone.now().time() > special_permission.last_time:
        #             return False
        return permitted

    def has_add_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True

        # is_have_panding = LeaveManagement.objects.filter(
        #     manager=request.user.employee, status="pending"
        # ).exists()

        permissons = super().has_add_permission(request)
        special_permission = EnableDailyUpdateNow.objects.first()

        # if is_have_panding:
        #     return False

        if self.today > self.deadline:
            if self.is_rating_completed(request) == False:
                return False

        if not (
            request.user.employee.lead
            or request.user.employee.manager
            or request.user.employee.sqa
        ):
            if special_permission is not None:
                if special_permission.enableproject == True:
                    return True
                if special_permission.last_time > timezone.now().time():
                    return True
            return False
        return permissons

    @admin.display(description="Status")
    def status_col(self, obj):
        color = "red"
        if obj.status == "approved":
            color = "green"
        return format_html(f'<b style="color: {color}">{obj.get_status_display()}</b>')

    @admin.action(description="Approve selected status daily project updates")
    def update_status_approve(modeladmin, request, queryset):
        if request.user.has_perm(
            "project_management.can_approve_or_edit_daily_update_at_any_time"
        ):
            qs_count = queryset.update(status="approved")
        elif request.user.employee.manager or request.user.employee.lead:
            # if len(
            #     LeaveManagement.objects.filter(
            #         manager=request.user.employee, status="pending"
            #     )
            # ):
            #     return messages.error(
            #         request,
            #         "You have pending leave application(s). Please approve first.",
            #     )

            qs_count = queryset.filter(manager_id=request.user.employee.id).update(
                status="approved"
            )

        messages.success(request, f"Marked Approved {qs_count} daily update(s).")

# for developers if needed to add dummy data into daily project updates
    # from django.contrib import admin, messages
    # from django.db import transaction
    # import random
    # import datetime
    # from datetime import timedelta
    # @admin.action(description="Generate 250 random daily project updates")
    # def generate_random_updates(modeladmin, request, queryset=None):
    #     employees = ["Nahar Sara", "Partho Debnath", "Arafat", "Farhina Tarhina"]
    #     projects = [
    #         "UIMC",
    #         "How to fetch all data from database using Django",
    #         "Bulk.ly",
    #         "Anj",
    #         "DWPS",
    #         "p2",
    #         "VIP",
    #     ]
    #     managers = ["Fatin", "Pappu"]
    #     random_updates = [
    #         "Completed feature implementation and bug fixes.",
    #         "Worked on UI enhancements and tested functionality.",
    #         "Integrated new API endpoints and optimized queries.",
    #         "Conducted code review and updated documentation.",
    #         "Fixed bugs and improved performance for backend services.",
    #         "Developed new module and ran unit tests.",
    #     ]

    #     def generate_random_date(start_year=2023, end_year=2024):
    #         start_date = datetime.datetime(start_year, 1, 1)
    #         print(start_date)
    #         end_date = datetime.datetime(end_year, 12, 31)
    #         print(end_date)
    #         delta = end_date - start_date
    #         print(delta)
    #         random_days = random.randint(0, delta.days)
    #         print(random_days)
    #         return start_date + timedelta(days=random_days)

    #     try:
    #         with transaction.atomic():
    #             print("Starting to generate random updates...")

    #             # Fetch or create employees
    #             employee_objects = []
    #             for emp_name in employees:
    #                 print(f"Creating/fetching employee: {emp_name}")
    #                 emp, _ = Employee.objects.get_or_create(
    #                     full_name=emp_name,
    #                     defaults={"active": True},
    #                 )
    #                 employee_objects.append(emp)

    #             # Fetch or create managers
    #             manager_objects = []
    #             for mgr_name in managers:
    #                 print(f"Creating/fetching manager: {mgr_name}")
    #                 mgr, _ = Employee.objects.get_or_create(
    #                     full_name=mgr_name,
    #                     defaults={"active": True, "manager": True},
    #                 )
    #                 manager_objects.append(mgr)

    #             # Fetch or create projects
    #             project_objects = []
    #             for proj_name in projects:
    #                 print(f"Creating/fetching project: {proj_name}")
    #                 proj, _ = Project.objects.get_or_create(
    #                     title=proj_name,
    #                     defaults={"active": True},
    #                 )
    #                 project_objects.append(proj)

    #             # Generate 250 random records
    #             num_records = 250
    #             for i in range(num_records):
    #                 print(f"Generating record {i+1}/{num_records}...")
    #                 employee = random.choice(employee_objects)
    #                 print(f"Selected employee: {employee.full_name}")
    #                 manager = random.choice(manager_objects)
    #                 print(f"Selected manager: {manager.full_name}")
    #                 project = random.choice(project_objects)
    #                 print(f"Selected project: {project.title}")
    #                 hours = round(random.uniform(5.0, 8.0), 1)
    #                 print(f"Generated hours: {hours}")
    #                 update_text = random.choice(random_updates)
    #                 print(f"Generated update: {update_text}")
    #                 updates_json = [[update_text, f"{hours:.1f}H"]]
    #                 created_at = generate_random_date()  # Naive datetime for USE_TZ = False
    #                 print(f"Generated created_at: {created_at}")

    #                 print(f"Creating record {i+1}/{num_records}: {employee.full_name} on {project.title}")

    #                 DailyProjectUpdate.objects.create(
    #                     employee=employee,
    #                     manager=manager,
    #                     project=project,
    #                     hours=hours,
    #                     update=update_text,
    #                     updates_json=updates_json,
    #                     management_updates=f"Additional note for {project.title}",
    #                     status="approved",
    #                     note=f"Approved by {manager.full_name}",
    #                     created_at=created_at,
    #                 )

    #             messages.success(
    #                 request, f"Successfully created {num_records} random daily project updates."
    #             )
    #     except Exception as e:
    #         print(f"Error details: {str(e)}")
    #         messages.error(
    #             request, f"Error occurred while generating updates: {str(e)}"
    #         )

    @admin.action(description="Pending selected status daily project updates")
    def update_status_pending(modeladmin, request, queryset):
        if request.user.has_perm(
            "project_management.can_approve_or_edit_daily_update_at_any_time"
        ):
            qs_count = queryset.update(status="pending")
        elif request.user.employee.manager or request.user.employee.lead:
            qs_count = queryset.filter(manager_id=request.user.employee.id).update(
                status="pending"
            )
            queryset = queryset.filter(manager_id=request.user.employee.id)

        messages.success(request, f"Marked Pending {qs_count} daily update(s).")

    @admin.action(description="Export selected update(s) in .xlsx file")
    def export_selected(modeladmin, request, queryset):
        # ic(queryset)

        project_name = queryset[0].project.title.replace(" ", "_")
        start_date = request.GET.get("created_at__date__gte", "not_selected")
        end_date = request.GET.get("created_at__date__lte", "not_selected")
        date_range = (
            f"{start_date}_to_{end_date}"
            if start_date != "not_selected" and end_date != "not_selected"
            else "selective"
        )
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            f'attachment; filename="{project_name}__{date_range}__exported.xlsx"'
        )

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        # sheet.column_dimensions['A'].width = 13
        sheet.column_dimensions["A"].width = 13
        sheet.column_dimensions["B"].width = 98
        sheet.column_dimensions["C"].width = 13
        sheet.column_dimensions["D"].width = 13
        sheet.column_dimensions["E"].width = 20

        # Customize this section to format and populate the worksheet with your data
        # For example:
        # sheet.append(['  Week  ', '  Date  ', '  Updates  ', '  Task Hours  ', '  Day Hours  ', '  Weekly Hours  '])
        sheet.append(["  Date  ", "  Updates  ", "  Task Hours  ", "  Day Hours  "])

        week_starting_date: str
        week_starting_index: int

        start_ref = 1
        total_hours = 0
        for index, obj in enumerate(queryset, start=2):
            ic("outside if, ", obj.updates_json)
            if (
                obj.updates_json is None
                or obj.updates_json.__len__() == 0
                or obj.status != "approved"
            ):
                ic(obj.updates_json)
                continue
            total_hours += obj.hours
            for index_update, update in enumerate(obj.updates_json):
                sheet.append(
                    [
                        obj.created_at.strftime("%d-%m-%Y"),
                        update[0],
                        update[1],
                        obj.hours,
                    ]
                )

            start_merge = 1 + start_ref
            end_merge = start_merge + index_update
            start_ref = end_merge
            date_cells = f"A{start_merge}:A{end_merge}"
            day_hour_cells = f"D{start_merge}:D{end_merge}"
            sheet.merge_cells(date_cells)
            sheet.merge_cells(day_hour_cells)
            # ic(index, week_cells, date_cells, day_hour_cells)

        sheet.append(["", "", "Total: ", f"{total_hours} Hours"])

        # Make styles
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(cell):
                cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                cell.fill = PatternFill(
                    start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    # text_rotation=0,
                    # wrap_text=False,
                    # shrink_to_fit=True,
                    indent=0,
                )

        for cell in sheet.iter_rows(min_row=2):
            for index, cell in enumerate(cell):
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if index == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="top")

        for cell in sheet.iter_rows(min_row=sheet.max_row):
            ic(cell)
            for index, cell in enumerate(cell):
                ic(index, cell)
                if index == 3 or index == 2:
                    ic("if ", index, cell)
                    cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                    cell.fill = PatternFill(
                        start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", indent=0
                    )

        # Save the workbook to the response
        wb.save(response)

        return response

    @admin.action(description="Export selected update(s) in .xlsx file [Merged Team]")
    def export_selected_merged(modeladmin, request, queryset):
        merged_set = []
        for obj in queryset:
            if merged_set.__len__() > 0:
                ic(
                    merged_set[-1].get("created_at").date() == obj.created_at.date(),
                    merged_set[-1].get("created_at").date(),
                    obj.created_at.date(),
                )
                if merged_set[-1].get("created_at").date() == obj.created_at.date():
                    ic("inside merge")
                    tmp_obj = {
                        "created_at": obj.created_at,
                        "updates_json": merged_set[-1].get("updates_json")
                        + obj.updates_json,
                        "hours": obj.hours + merged_set[-1].get("hours"),
                        "status": obj.status,
                    }
                    merged_set[-1] = tmp_obj
                else:
                    tmp_obj = {
                        "created_at": obj.created_at,
                        "updates_json": obj.updates_json,
                        "hours": obj.hours,
                        "status": obj.status,
                    }
                    merged_set.append(tmp_obj)
            else:
                tmp_obj = {
                    "created_at": obj.created_at,
                    "updates_json": obj.updates_json,
                    "hours": obj.hours,
                    "status": obj.status,
                }
                merged_set.append(tmp_obj)
        ic(merged_set)

        project_name = queryset[0].project.title.replace(" ", "_")
        start_date = request.GET.get("created_at__date__gte", "not_selected")
        end_date = request.GET.get("created_at__date__lte", "not_selected")
        date_range = (
            f"{start_date}_to_{end_date}"
            if start_date != "not_selected" and end_date != "not_selected"
            else "selective"
        )
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = (
            f'attachment; filename="{project_name}__{date_range}__exported.xlsx"'
        )

        # Create a new workbook and add a worksheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        # sheet.column_dimensions['A'].width = 13
        sheet.column_dimensions["A"].width = 13
        sheet.column_dimensions["B"].width = 98
        sheet.column_dimensions["C"].width = 13
        sheet.column_dimensions["D"].width = 13
        sheet.column_dimensions["E"].width = 20

        # Customize this section to format and populate the worksheet with your data
        # For example:
        # sheet.append(['  Week  ', '  Date  ', '  Updates  ', '  Task Hours  ', '  Day Hours  ', '  Weekly Hours  '])
        sheet.append(["  Date  ", "  Updates  ", "  Task Hours  ", "  Day Hours  "])

        week_starting_date: str
        week_starting_index: int

        start_ref = 1
        total_hours = 0
        for index, obj in enumerate(merged_set, start=2):
            # ic('outside if, ', obj.get('updates_json'))
            if (
                obj.get("updates_json") is None
                or obj.get("updates_json").__len__() == 0
                or obj.get("status") != "approved"
            ):
                # ic(obj.get('updates_json'))
                continue
            total_hours += obj.get("hours")
            for index_update, update in enumerate(obj.get("updates_json")):
                sheet.append(
                    [
                        obj.get("created_at").strftime("%d-%m-%Y"),
                        update[0],
                        update[1],
                        obj.get("hours"),
                    ]
                )

            start_merge = 1 + start_ref
            end_merge = start_merge + index_update
            start_ref = end_merge
            date_cells = f"A{start_merge}:A{end_merge}"
            day_hour_cells = f"D{start_merge}:D{end_merge}"
            sheet.merge_cells(date_cells)
            sheet.merge_cells(day_hour_cells)

            # ic(index, week_cells, date_cells, day_hour_cells)

        sheet.append(["", "", "Total: ", f"{total_hours} Hours"])

        # Make styles
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for index, cell in enumerate(cell):
                cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                cell.fill = PatternFill(
                    start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    # text_rotation=0,
                    # wrap_text=False,
                    # shrink_to_fit=True,
                    indent=0,
                )

        for cell in sheet.iter_rows(min_row=2):
            for index, cell in enumerate(cell):
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if index == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="top")

        for cell in sheet.iter_rows(min_row=sheet.max_row):
            # ic(cell)
            for index, cell in enumerate(cell):
                # ic(index, cell)
                if index == 3 or index == 2:
                    # ic('if ', index, cell)
                    cell.font = Font(name="Arial", size=12, bold=True, color="ffffff")
                    cell.fill = PatternFill(
                        start_color="6aa84f", end_color="6aa84f", fill_type="solid"
                    )
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", indent=0
                    )

        # Save the workbook to the response
        wb.save(response)

        return response

    @admin.action(description="Export today's update(s) in .txt file")
    def export_updated_in_txt(modeladmin, request, queryset):
        project_name = queryset[0].project.title.replace(" ", "_")
        date = queryset.first().created_at.strftime("%d-%m-%Y")
        # Annotate the queryset to count the distinct dates
        date_count = (
            queryset.annotate(date=TruncDate("created_at"))
            .values("date")
            .annotate(count=Count("id"))
        )

        if date_count.count() > 1:
            return messages.error(
                request, "Only one day's update can be exported here."
            )
            # return HttpResponseBadRequest("Dates are not the same")
        total_hour = 0

        update_list_str = ""
        for index, obj in enumerate(queryset):
            if obj.updates_json is None:
                continue
            updates = ""
            commit_links = ""
            for index, update in enumerate(obj.updates_json):
                total_hour += float(update[1])
                updates += f"{update[0]} - {update[1]}H.\n"
                commit_links += f"{index+1}. {update[2]}\n"

            tmp_add = (
                f"{obj.employee.full_name}\n\n"
                + f"{updates}\n\n"
                + "Associated Links: \n"
                + f"{commit_links}\n"
                + "-------------------------------------------------------------\n\n"
            )

            update_list_str += tmp_add
        all_updates = (
            "Today's Update\n"
            + "-----------------\n"
            + f"{date}\n\n"
            + f"Total Hours: {round(total_hour, 3)}H\n\n"
        )
        all_updates += update_list_str
        response = HttpResponse(all_updates, content_type="text/plain")
        response["Content-Disposition"] = (
            f'attachment; filename="{project_name}_{date}.txt"'
        )

        return response

    @admin.action(description="Send Report to slack")
    def send_report_to_slack(self, modeladmin, request, queryset, **kwargs):
        date = queryset.first().created_at.strftime("%d-%m-%Y")
        # Annotate the queryset to count the distinct dates
        projects = EmployeeProject.objects.filter(
            employee=request.user.employee, project=queryset[0].project
        )
        if not projects.exists() and not request.user.is_superuser:
            return messages.error(request, "You are not assign this project")

        date_count = (
            queryset.annotate(date=TruncDate("created_at"))
            .values("date")
            .annotate(count=Count("id"))
        )

        if date_count.count() > 1:
            return messages.error(request, "Only one day's update can send to slack.")

        # Annotate the queryset to count the distinct project
        project_count = queryset.values("project").annotate(count=Count("project"))

        if project_count.count() > 1:
            return messages.error(request, "Only one project select to send report.")
        try:
            user_type = ""
            if request.user.is_superuser:
                user_type = "admin"
            elif request.user.employee.manager:
                user_type = "manager"
            elif request.user.employee.lead:
                user_type = "lead"
            elif request.user.employee.top_one_skill.skill.title.lower() == "sqa":
                user_type = "sqa"
            to_report = ProjectReport.objects.get(
                project=queryset[0].project, type=user_type
            )
        except ProjectReport.DoesNotExist:
            return messages.error(
                request, "slack credential not set yet for this project"
            )
        total_hour = 0
        updates_list = []
        links_list = []
        for obj in queryset:
            if obj.updates_json is None or obj.status == "pending":
                continue

            for update in obj.updates_json:
                total_hour += float(update[1])
                updates_list.append(
                    f"{update[0]} ({update[1]}H)"
                )  # Combine update and hours
                links_list.append(update[2]) if len(update) > 2 else ""

        # Prepare the updates section
        updates_section = "\n".join(
            [f"{i}. {update}" for i, update in enumerate(updates_list, start=1)]
        )

        # Prepare the links section
        links_section = "\n".join(
            [f"{i}. {link}" for i, link in enumerate(links_list, start=1)]
        )

        # Combine updates, links, and total hours in the final format
        formatted_message = (
            # f"\nTotal Hours: {round(total_hour, 3)}H\n\n"
            # "Updates:\n" + updates_section + "\n\n"
            #                                  "Links:\n" + links_section
            "Updates:\n" + updates_section + "\n\n"
            "Links:\n" + links_section
        )

        # Add other sections as needed

        # Combine all sections
        all_sections = (
            f"Today's Update\n-----------------\n\nTotal Hours: {round(total_hour, 3)}H\n\n"
            "\n"
            f"{formatted_message}"
        )

        response = send_report_slack(
            token=to_report.api_token, channel=to_report.send_to, message=all_sections
        )

        if response.get("ok"):
            return messages.success(
                request, f"Report send to #{to_report.send_to} channel"
            )
        else:
            return messages.error(
                request, f"#{to_report.send_to} {response.get('error')}"
            )

    def has_delete_permission(self, request, obj=None):
        permitted = super().has_delete_permission(request, obj=obj)
        if (
            not request.user.is_superuser
            and obj
            and obj.employee != request.user.employee
        ):
            permitted = False
        return permitted

    def is_rating_completed(self, request):
        # today = timezone.now()
        # user_id = request.user.id
        # ratings = EmployeeRating.objects.filter(created_by_id=user_id)

        return True

    def _is_rating_completed(self, request):
        today = timezone.now()
        user_id = request.user.id
        # ratings = EmployeeRating.objects.filter(created_by_id=user_id)
        projects = Project.objects.filter(
            employeeproject__employee__user__id=user_id,
        )

        # print("*"*50)
        associated_employees_qs = Employee.objects.none()

        for project in projects:
            associated_employees_qs = (
                associated_employees_qs
                | project.associated_employees.exclude(user__id=user_id)
            )

        # unique_associated_employees = associated_employees_qs.distinct()

        # print(unique_associated_employees)

        is_completed = True
        # for employee in unique_associated_employees:
        #     has_rating = ratings.filter(
        #         Q(created_by_id=user_id)
        #         & Q(employee_id=employee.id)
        #         & Q(month=today.month)
        #         & Q(year=today.year)
        #     ).exists()

        #     # print(f"{employee}--------------------{has_rating}-----------{is_completed}")
        #     if not has_rating:
        #         is_completed = False
        #         # print(is_completed)
        #         break
        # # print("*"*50)

        return is_completed

    def save_model(self, request, obj, form, change) -> None:
        print(f"{self.start_of_month}-----------------{self.deadline}")
        # cache.delete('dailyprojectupdate_changelist')

        if not change:
            if form.cleaned_data.get("employee"):
                employee = form.cleaned_data.get("employee")
            else:
                employee = request.user.employee

            if self.today > self.deadline:
                if self.is_rating_completed(request) == False:
                    messages.error(
                        request,
                        "You have to complete your employee rating first to add daily update",
                    )
                    return

            if len(employee.leave_management_manager.filter(status="pending")) > 0:
                messages.error(
                    request,
                    "You have pending leave application(s). Please approve first.",
                )
                return
            update_obj = DailyProjectUpdate.objects.filter(
                employee=employee,
                project=form.cleaned_data.get("project"),
                manager=form.cleaned_data.get("manager"),
                created_at__date=timezone.now().date(),
            )
            if update_obj.exists():
                messages.error(
                    request, "Already you have given today's update for this project"
                )
                return
        if not obj.employee_id:
            obj.employee_id = request.user.employee.id

        json_updates = form.cleaned_data.get("updates_json")
        if json_updates:
            total_hour = 0
            for index, item in enumerate(json_updates):
                if float(item[1]) > 4:
                    json_updates[index][1] = 4
                    total_hour += 4
                else:
                    total_hour += float(item[1])

            # total_hour = sum(float(item[1]) if float(item[1]) <= 2 else 2 for item in json_updates)

            obj.hours = total_hour
        else:
            total_hour = request.POST.get("hours")
            # ic(total_hour)

        obj.hours = form.cleaned_data.get("hours")
        super().save_model(request, obj, form, change)
        total_hour = form.cleaned_data.get("hours")
        if change == False:
            return DailyProjectUpdateHistory.objects.create(
                # hours=request.POST.get("hours"), daily_update=obj
                hours=total_hour,
                daily_update=obj,
            )

        requested_hours = float(request.POST.get("hours"))
        if requested_hours != obj.hours or obj.created_by is not request.user:
            return DailyProjectUpdateHistory.objects.create(
                # hours=request.POST.get("hours"), daily_update=obj
                hours=total_hour,
                daily_update=obj,
            )

    def get_actions(self, request):
        actions = super().get_actions(request)
        # is_have_pending = LeaveManagement.objects.filter(
        #     manager=request.user.employee, status="pending"
        # ).exists()

        # if is_have_pending:
        #     # Remove both "update_status_approve" and "update_status_pending" actions if there are pending leave approvals
        #     actions.pop("update_status_approve", None)
        #     actions.pop("update_status_pending", None)

        return actions

    def response_add(self, request, obj, post_url_continue=None):
        if obj.pk:
            return super().response_add(
                request, obj, post_url_continue=post_url_continue
            )
        else:
            return redirect("/admin/project_management/dailyprojectupdate/")

    def delete_model(self, request, obj):
        super().delete_model(request, obj)
        # cache.delete('dailyprojectupdate_changelist')
