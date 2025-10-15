import math
from datetime import date, timedelta

import openpyxl  # from networkx import project
from dateutil.relativedelta import relativedelta
from django import forms
from django.contrib import admin
from django.contrib import messages as message
from django.contrib.admin.views.main import ERROR_FLAG
from django.db import models
from django.db.models import (
    Case,
    DateField,
    DecimalField,
    DurationField,
    ExpressionWrapper,
    F,
    Prefetch,
    Sum,
    When,
    Count,
)
from django.db.models.functions import Coalesce, TruncMonth
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect
from django.template.loader import get_template
from django.utils import timezone
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils.timesince import timesince
from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter

from account.models import Income
from config.utils.pdf import PDF
from project_management.models import (
    Client,
    ClientAttachment,
    # ClientExperience,
    ClientFeedbackEmail,
    ClientHistory,
    ClientInvoiceDate,
    ClientReview,
    ClientSource,
    ClientStatus,
    Country,
    CurrencyType,
    DecisionMaker,
    InvoiceType,
    PaymentMethod,
    Project,
)
from website.models_v2.services import ServicePage


class ClientInvoiceDateInline(admin.StackedInline):
    model = ClientInvoiceDate
    extra = 1


class ActiveProjectFilter(admin.SimpleListFilter):
    title = _("Active Project")
    parameter_name = "active_project"

    def lookups(self, request, model_admin):
        project_title = Project.objects.filter(active=True).values_list(
            "title", flat=True
        )
        return tuple((title, title) for title in project_title)

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(project__title=self.value())
        return queryset


@admin.register(ClientReview)
class ClientReviewAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ["review"]

    def has_module_permission(self, request: HttpRequest) -> bool:
        return False


@admin.register(PaymentMethod)
class PaymentMethodAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ["name"]

    def has_module_permission(self, request: HttpRequest) -> bool:
        return False


@admin.register(InvoiceType)
class InvoiceTypeAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ["name"]

    def has_module_permission(self, request: HttpRequest) -> bool:
        return False


@admin.register(Country)
class CountryAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ["name"]

    def has_module_permission(self, request: HttpRequest) -> bool:
        return False


class ClientAttachmentInline(admin.TabularInline):
    model = ClientAttachment
    extra = 1


class ClientForm(forms.ModelForm):
    class Meta:
        model = Client
        fields = (
            "name",
            # "web_name",
            "email",
            "invoice_cc_email",
            "designation",
            "company_name",
            "logo",
            "is_need_feedback",
            "client_feedback",
            "image",
            "linkedin_url",
            # "bill_from",
            # "cc_email",
            "address",
            "country",
            "notes",
            "is_hour_breakdown",
            "hourly_rate",
            "active_from",
            "payment_method",
            "invoice_type",
            "review",
        )

        widgets = {
            "invoice_cc_email": forms.Textarea(
                attrs={
                    "cols": 100,
                    "rows": 5,
                    "placeholder": "client1@gmail.com,client2@gmail.com",
                }
            ),
        }
        
        
    def clean(self):
        cleaned = super().clean()
        instance_pk = getattr(self.instance, "pk", None)
        if instance_pk:                      
            if not cleaned.get("active"):
                if not cleaned.get("remark"):
                    self.add_error("remark", "Remark is required when de-activating a client.")
                if not cleaned.get("client_feedback"):
                    self.add_error("client_feedback", "Client feedback is required when de-activating a client.")
        return cleaned


@admin.register(CurrencyType)
class CurrencyTypeAdmin(admin.ModelAdmin):
    list_display = (
        "currency_name",
        "currency_code",
        "icon",
        "is_active",
        "is_default",
        "exchange_rate",
    )
    list_filter = ("is_active", "is_default")
    search_fields = ("currency_name", "currency_code")
    # readonly_fields = ('exchange_rate',)

    # def get_readonly_fields(self, request, obj=None):
    #     if obj and obj.is_default:
    #         return 'currency_name', 'currency_code', 'icon', 'is_default'
    #     return super().get_readonly_fields(request, obj)

    def has_delete_permission(self, request, obj=None):
        # Prevent deletion of default currency
        if obj and obj.is_default:
            return False
        return super().has_delete_permission(request, obj)


@admin.register(ClientSource)
class ClientSourceAdmin(admin.ModelAdmin):
    pass


@admin.register(DecisionMaker)
class DecisionMakerAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ["name"]

    def has_module_permission(self, request: HttpRequest) -> bool:
        return False


@admin.register(Client)
class ClientAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "get_attachments",
        "get_hourly_rate",
        "get_project_income",
        "get_inactive_from",
        "get_duration",
        "get_referrals_name",
        # "web_name",
        # "get_project_name",
        # "email",
        # 'hourly_rate_display',  # Add this
        # "linkedin_url",
        # "country",
        # "currency",
        "payment_method",
        "invoice_type",
        "get_source",
        "get_remark",
        # "get_client_review",
    )
    # fields = (
    #     "name",
    #     "active",
    #     "source",
    #     "hourly_rate",
    #     # "inactive_from",
    #     "refered_by",
    #     "payment_method",
    #     "invoice_type",
    #     "remark",
    #     "email",
    #     "country",
    #     "review",
    #     "invoice_cc_email",
    #     "designation",
    #     "company_name",
    #     "logo",
    #     # "is_need_feedback",
    #     "client_feedback",
    #     "image",
    #     "linkedin_url",
    #     "bill_from",
    #     # "cc_email",
    #     "address",
    #     # "active_from",
    #     "notes",
    #     "is_hour_breakdown",
    #     "currency",
    # )
    list_filter = [
        # "is_need_feedback",
        "active",
        # "review",
        "payment_method",
        "source",
        "invoice_type",
        "currency",
    ]
    inlines = (ClientAttachmentInline,)
    search_fields = [
        "name",
        "web_name",
        "currency__currency_name",
        "currency__currency_code",
        "currency__icon",
    ]
    autocomplete_fields = ["country", "payment_method", "refered_by"]
    form = ClientForm
    actions = ["mark_as_in_active", "export_to_excel", "export_to_pdf"]

    custom_filters = ["from_date", "to_date"]
    fieldsets = (
        (
            "Identification Information",
            {
                "fields": (
                    "name",
                    "active",
                    "source",
                    "client_value",
                    "refered_by",
                    "email",
                    "country",
                    "company_name",
                    "linkedin_url",
                    "decision_maker",
                ),
            },
        ),
        (
            "Billing Information",
            {
                "fields": (
                    "hourly_rate",
                    "payment_method",
                    "invoice_type",
                    "currency",
                    "bill_from",
                    "invoice_cc_email",
                    "exclude_client_name",
                    "address",
                    "notes",
                ),
            },
        ),
        (
            "Engagement & Relationship",
            {
                "fields": (
                    "remark",  # Issues / Concerns Raised
                    "client_feedback",  # Client Feedback / Testimonial
                ),
            },
        ),
        # (
        #     "Follow-up & Opportunities",
        #     {
        #         "fields": (
        #             "meeting_date",  # Last Engagement Date
        #             "next_follow_up_date",  # Next Follow-up Date
        #             "upsell_opportunity",  # Multi-select checkboxes
        #             "referral_potential",  # Yes/No/Maybe dropdown
        #         ),
        #     },
        # ),
        (
            "Other",
            {
                "fields": (
                    # "active",
                    "logo",
                    "image",
                    "is_need_feedback",
                    "is_hour_breakdown",
                    "active_from",
                    "inactive_from",
                    "follow_up_date",
                    "web_name",
                    "designation",
                    "review",
                ),
                "classes": ("collapse",),  # optional – collapsible section
            },
        ),
        # …additional fieldsets go here…
    )

    # Restrict actions based on permissions
    def get_actions(self, request):
        actions = super().get_actions(request)
        # Remove actions if the user is not a superuser and lacks the specific permissions
        if not request.user.is_superuser:
            if not request.user.has_perm(
                "project_management.can_mark_as_inactive"
            ):
                actions.pop("mark_as_in_active", None)
            if not request.user.has_perm(
                "project_management.can_export_to_excel"
            ):
                actions.pop("export_to_excel", None)

            if not request.user.has_perm(
                "project_management.can_export_to_pdf"
            ):
                actions.pop("export_to_pdf", None)
        return actions
    
    @admin.display(description="Attachments", ordering="attachment_count")
    def get_attachments(self, obj):
        attachments = obj.clientattachment_set.all()
        if not attachments:
            return "—"
        
        icons = []
        for attachment in attachments:
            if attachment.attachment:
                url = attachment.attachment.url
                icons.append(f'<a href="{url}" target="_blank" title="{attachment.name or "Attachment"}">📄</a>')
        
        return mark_safe(" ".join(icons))

    @admin.action(description="Export selected clients to PDF")
    def export_to_pdf(self, request, queryset):
        if not request.user.is_superuser and not request.user.has_perm(
            "project_management.can_export_to_pdf"
        ):
            self.message_user(
                request,
                "You do not have permission to export clients to PDF.",
                # messages.ERROR,
            )
            return
        queryset = queryset.prefetch_related(
            Prefetch(
                "project_set",
                queryset=Project.objects.all(),
                to_attr="projects",
            ),
            Prefetch(
                "upsell_opportunity",
                queryset=ServicePage.objects.all(),
                to_attr="upsell_services",
            ),
        ).select_related(
            "currency",
            "invoice_type",
            "payment_method",
            "country",
            "source",
        )
        projects = getattr(queryset.first(), "projects", None)
        project_name = [project.title for project in projects]
        services = set()
        for project in projects:
            if project.services.exists():
                s = {service.title for service in project.services.all()}
                services.union(s)

        upsell_services = getattr(queryset.first(), "upsell_services", None)
        upsell_list = (
            [service.title for service in upsell_services]
            if upsell_services
            else []
        )
        pdf = PDF()
        pdf.template_path = "admin/client_pdf.html"
        pdf.context = {
            "client": queryset.first(),
            "project_names": ", ".join(project_name),
            "services": ", ".join(services),
            "today": date.today(),
            "upsell_services": "\n".join(upsell_list),
        }
        pdf.file_name = f"{queryset.first().name}_client_info"

        return pdf.render_to_pdf(download=True)

    # Dynamically exclude hourly_rate field in detail view for users with exclude_hourly_rate permission
    def get_fields(self, request, obj=None):
        fields = list(super().get_fields(request, obj))
        if not request.user.is_superuser and request.user.has_perm(
            "project_management.exclude_hourly_rate"
        ):
            if "hourly_rate" in fields:
                fields.remove("hourly_rate")
            if not request.user.is_superuser and not request.user.has_perm(
                "project_management.can_see_all_field"
            ):
                fields.remove("active")
        return tuple(fields)

    def changelist_view(self, request, extra_context=None):
        request.GET._mutable = True
        from_date = request.GET.pop("from_date", [""])[0]
        to_date = request.GET.pop("to_date", [""])[0]
        request.GET._mutable = False

        if ERROR_FLAG in request.GET:
            return HttpResponseRedirect(request.path)

        extra_context = extra_context or {}
        extra_context.update({"from_date": from_date, "to_date": to_date})
        return super().changelist_view(request, extra_context=extra_context)

    def lookup_allowed(self, lookup, value):
        if lookup in self.custom_filters:
            return True
        return super().lookup_allowed(lookup, value)

    @admin.display(description="Project Name")
    def get_project_name(self, obj):
        project_name = obj.project_set.all().values_list("title", flat=True)
        return format_html("<br>".join(project_name))

    @admin.display(description="Inactive", ordering="inactive_from")
    def get_inactive_from(self, client_obj):
        return client_obj.inactive_from

    @admin.display(description="source", ordering="source")
    def get_source(self, client_obj):
        return client_obj.source

    @admin.display(description="Referrals", ordering="refered_by")
    def get_referrals_name(self, obj):
        client_referrals = Client.objects.filter(refered_by=obj).values_list(
            "name", flat=True
        )
        return format_html("<br>".join(client_referrals))


    # Inside your ClientAdmin class

    from django.utils.html import format_html
    from django.utils import timezone
    from datetime import timedelta

   
    @admin.display(description="Hourly Rate", ordering="hourly_rate")
    def get_hourly_rate(self, obj):
        # Use prefetched history if available
        histories = getattr(obj, 'prefetched_hourly_rate_history', None)
        if histories is None:
            histories = obj.hourly_rate_history.all().order_by('-starting_date')

        rate_lines = []

        # Handle legacy case: no history but hourly_rate is set
        if not histories and obj.hourly_rate is not None:
            currency_icon = obj.currency.icon if obj.currency else ""
            rate_value = obj.hourly_rate
            date_str = obj.active_from.strftime("%d %b %y") if obj.active_from else "—"
            display_text = f"{rate_value} ({date_str})" if currency_icon else f"{rate_value} ({date_str})"
            # display_text = f"{rate_value} ({date_str})"

            if obj.active_from:
                six_months_ago = timezone.now().date() - timedelta(days=180)
                if obj.active_from <= six_months_ago:
                    display_text = format_html(
                        '<span style="color: red; font-weight: bold;">{}</span>',
                        display_text
                    )
                else:
                    display_text = format_html('<span>{}</span>', display_text)
            else:
                display_text = format_html('<span>{}</span>', display_text)

            rate_lines.append(display_text)


        # Process historical records
        for idx, history in enumerate(histories):
            currency_icon = obj.currency.icon if obj.currency else ""
            rate_value = history.hourly_rate
            date_str = history.starting_date.strftime("%d %b %y") if history.starting_date else "—"
            display_text = f"{rate_value} ({date_str})" if currency_icon else f"{rate_value} ({date_str})"

            if idx == 0:  # Current rate
                if history.starting_date:
                    six_months_ago = timezone.now().date() - timedelta(days=180)
                    if history.starting_date <= six_months_ago:
                        display_text = format_html(
                            '<span style="color: red; font-weight: bold;">{}</span>',
                            display_text
                        )
                    else:
                        display_text = format_html('<span>{}</span>', display_text)
                else:
                    display_text = format_html('<span>{}</span>', display_text)

            rate_lines.append(display_text)

        # Final rate_display string for template — join with <br>
        final_rate_display = format_html("<br>".join(rate_lines)) if rate_lines else "-"

        # Render your original template — now with enhanced rate_display
        html_template = get_template(
            "admin/project_management/list/client_info.html"
        )
        html_content = html_template.render(
            {
                "projects": obj.project_set.all(),
                "client_obj": obj,
                "rate_display": final_rate_display,
                "client_review": obj.review.all().values_list("name", flat=True),
                "client_hourly_rate": obj.hourly_rate,  
            }
        )

        try:
            return format_html(html_content)
        except Exception:
            return "-"



    @admin.display(description="Age", ordering="created_at")
    def get_client_age(self, obj):
        return timesince(obj.created_at)

    @admin.display(description="Remark", ordering="remark")
    def get_remark(self, obj):
        if not obj.remark:
            return "-"
        html_template = get_template("admin/client_remark.html")

        html_content = html_template.render(
            {
                "remark": obj.remark,
            }
        )

        return mark_safe(html_content)

    def get_list_filter(self, request: HttpRequest):
        if request.user.has_perm("project_management.view_client"):
            return super().get_list_filter(request)
        self.list_filter.remove("payment_method")
        return self.list_filter

    @admin.action(description="Mark as In-active")
    def mark_as_in_active(self, request, queryset):
        if not request.user.is_superuser and not request.user.has_perm(
            "your_app.can_mark_as_inactive"
        ):
            self.message_user(
                request,
                "You do not have permission to mark clients as inactive.",
                # messages.ERROR,
            )
            return
        try:
            queryset.update(active=False)
            self.message_user(
                request,
                "Selected clients are marked as active.",
                message.SUCCESS,
            )
        except Exception:
            self.message_user(
                request,
                "Selected clients are not marked as active.",
                message.ERROR,
            )

    @admin.display(description="Income", ordering="total_income")
    def get_project_income(self, client_object):
        """
        Display the total income (bold) followed by the income for the last 4 months (with month names),
        each on a new line. If a month's income drops by more than 500 compared to the previous month
        (or March for April), display it in red. Use ceiling rounding for all values.
        Format: $ total_income\nJuly income\nJune income\nMay income\nApril income
        Example: $ 1000\nJuly 500\nJune 400\nMay 1000\nApril 700
        """
        # Get the current date and define the start of the last 4 months and 5th month
        current_date = timezone.now().date()
        four_months_ago = current_date - relativedelta(months=4)
        four_months_ago = four_months_ago.replace(
            day=1
        )  # Start from the 1st of the month
        fifth_month_start = (current_date - relativedelta(months=4)).replace(
            day=1
        )
        fifth_month_end = (
            fifth_month_start + relativedelta(months=1) - relativedelta(days=1)
        )

        # Define the month keys and names for the last 4 months
        month_0 = current_date.replace(day=1)  # Current month (July 2025)
        month_1 = (current_date - relativedelta(months=1)).replace(
            day=1
        )  # June 2025
        month_2 = (current_date - relativedelta(months=2)).replace(
            day=1
        )  # May 2025
        month_3 = (current_date - relativedelta(months=3)).replace(
            day=1
        )  # April 2025
        month_names = [
            month_0.strftime("%B"),
            month_1.strftime("%B"),
            month_2.strftime("%B"),
            month_3.strftime("%B"),
        ]

        # Get all projects associated with the client
        client_project_ids = client_object.project_set.all().values_list(
            "id", flat=True
        )

        # Get total income from the annotated queryset and apply ceiling
        total_income = math.ceil(
            float(getattr(client_object, "total_income", 0.0))
        )

        # Aggregate monthly incomes in a single query for the last 4 months
        monthly_incomes_qs = (
            Income.objects.filter(
                project_id__in=client_project_ids,
                status="approved",
                date__gte=four_months_ago,
                date__lte=current_date,
            )
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(
                total=Coalesce(
                    Sum(
                        ExpressionWrapper(
                            F("hours") * F("hour_rate"),
                            output_field=DecimalField(
                                max_digits=10, decimal_places=2
                            ),
                        )
                    ),
                    0.0,
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
        )

        # Aggregate pending payments for the last 4 months
        pending_incomes_qs = (
            Income.objects.filter(
                project_id__in=client_project_ids,
                status="pending",
                date__gte=four_months_ago,
                date__lte=current_date,
            )
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(
                total=Coalesce(
                    Sum(
                        ExpressionWrapper(
                            F("hours") * F("hour_rate"),
                            output_field=DecimalField(
                                max_digits=10, decimal_places=2
                            ),
                        )
                    ),
                    0.0,
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                )
            )
        )

        # Aggregate income for the 5th month (March 2025)
        fifth_month_income_qs = Income.objects.filter(
            project_id__in=client_project_ids,
            status="approved",
            date__gte=fifth_month_start,
            date__lte=fifth_month_end,
        ).aggregate(
            total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("hours") * F("hour_rate"),
                        output_field=DecimalField(
                            max_digits=10, decimal_places=2
                        ),
                    )
                ),
                0.0,
                output_field=DecimalField(max_digits=10, decimal_places=2),
            )
        )
        fifth_month_income = float(fifth_month_income_qs["total"])

        # Create a dictionary of monthly incomes
        monthly_incomes = {
            item["month"]: float(item["total"]) for item in monthly_incomes_qs
        }
        pending_incomes = {
            item["month"]: float(item["total"]) for item in pending_incomes_qs
        }

        # Get incomes for the last 4 months using dictionary lookups and apply ceiling
        incomes = [
            math.ceil(monthly_incomes.get(month_0, 0.0)),
            math.ceil(monthly_incomes.get(month_1, 0.0)),
            math.ceil(monthly_incomes.get(month_2, 0.0)),
            math.ceil(monthly_incomes.get(month_3, 0.0)),
        ]

        pending = [
            math.ceil(pending_incomes.get(month_0, 0.0)),
            math.ceil(pending_incomes.get(month_1, 0.0)),
            math.ceil(pending_incomes.get(month_2, 0.0)),
            math.ceil(pending_incomes.get(month_3, 0.0)),
        ]

        # Determine which months have a drop greater than 500 compared to the previous month
        red_flags = [
            incomes[0] < incomes[1] - 500
            if incomes[1] > 0
            else False,  # July vs June
            incomes[1] < incomes[2] - 500
            if incomes[2] > 0
            else False,  # June vs May
            incomes[2] < incomes[3] - 500
            if incomes[3] > 0
            else False,  # May vs April
            incomes[3] < fifth_month_income - 500
            if fifth_month_income > 0
            else False,  # April vs March
        ]

        month_lines = [
            f"{incomes[i]}{f' ({pending[i]})' if pending[i] > 0 else ''}"
            for i in range(4)
        ]

        # # Format the output with HTML: bold total_income, month names with incomes, red for significant drops
        # formatted_income = format_html(
        #     '<span style="font-weight: bold; font-size: calc(1rem);">$ {}</span><br>'
        #     '<span style="color: {}">{}</span><br>'
        #     '<span style="color: {}">{}</span><br>'
        #     '<span style="color: {}">{}</span><br>'
        #     '<span style="color: {}">{}</span>',
        #     total_income,
        #     "red" if red_flags[0] else "inherit", incomes[0],
        #     "red" if red_flags[1] else "inherit", incomes[1],
        #     "red" if red_flags[2] else "inherit", incomes[2],
        #     "red" if red_flags[3] else "inherit", incomes[3],
        # )
        formatted_income = format_html(
            '<span style="font-weight: bold; font-size: calc(1rem);">{}</span><br>'
            '<span style="color: {}">{}</span><br>'
            '<span style="color: {}">{}</span><br>'
            '<span style="color: {}">{}</span><br>'
            '<span style="color: {}">{}</span>',
            total_income,
            "red" if red_flags[0] else "inherit",
            month_lines[0],
            "red" if red_flags[1] else "inherit",
            month_lines[1],
            "red" if red_flags[2] else "inherit",
            month_lines[2],
            "red" if red_flags[3] else "inherit",
            month_lines[3],
        )
        return formatted_income

    @admin.display(description="Duration", ordering="duration_in_days")
    def get_duration(self, client_object):
        """
        get active duration, Example: "1y 5m 10d"
        """
        current_date = timezone.now().date()

        active_from = client_object.active_from or current_date
        inactive_from = client_object.inactive_from or current_date

        total_days = (inactive_from - active_from).days
        years, remainder = divmod(total_days, 365)
        months, days = divmod(remainder, 30)

        time_list = []
        if years:
            time_list.append(f"{years}y")
        if months:
            time_list.append(f"{months}m")
        if days:
            time_list.append(f"{days}d")
        return " ".join(time_list) or None

    # def has_module_permission(self, request):
    #     return False



    # def get_queryset(self, request):
    #     queryset = super().get_queryset(request)
    #     current_date = timezone.now().date()
    #     income_filter = models.Q(project__income__status="approved")
    #     active_from = F("active_from")
    #     inactive_from = F("inactive_from")

    #     queryset = queryset.annotate(
    #         duration_in_days=ExpressionWrapper(
    #             expression=Coalesce(inactive_from, current_date)
    #             - Coalesce(active_from, current_date),
    #             output_field=DurationField(),
    #         ),
    #         total_income=Coalesce(
    #             Sum(
    #                 ExpressionWrapper(
    #                     F("project__income__hours")
    #                     * F("project__income__hour_rate"),
    #                     output_field=DecimalField(
    #                         max_digits=10, decimal_places=2
    #                     ),
    #                 ),
    #                 filter=income_filter,
    #                 output_field=DecimalField(max_digits=10, decimal_places=2),
    #             ),
    #             0.0,
    #             output_field=DecimalField(max_digits=10, decimal_places=2),
    #         ),
    #     ).prefetch_related(
    #         Prefetch(
    #             "hourly_rate_history",
    #             queryset=ClientHistory.objects.all().order_by('-starting_date'),
    #             to_attr="prefetched_hourly_rate_history"
    #         )
    #     )

    #     return queryset

    from django.db.models import Count

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        current_date = timezone.now().date()
        income_filter = models.Q(project__income__status="approved")
        active_from = F("active_from")
        inactive_from = F("inactive_from")

        queryset = queryset.annotate(
            duration_in_days=ExpressionWrapper(
                expression=Coalesce(inactive_from, current_date)
                - Coalesce(active_from, current_date),
                output_field=DurationField(),
            ),
            total_income=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F("project__income__hours")
                        * F("project__income__hour_rate"),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    ),
                    filter=income_filter,
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
                0.0,
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
            attachment_count=Count('clientattachment'),  # ← Add this
        ).prefetch_related(
            Prefetch(
                "hourly_rate_history",
                queryset=ClientHistory.objects.all().order_by('-starting_date'),
                to_attr="prefetched_hourly_rate_history"
            )
        )

        return queryset


    def get_list_display(self, request):
        list_display = list(super().get_list_display(request))

        if request.user.is_superuser:
            return tuple(list_display)
        if request.user.has_perm("project_management.exclude_income"):
            if "get_project_income" in list_display:
                list_display.remove("get_project_income")
        if request.user.has_perm("project_management.exclude_hourly_rate"):
            if "get_hourly_rate" in list_display:
                list_display.remove("get_hourly_rate")
        return tuple(list_display)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        if not obj.active and obj.inactive_from is None:
            obj.inactive_from = timezone.now().date()
            obj.save()
            obj.project_set.update(active=False)
        elif obj.active:
            if obj.active_from is None:
                obj.active_from = timezone.now().date()
            obj.inactive_from = None
            obj.save()

    @admin.action(description="Export selected clients to Excel")
    def export_to_excel(self, request, queryset):
        if not request.user.is_superuser and not request.user.has_perm(
            "your_app.can_export_to_excel"
        ):
            self.message_user(
                request,
                "You do not have permission to export clients to Excel.",
                # messages.ERROR,
            )
            return
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"

        # Define headers based on list_display
        headers = [
            "Name",
            "Projects",
            "Email",
            "Country",
            "Hourly Rate",
            "Income",
            "Inactive From",
            "Duration",
            "Referrals",
            "Payment Method",
            "Invoice Type",
            "Source",
            # "Remark",
        ]
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header
            ws[f"{get_column_letter(col_num)}1"].font = openpyxl.styles.Font(
                bold=True
            )

        for row_num, client in enumerate(queryset, 2):
            ws[f"{get_column_letter(1)}{row_num}"] = client.name
            ws[f"{get_column_letter(2)}{row_num}"] = (
                ", ".join(
                    client.project_set.all().values_list("title", flat=True)
                )
                if client.project_set.exists()
                else ""
            )
            ws[f"{get_column_letter(3)}{row_num}"] = client.email or ""
            ws[f"{get_column_letter(4)}{row_num}"] = (
                str(client.country) if client.country else ""
            )
            ws[f"{get_column_letter(5)}{row_num}"] = (
                f"{client.currency.icon if client.currency else ''} {client.hourly_rate}"
                if client.hourly_rate is not None
                else "-"
            )
            ws[f"{get_column_letter(6)}{row_num}"] = self.get_project_income(
                client
            )
            ws[f"{get_column_letter(7)}{row_num}"] = (
                client.inactive_from.strftime("%Y-%m-%d")
                if client.inactive_from
                else ""
            )
            ws[f"{get_column_letter(8)}{row_num}"] = (
                self.get_duration(client) or ""
            )
            ws[f"{get_column_letter(9)}{row_num}"] = (
                self.get_referrals_name(client).replace("<br>", ", ")
                if "<br>" in self.get_referrals_name(client)
                else self.get_referrals_name(client)
            )
            ws[f"{get_column_letter(10)}{row_num}"] = (
                str(client.payment_method) if client.payment_method else ""
            )
            ws[f"{get_column_letter(11)}{row_num}"] = (
                str(client.invoice_type) if client.invoice_type else ""
            )
            ws[f"{get_column_letter(12)}{row_num}"] = client.source or ""

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="clients_export_{date.today().strftime("%Y%m%d")}.xlsx"'
        )
        wb.save(response)
        return response

    class Media:
        css = {"all": ("css/list.css", "css/daily-update.css")}



@admin.register(ClientHistory)
class ClientHistoryAdmin(admin.ModelAdmin):
    list_display = (
        "client",
        "starting_date",
        "hourly_rate",
        "end_date",
        "created_at",
    )

    # def has_module_permission(self, request):
    #     return False





@admin.register(ClientFeedbackEmail)
class ClientFeedbackEmailAdmin(admin.ModelAdmin):
    list_display = ("subject", "feedback_type")

    def has_module_permission(self, request):
        return False


# class ClientExperienceForm(forms.ModelForm):
#     class Meta:
#         model = ClientExperience
#         fields = "__all__"

#     def clean(self):
#         status = self.cleaned_data.get("status")
#         follow_up_date = self.cleaned_data.get("follow_up_date")
#         meeting_date = self.cleaned_data.get("meeting_date")
#         if status == 1 and not follow_up_date:
#             raise forms.ValidationError(
#                 "Follow up date is required. When status is follow up."
#             )
#         elif status == 2 and not meeting_date:
#             raise forms.ValidationError(
#                 "Meeting date is required. When status is meeting."
#             )
#         super().clean()


# @admin.register(ClientExperience)
# class ClientExperienceAdmin(admin.ModelAdmin):
#     list_display = (
#         "name",
#         # "web_name",
#         "get_project_name",
#         "email",
#         # "linkedin_url",
#         # "get_client_review",
#         "country",
#         "get_hourly_rate",
#         "payment_method",
#         "invoice_type",
#         "get_client_age",
#         "status",
#     )
#     fields = (
#         # "name",
#         # "active",
#         # "email",
#         # "invoice_cc_email",
#         # "designation",
#         # "company_name",
#         # "logo",
#         # "is_need_feedback",
#         # "client_feedback",
#         # "image",
#         # "linkedin_url",
#         # "bill_from",
#         # # "cc_email",
#         # "address",
#         # "country",
#         # "notes",
#         # "is_hour_breakdown",
#         # "hourly_rate",
#         # "active_from",
#         # "payment_method",
#         # "invoice_type",
#         # "review",
#         "status",
#         "follow_up_date",
#         "meeting_date",
#     )
#     readonly_fields = (
#         "name",
#         "active",
#         "email",
#         "invoice_cc_email",
#         "designation",
#         "company_name",
#         "logo",
#         "is_need_feedback",
#         "client_feedback",
#         "image",
#         "linkedin_url",
#         "bill_from",
#         # "cc_email",
#         "address",
#         "country",
#         "notes",
#         "is_hour_breakdown",
#         "hourly_rate",
#         "active_from",
#         "payment_method",
#         "invoice_type",
#         "review",
#     )
#     list_filter = [
#         "is_need_feedback",
#         "active",
#         "status",
#         "review",
#         # "payment_method",
#         # "invoice_type",
#     ]
#     # inlines = (ClientInvoiceDateInline, ClientAttachmentInline)
#     search_fields = ["name", "web_name"]
#     # autocomplete_fields = ["country", "payment_method"]
#     form = ClientExperienceForm
#     # actions = ["mark_as_in_active"]

#     def get_queryset(self, request):
#         return (
#             super()
#             .get_queryset(request)
#             .prefetch_related("project_set")
#             .annotate(
#                 sort_date=Case(
#                     When(status=ClientStatus.FOLLOW_UP, then="follow_up_date"),
#                     When(status=ClientStatus.MEETING, then="meeting_date"),
#                     default=None,
#                     output_field=DateField(),
#                 )
#             )
#             .order_by("sort_date")
#         )

#     @admin.display(description="Project Name")
#     def get_project_name(self, obj):
#         project_name = obj.project_set.all().values_list("title", flat=True)

#         return format_html("<br>".join(project_name))

#     @admin.display(description="Hourly", ordering="hourly_rate")
#     def get_hourly_rate(self, obj):
#         if obj.is_active_over_six_months:
#             return format_html(
#                 f"<span style='color: red;'>{obj.hourly_rate}</span>"
#             )
#         return obj.hourly_rate

#     @admin.display(description="Age", ordering="created_at")
#     def get_client_age(self, obj):
#         return timesince(obj.created_at)

#     def get_list_filter(self, request: HttpRequest):
#         if request.user.has_perm("project_management.view_client"):
#             return super().get_list_filter(request)
#         self.list_filter.remove("payment_method")
#         return self.list_filter

#     @admin.action(description="Mark as In-active")
#     def mark_as_in_active(self, request, queryset):
#         try:
#             queryset.update(active=False)
#             self.message_user(
#                 request,
#                 "Selected clients are marked as active.",
#                 message.SUCCESS,
#             )
#         except Exception:
#             self.message_user(
#                 request,
#                 "Selected clients are not marked as active.",
#                 message.ERROR,
#             )

#     # get_project_name.short_description = "Project Name"
#     @admin.display(description="Client Review")
#     def get_client_review(self, obj):
#         client_review = obj.review.all().values_list("name", flat=True)

#         return format_html("<br>".join(client_review))

#     # def has_module_permission(self, request):
#     #     return False

#     def save_model(self, request, obj, form, change):
#         if obj.status == ClientStatus.MEETING:
#             obj.follow_up_date = None
#         elif obj.status == ClientStatus.FOLLOW_UP:
#             obj.meeting_date = None
#         super().save_model(request, obj, form, change)
#         super().save_model(request, obj, form, change)
