from math import floor

from django.conf import settings
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from account.models import SalaryDisbursement
from config.utils.pdf import PDF
# from employee.models.bank_account import BEFTN


class FestivalBonusAction(admin.ModelAdmin):
    actions = (
        # "export_city_bank_beftn",
        "export_bankasia_salary_acc_dis_excel",
        "export_salary_account_dis_excel",
        "export_bonus_account_dis_pdf",
        "export_city_live_excel",
    )
    
    
    def get_actions(self, request):
        actions = super().get_actions(request)
            
        if request.user.is_superuser or request.user.has_perm("account.can_see_salary_on_salary_sheet"):
            return actions
        return tuple()
    
    @admin.action(description="Export City Live Excel")
    def export_city_live_excel(self, request, queryset):
        wb = Workbook()
        work_sheet = wb.active
        work_sheet.title = "City Live Export"
        work_sheet.append(
            [
                "Reason",
                "Sender Account No",
                "Receiving Bank Routing No",
                "Beneficiary Bank Account  No",
                "Account Type",
                "Amount",
                "Receiver ID",	
                "Receiver Name",
                "Remarks",
                "Receiver Mobile Number",
                "Receiver Email Address",
            ]
        )

        for bonus_sheet in queryset:
            print(dir(bonus_sheet))
            for employee_bonus in bonus_sheet.employeefestivalbonus_set.all():

                bank_account = employee_bonus.employee.bankaccount_set.filter(
                    default=True, is_approved=True
                ).last()
                if bank_account:
                    work_sheet.append(
                        [
                            "",
                            "",
                            "",
                            bank_account.account_number,
                            "",
                            str(int(employee_bonus.amount)),
                            "",
                            employee_bonus.employee.full_name,
                            "",
                            "",
                            "",
                        ]
                    )

        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=City_Live.xlsx"
        return response

    # @admin.action(description="Export City Bank BEFTN")
    # def export_city_bank_beftn(self, request, queryset):
    #     wb = Workbook()
    #     work_sheet = wb.active
    #     work_sheet.title = "BEFTN Export"

    #     # Add headers including 'Total Salary' at the end
    #     work_sheet.append(
    #         [
    #             "Date",
    #             "Account No",
    #             "Employee Name",
    #             "BDT",
    #             "Amount",
    #             "Originating Bank Routing No.",
    #             "Routing No",
    #             "Originating Bank Account No.",
    #             "Originating Account Name",
    #             "Remarks",
    #         ]
    #     )

    #     for festival_bonus_sheet in queryset:
    #         # beftn = BEFTN.objects.last()

    #         total_salary = 0
    #         for (
    #             employee_festival_bonus
    #         ) in festival_bonus_sheet.employeefestivalbonus_set.all():
    #             # Get the bank account information
    #             bank_account = employee_festival_bonus.employee.bankaccount_set.filter(
    #                 default=True, is_approved=True
    #             ).last()

    #             if bank_account and employee_festival_bonus.amount > 0:
    #                 # Append the employee salary data with total salary at the end
    #                 total_salary += int(employee_festival_bonus.amount)
    #                 work_sheet.append(
    #                     [
    #                         festival_bonus_sheet.date.strftime("%d-%m-%Y"),  # Date
    #                         bank_account.account_number,  # Account No
    #                         employee_festival_bonus.employee.full_name,
    #                         "BDT",  # Currency
    #                         str(int(employee_festival_bonus.amount)),  # Amount
    #                         beftn.originating_bank_routing_number,  # Originating Bank Routing No.
    #                         beftn.routing_no,  # Routing No
    #                         beftn.originating_bank_account_number,  # Originating Bank Account No.
    #                         beftn.originating_bank_account_name,  # Originating Account Name
    #                         f'Festival Bonus of {festival_bonus_sheet.date.strftime("%b, %Y")}',  # Remarks
    #                     ]
    #                 )
    #     work_sheet.append(
    #         [
    #             "",
    #             "",
    #             "",
    #             "",
    #             f"Total={int(total_salary)}",
    #             "",
    #             "",
    #             "",
    #             "",
    #             "",  # Total salary row below "Amount"
    #         ]
    #     )
    #     # Prepare the response with the Excel file
    #     response = HttpResponse(
    #         content=save_virtual_workbook(wb), content_type="application/ms-excel"
    #     )
    #     response["Content-Disposition"] = "attachment; filename=City_Bank_BEFTN.xlsx"
    #     return response

    @admin.action(description="Export Bank Asia Salary Account Disbursements (Excel)")
    def export_bankasia_salary_acc_dis_excel(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(
            disbursement_type="salary_account"
        ).first()
        return self.export_in_xl_bankasia(
            queryset,
            ("employee__in", salary_disbursement.employee.all()),
        )

    @admin.action(description="Export Salary Account Disbursements (Excel)")
    def export_salary_account_dis_excel(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(
            disbursement_type="salary_account"
        ).first()
        return self.export_in_xl_dbbl(
            queryset,
            ("employee__in", salary_disbursement.employee.all()),
        )

    @admin.action(description="Export Bonus Account Disbursements (PDF)")
    def export_bonus_account_dis_pdf(self, request, queryset):
        return self.bonus_pdf(
            queryset=queryset,
            filter=("disbursement_type", "salary_account"),
            bank={
                "ref": "Mediuswareltd",
                "account_name": "Mediusware Ltd.",
                "account_number": "1481100038741",
            },
        )

    def bonus_pdf(self, queryset, filter=None, bank=None):
        salary_disbursement = SalaryDisbursement.objects.filter(filter).first()

        festival_bonus_sheet = queryset.first()
        employee_festival_bonus_set = (
            festival_bonus_sheet.employeefestivalbonus_set.filter(
                employee__in=salary_disbursement.employee.all().values_list(
                    "id", flat=True
                )
            )
        )

        pdf = PDF()
        pdf.context = {
            "festival_bonus_sheet": festival_bonus_sheet,
            "employee_festival_bonus_set": employee_festival_bonus_set,
            "bank": bank,
            "seal": f"{settings.STATIC_ROOT}/stationary/sign_md.png",
        }
        pdf.template_path = "letters/bonus_pdf_v2.html"
        return pdf.render_to_pdf()

    def export_in_xl_bankasia(self, queryset, query_filter=None):
        wb = Workbook()
        work_sheets = {}
        for festival_bonus_sheet in queryset:

            festival_bonus_sheet.total_value = 0

            work_sheet = wb.create_sheet(title=str(festival_bonus_sheet.date))

            work_sheet.append(
                [
                    "Employee Name",
                    "Account Number",
                    "Dr./Cr.",
                    "Transaction Amount",
                    "Chqser",
                    "Chqnum",
                    "Chqdat",
                    "Remarks",
                ]
            )

            work_sheet.append(
                [
                    settings.COMPANY_ACCOUNT_NAME,
                    settings.COMPANY_ACCOUNT_NO,
                    "D",
                    "0",
                    "",
                    "",
                    "",
                    f'Festival Bonus for {festival_bonus_sheet.date.strftime("%b, %Y")}',
                ]
            )

            employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set
            if query_filter is not None:
                employee_festival_bonuses = (
                    festival_bonus_sheet.employeefestivalbonus_set.filter(
                        query_filter
                    ).all()
                )

            for employee_festival_bonus in employee_festival_bonuses.all():

                festival_bonus_sheet.total_value += floor(
                    employee_festival_bonus.amount
                )
                bank_account = employee_festival_bonus.employee.bankaccount_set.filter(
                    default=True, is_approved=True
                ).last()

                work_sheet.append(
                    [
                        employee_festival_bonus.employee.full_name,
                        bank_account.account_number if bank_account else "-",
                        "C",
                        floor(employee_festival_bonus.amount),
                        "",
                        "",
                        "",
                        f'Festival Bonus for {festival_bonus_sheet.date.strftime("%b, %Y")}',
                    ]
                )

            work_sheet["D2"] = festival_bonus_sheet.total_value
            work_sheets[str(festival_bonus_sheet.id)] = work_sheet
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=FestivalBonusSheet.xlsx"
        return response

    def export_in_xl_dbbl(self, queryset, query_filter=None):
        """

        @param queryset:
        @param query_filter:
        @return:
        """
        wb = Workbook()
        work_sheets = {}
        for festival_bonus_sheet in queryset:

            festival_bonus_sheet.total_value = 0

            work_sheet = wb.create_sheet(title=str(festival_bonus_sheet.date))

            work_sheet.append(
                ["name", "Basic Salary", "Bonus Amount", "Bank Name", "Bank Number"]
            )

            employee_festival_bonuses = festival_bonus_sheet.employeefestivalbonus_set
            if query_filter is not None:
                employee_festival_bonuses = (
                    festival_bonus_sheet.employeefestivalbonus_set.filter(
                        query_filter
                    ).all()
                )

            for employee_festival_bonus in employee_festival_bonuses.all():
                festival_bonus_sheet.total_value += floor(
                    employee_festival_bonus.amount
                )
                bank_account = employee_festival_bonus.employee.bankaccount_set.filter(
                    default=True, is_approved=True
                ).last()

                salary_history = (
                    employee_festival_bonus.employee.salaryhistory_set.filter(
                        active_from__lte=festival_bonus_sheet.date.replace(day=1)
                    ).last()
                )

                basic_salary = 0
                if salary_history:
                    basic_salary = (
                        salary_history.payable_salary / 100
                    ) * employee_festival_bonus.employee.pay_scale.basic

                work_sheet.append(
                    [
                        employee_festival_bonus.employee.full_name,
                        basic_salary,
                        employee_festival_bonus.amount,
                        bank_account.bank.name if bank_account else "",
                        bank_account.account_number if bank_account else "",
                    ]
                )
            work_sheet.append(["", "Total", festival_bonus_sheet.total_value])
            work_sheets[str(festival_bonus_sheet.id)] = work_sheet
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=FestivalBonusSheet.xlsx"
        return response
