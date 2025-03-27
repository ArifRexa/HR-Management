from math import floor
import math

from django.conf import settings
from django.contrib import admin, messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from account.models import SalaryDisbursement
from config.utils.pdf import PDF
from employee.models.bank_account import BEFTN, BankAccount


class SalarySheetAction(admin.ModelAdmin):
    actions = (
        "export_excel",
        "export_city_bank_npsb",
        "export_city_bank_npsb_excel",
        "export_city_bank_beftn",
        "export_bankasia_salary_acc_dis_excel",
        "export_salary_account_dis",
        "export_salary_account_dis_pdf",
        "export_bonus_account_dis_pdf",
    )

    def get_actions(self, request):
        actions = super().get_actions(request)
        if request.user.is_superuser:
            return actions
        return tuple()

    @admin.action(description="Export in Excel")
    def export_excel(self, request, queryset):
        return self.export_in_xl(queryset)

    @admin.action(description="Export City Bank BEFTN")
    def export_city_bank_beftn(self, request, queryset):
        wb = Workbook()
        work_sheet = wb.active
        work_sheet.title = "BEFTN Export"

        # Add headers including 'Total Salary' at the end
        work_sheet.append(
            [
                "Date",
                "Account No",
                "Employee Name",
                "BDT",
                "Amount",
                "Originating Bank Routing No.",
                "Routing No",
                "Originating Bank Account No.",
                "Originating Account Name",
                "Remarks",
            ]
        )

        for salary_sheet in queryset:
            beftn = BEFTN.objects.last()

            # Calculate total salary for the salary sheet
            # total_salary = salary_sheet.employeesalary_set.filter(
            #     employee__bankaccount__default=True,
            #     employee__bankaccount__is_approved=True
            # ).aggregate(Sum('gross_salary'))['gross_salary__sum']
            total_salary = 0
            for employee_salary in salary_sheet.employeesalary_set.all():
                # Get the bank account information
                bank_account = employee_salary.employee.bankaccount_set.filter(
                    default=True, is_approved=True
                ).last()

                if bank_account and employee_salary.gross_salary > 0:
                    # Append the employee salary data with total salary at the end
                    total_salary += int(employee_salary.gross_salary)
                    work_sheet.append(
                        [
                            salary_sheet.date.strftime("%d-%m-%Y"),  # Date
                            bank_account.account_number,  # Account No
                            employee_salary.employee.full_name,
                            "BDT",  # Currency
                            str(int(employee_salary.gross_salary)),  # Amount
                            beftn.originating_bank_routing_number,  # Originating Bank Routing No.
                            beftn.routing_no,  # Routing No
                            beftn.originating_bank_account_number,  # Originating Bank Account No.
                            beftn.originating_bank_account_name,  # Originating Account Name
                            f'Salary of {salary_sheet.date.strftime("%b, %Y")}',  # Remarks
                        ]
                    )
        work_sheet.append(
            [
                "",
                "",
                "",
                "",
                f"Total={int(total_salary)}",
                "",
                "",
                "",
                "",
                "",  # Total salary row below "Amount"
            ]
        )
        # Prepare the response with the Excel file
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=City_Bank_BEFTN.xlsx"
        return response

    @admin.action(description="Export City Bank NPSB(Excel)")
    def export_city_bank_npsb_excel(self, request, queryset):
        wb = Workbook()
        work_sheet = wb.active
        work_sheet.title = "NPSB Export"

        # Add headers including 'Total Salary' at the end
        work_sheet.append(
            [
                "SL No",
                "Name",
                "A/C No.",
                "Amount in Tk.",
            ]
        )
        salary_sheets = queryset

        total = 0
        salary_data = []
        for sheet in salary_sheets:
            employee_salaries = sheet.employeesalary_set.all()
            for emp_salary in employee_salaries:
                employee_bank = BankAccount.objects.filter(
                    employee=emp_salary.employee, default=True, is_approved=True
                ).first()
                if emp_salary.gross_salary <= 0 or not employee_bank:
                    continue
                gross_salary = emp_salary.gross_salary or 0.0
                salary_data.append(
                    {
                        "sheet_date": sheet.date,
                        "name": emp_salary.employee.full_name,
                        "account_number": (
                            employee_bank.account_number if employee_bank else "-"
                        ),
                        "gross_salary": math.ceil(gross_salary),
                    }
                )
                total += math.ceil(gross_salary)

        for index, data in enumerate(salary_data, start=1):
            work_sheet.append(
                [
                    index,
                    data["name"],
                    data["account_number"],
                    data["gross_salary"],
                ]
            )

        work_sheet.append(
            [
                "",
                "",
                "Total",
                f"{int(total)}",
            ]
        )

        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=City_Bank_npsb.xlsx"
        return response

    @admin.action(description="Export City Bank NPSB")
    def export_city_bank_npsb(self, request, queryset):

        salary_sheets = queryset

        salary_data = []
        for sheet in salary_sheets:
            employee_salaries = sheet.employeesalary_set.all()
            total = 0
            for emp_salary in employee_salaries:
                employee_bank = BankAccount.objects.filter(
                    employee=emp_salary.employee, default=True, is_approved=True
                ).first()
                if emp_salary.gross_salary <= 0 or not employee_bank:
                    continue
                gross_salary = emp_salary.gross_salary or 0.0
                salary_data.append(
                    {
                        "sheet_date": sheet.date,
                        "name": emp_salary.employee.full_name,
                        "account_number": (
                            employee_bank.account_number if employee_bank else "-"
                        ),
                        "gross_salary": math.ceil(gross_salary),
                    }
                )
                total += math.ceil(gross_salary)
        pdf = PDF(
            template_path="admin/city_bank_npsb.html",
            context={
                "salary_sheets": salary_data,
                "total": total,
                "month": sheet.date.strftime("%B"),
                "bank": {
                    "ref": "Mediusware",
                    "account_name": "Md. Shahinur Rahman",
                    "account_number": "1481510160023",
                },
            },
        )

        return pdf.render_to_pdf()

    @admin.action(description="Export Bank Asia Salary Account Disbursements (Excel)")
    def export_bankasia_salary_acc_dis_excel(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(
            disbursement_type="salary_account"
        ).first()
        return self.export_in_xl_bankasia(
            queryset,
            ("employee__in", salary_disbursement.employee.all()),
        )

    @admin.action(description="Export Personal Account Disbursements (Excel)")
    def export_personal_account_dis(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(
            disbursement_type="personal_account"
        ).first()
        return self.export_in_xl(
            queryset, ("employee__in", salary_disbursement.employee.all())
        )

    @admin.action(description="Export Salary Account Disbursements (Excel)")
    def export_salary_account_dis(self, request, queryset):
        salary_disbursement = SalaryDisbursement.objects.filter(
            disbursement_type="salary_account"
        ).first()
        return self.export_in_xl(
            queryset, ("employee__in", salary_disbursement.employee.all())
        )

    @admin.action(description="Export Personal Account Disbursements (PDF)")
    def export_personal_account_dis_pdf(self, request, queryset):
        return self.export_in_pdf(
            queryset=queryset,
            # TODO : i don't know whey it's salary account it's suppose
            # to personal account. as shahin vai asked to change so i change
            filter=("disbursement_type", "salary_account"),
            bank={
                "ref": "Mediusware",
                "account_name": "Md. Shahinur Rahman",
                "account_number": "1481510160023",
            },
        )

    @admin.action(description="Export Salary Account Disbursements (PDF)")
    def export_salary_account_dis_pdf(self, request, queryset):
        return self.export_in_pdf(
            queryset=queryset,
            filter=("disbursement_type", "salary_account"),
            bank={
                "ref": "Mediuswareltd",
                "account_name": "Mediusware Ltd.",
                "account_number": "1481100038741",
            },
        )

    @admin.action(description="Export Bonus Account Disbursements (PDF)")
    def export_bonus_account_dis_pdf(self, request, queryset):
        print(queryset.filter(festival_bonus=False).exists())
        if queryset.filter(festival_bonus=False).exists():
            return self.message_user(
                request,
                "In this salary sheet does not include festival bonus !!!",
                level=messages.WARNING,
            )

        return self.bonus_pdf(
            queryset=queryset,
            filter=("disbursement_type", "salary_account"),
            bank={
                "ref": "Mediuswareltd",
                "account_name": "Mediusware Ltd.",
                "account_number": "1481100038741",
            },
        )

    def export_in_pdf(self, queryset, filter=None, bank=None):
        salary_disbursement = SalaryDisbursement.objects.filter(filter).first()
        pdf = PDF()
        pdf.context = {
            "salary_sheet": queryset.first(),
            "employee_salary_set": queryset.first()
            .employeesalary_set.filter(employee__in=salary_disbursement.employee.all())
            .all(),
            "bank": bank,
            "seal": f"{settings.STATIC_ROOT}/stationary/sign_md.png",
        }
        pdf.template_path = "letters/bank_salary.html"
        return pdf.render_to_pdf()

    def bonus_pdf(self, queryset, filter=None, bank=None):
        salary_disbursement = SalaryDisbursement.objects.filter(filter).first()
        pdf = PDF()
        pdf.context = {
            "salary_sheet": queryset.first(),
            "employee_salary_set": queryset.first()
            .employeesalary_set.filter(employee__in=salary_disbursement.employee.all())
            .all(),
            "bank": bank,
            "seal": f"{settings.STATIC_ROOT}/stationary/sign_md.png",
        }
        pdf.template_path = "letters/bonus_pdf.html"
        return pdf.render_to_pdf()

    def export_in_xl(self, queryset, query_filter=None):
        """

        @param queryset:
        @param query_filter:
        @return:
        """
        wb = Workbook()
        work_sheets = {}
        for salary_sheet in queryset:
            salary_sheet.total_value = 0
            work_sheet = wb.create_sheet(title=str(salary_sheet.date))
            work_sheet.append(
                [
                    "name",
                    "Net Salary",
                    "Overtime",
                    "Project Bonus",
                    "Leave Bonus",
                    "Festival Bonus",
                    "Gross Salary",
                    "Bank Name",
                    "Bank Number",
                ]
            )
            employee_salaries = salary_sheet.employeesalary_set
            if query_filter is not None:
                employee_salaries = salary_sheet.employeesalary_set.filter(
                    query_filter
                ).all()
            for employee_salary in employee_salaries.all():
                # if employee_salary.employee.
                salary_sheet.total_value += floor(employee_salary.gross_salary)
                bank_account = employee_salary.employee.bankaccount_set.filter(
                    default=True
                ).first()
                work_sheet.append(
                    [
                        employee_salary.employee.full_name,
                        employee_salary.net_salary,
                        employee_salary.overtime,
                        employee_salary.project_bonus,
                        employee_salary.leave_bonus,
                        employee_salary.festival_bonus,
                        floor(employee_salary.gross_salary),
                        bank_account.bank.name if bank_account else "",
                        bank_account.account_number if bank_account else "",
                    ]
                )
            work_sheet.append(["", "", "", "", "", "Total", salary_sheet.total_value])
            work_sheets[str(salary_sheet.id)] = work_sheet
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=SalarySheet.xlsx"
        return response

    def export_in_xl_bankasia(self, queryset, query_filter=None):
        wb = Workbook()
        work_sheets = {}
        for salary_sheet in queryset:

            salary_sheet.total_value = 0

            work_sheet = wb.create_sheet(title=str(salary_sheet.date))

            work_sheet.append(
                [
                    "Employee Name",
                    "Account Number",
                    "Dr./Cr.",
                    "Transaction Amount",
                    "Chqser",
                    "Chqnum",
                    "Chqdat",
                    "Remarks",
                ]
            )

            work_sheet.append(
                [
                    settings.COMPANY_ACCOUNT_NAME,
                    settings.COMPANY_ACCOUNT_NO,
                    "D",
                    "0",
                    "",
                    "",
                    "",
                    f'Salary of {salary_sheet.date.strftime("%b, %Y")}',
                ]
            )

            employee_festival_bonuses = salary_sheet.employeesalary_set
            if query_filter is not None:
                employee_festival_bonuses = salary_sheet.employeesalary_set.filter(
                    query_filter
                ).all()

            for employee_salary in employee_festival_bonuses.all():

                salary_sheet.total_value += floor(employee_salary.gross_salary)
                bank_account = employee_salary.employee.bankaccount_set.filter(
                    is_approved=True, id=11
                ).last()

                work_sheet.append(
                    [
                        employee_salary.employee.full_name,
                        bank_account.account_number if bank_account else "-",
                        "C",
                        floor(employee_salary.gross_salary),
                        "",
                        "",
                        "",
                        f'Salary of {salary_sheet.date.strftime("%b, %Y")}',
                    ]
                )

            work_sheet["D2"] = salary_sheet.total_value
            work_sheets[str(salary_sheet.id)] = work_sheet
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = "attachment; filename=SalarySheet.xlsx"
        return response
