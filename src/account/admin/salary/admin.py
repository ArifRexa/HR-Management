import calendar
from datetime import datetime
from math import floor
import re
from typing import Any, Optional, Sequence

from django.contrib import admin
from django.contrib.humanize.templatetags.humanize import intcomma
from django.db.models import Sum
from django.http import HttpResponse
from django.http.request import HttpRequest
from django.utils.html import format_html
from num2words import num2words
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from account.admin.salary.actions import SalarySheetAction
from account.models import Loan, SalarySheet, EmployeeSalary,SalaryReport
from account.repository.SalarySheetRepository import SalarySheetRepository
from employee.models.employee import LateAttendanceFine
from decimal import Decimal
from django.db.models import Count
from employee.models import Employee

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from employee.templatetags.employee_helper import total_employee_count
from job_board.models import job

class EmployeeSalaryInline(admin.TabularInline):
    model = EmployeeSalary
    extra = 0
    template= "admin/employee_salary.html"
    exclude = [
        'provident_fund',
        'code_quality_bonus',
        'festival_bonus',
        'device_allowance',
        "loan_emi",
        'net_salary', 'overtime',
        'project_bonus', 'leave_bonus', #'festival_bonus', 
        'food_allowance',
        'gross_salary',
    ]
    readonly_fields = (
        'employee', 'formatted_net_salary', 'formatted_overtime',
        'formatted_project_bonus', 
        'formatted_food_allowance', 'formatted_festival_bonus', 'formatted_leave_bonus', 'get_tax_loan',"get_salary_loan",

        # 'provident_fund', 'code_quality_bonus',
        'get_late_fine',
        'formatted_gross_salary', #'get_details',

    )
    superadminonly_fields = (
        'net_salary',
        'leave_bonus',
        'gross_salary',
    )

    can_delete = False

    def formatted_net_salary(self, obj):
        return int(obj.net_salary)
    
    def formatted_overtime(self, obj):
        return int(obj.overtime)
    
    def formatted_project_bonus(self, obj):
        return int(obj.project_bonus)

    def formatted_leave_bonus(self, obj):
        return int(obj.leave_bonus)
    
    def formatted_food_allowance(self, obj):
        return int(obj.food_allowance)
    
    def formatted_festival_bonus(self, obj):
        return int(obj.festival_bonus)
    def formatted_gross_salary(self, obj):
        return int(obj.gross_salary)
    
    formatted_net_salary.short_description = "Net Salary"
    formatted_overtime.short_description = "Overtime"
    formatted_project_bonus.short_description = "Project Bonus"
    formatted_leave_bonus.short_description = "Leave Bonus"
    formatted_food_allowance.short_description = "Food Allowance"
    formatted_festival_bonus.short_description = "Festival Bonus"
    formatted_gross_salary.short_description = "Gross Bonus"


    def get_tax_loan(self, obj):
        salary_loan = obj.employee.loan_set.filter(
            start_date__month=obj.salary_sheet.date.month,
            end_date__year=obj.salary_sheet.date.year,
            loan_type="tds"
        )
        loan_amount = salary_loan.aggregate(Sum("emi"))
        return int(loan_amount["emi__sum"]) if loan_amount["emi__sum"] else 0
    get_tax_loan.short_description = 'Tax Loan'

    def get_exclude(self, request, obj=None):
        exclude = list(super().get_exclude(request, obj))
        if not request.user.is_superuser and not request.user.has_perm('account.can_see_salary_on_salary_sheet'):
            exclude.extend(self.superadminonly_fields)
        return exclude
    def get_late_fine(self, obj):
        
        fine = LateAttendanceFine.objects.filter(
                    employee=obj.employee,
                    month=obj.salary_sheet.date.month,
                    year=obj.salary_sheet.date.year,
                ).aggregate(fine=Sum('total_late_attendance_fine'))
        return int(fine.get('fine', 0)) if fine.get('fine') else 0
    get_late_fine.short_description = "Late Fine"

    # def get_salary_loan(self, obj):
    #     salary_loan = obj.employee.loan_set.filter(
    #         start_date__month=obj.salary_sheet.date.month,
    #         end_date__month=obj.salary_sheet.date.month,
    #         loan_type="salary"
    #     )
    #     loan_amount = salary_loan.aggregate(Sum("emi"))
    #     return int(loan_amount["emi__sum"]) if loan_amount["emi__sum"] else 0
    def get_salary_loan(self, obj):
        salary_date = obj.salary_sheet.date
        salary_month_start = datetime(salary_date.year, salary_date.month, 1).date()
        salary_month_end = datetime(
            salary_date.year,
            salary_date.month,
            calendar.monthrange(salary_date.year, salary_date.month)[1],
        ).date()

        loan_amount = obj.employee.loan_set.filter(
            start_date__lte=salary_month_end,
            end_date__gte=salary_month_start,
            loan_type="salary",
        ).aggregate(Sum("emi"))
        return int(loan_amount["emi__sum"]) if loan_amount["emi__sum"] else 0
    get_salary_loan.short_description = "Salary Loan"

    def get_readonly_fields(self, request, obj=None):
        readonly_fields = super().get_readonly_fields(request, obj)
        if not request.user.is_superuser and not request.user.has_perm('account.can_see_salary_on_salary_sheet'):
            readonly_fields = [field for field in readonly_fields if field not in self.superadminonly_fields]
        return readonly_fields

    @admin.display(description="More Info")
    def get_details(self, obj, *args, **kwargs):
        return format_html(f'<a href="/media/temp_emp_salary/{obj.employee.id}.txt" download>Download</a>')


    def has_add_permission(self, request, obj):
        return False
    
     

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        
        # Filter by the current SalarySheet
        salary_sheet_id = self.parent_model.objects.filter(id=request.resolver_match.kwargs.get('object_id')).values_list('id', flat=True).first()
        qs = qs.filter(salary_sheet_id=salary_sheet_id)
        
        # Calculate totals
        self.total_values = qs.aggregate(
            total_employees = Count('id'),
            total_net_salary=Sum('net_salary'),
            total_overtime=Sum('overtime'),
            total_project_bonus=Sum('project_bonus'),
            total_leave_bonus=Sum('leave_bonus'),
            total_food_allowance=Sum('food_allowance'),
            total_festival_bonus=Sum('festival_bonus'),
            total_gross_salary=Sum('gross_salary'),
        )
        
        # Initialize total_late_fine as Decimal
        total_late_fine = Decimal('0.00')
        total_tax_loan = Decimal('0.00')
        total_salary_loan = Decimal('0.00')

        # Calculate custom total for late fines manually
        for obj in qs:
            total_late_fine += Decimal(self.get_late_fine(obj))
            total_tax_loan += Decimal(self.get_tax_loan(obj))
            total_salary_loan+= Decimal(self.get_salary_loan(obj))
        
        self.total_values['total_late_fine'] = total_late_fine
        self.total_values['total_tax_loan'] = total_tax_loan
        self.total_values['total_salary_loan'] = total_salary_loan

        return qs
@admin.register(SalarySheet)
class SalarySheetAdmin(SalarySheetAction, admin.ModelAdmin):
    list_display = ('date', 'created_at', 'total', 'total_employee', 'festival_bonus')
    fields = ('date', 'festival_bonus')
    inlines = (EmployeeSalaryInline,)
    change_form_template = ('admin/salary_sheet.html')
    

    def save_model(self, request, salary_sheet, form, change):
        # TODO : add festival bonus
        salary = SalarySheetRepository(request.POST['date'])
        if 'festival_bonus' in request.POST and request.POST['festival_bonus'] == 'on':
            salary.festival_bonus = True
        salary.save()
    
    def get_list_display(self, request):
        list_display = list(super().get_list_display(request))
        if not request.user.is_superuser and not request.user.has_perm('account.can_see_salary_on_salary_sheet') and 'total' in list_display:
            print(request.user, request.user.has_perm('account.can_see_salary_on_salary_sheet'))
            list_display.remove('total')
        return tuple(list_display)
    
    def total(self, obj):
        total_value = EmployeeSalary.objects.filter(salary_sheet_id=obj.id).aggregate(Sum('gross_salary'))[
            'gross_salary__sum']
        if total_value:
            return format_html(
                f'<b>{intcomma(floor(total_value))}</b> <br>'
                f'{num2words(floor(total_value)).capitalize()}'
            )

    def total_employee(self, obj):
        return obj.employeesalary_set.count()
    
 # Override change_view to pass the calculated totals to the template
    def change_view(self, request, object_id, form_url='', extra_context=None):
        salary_sheet = self.get_object(request, object_id)
        inline_instance = self.get_inline_instances(request, salary_sheet)[0]
        inline_instance_queryset = inline_instance.get_queryset(request)

        # Add total values from EmployeeSalaryInline to the context
        extra_context = extra_context or {}
        extra_context['total_values'] = inline_instance.total_values

        return super().change_view(request, object_id, form_url, extra_context=extra_context)
    

# @admin.register(SalaryReport)
# class SalaryReportAdmin(admin.ModelAdmin):
#     list_display = ('start_date', 'end_date')
#     change_form_template = "admin/salary_report.html"
#     actions = ['export_as_excel']
#
#     def change_view(self, request, object_id, form_url='', extra_context=None):
#         salary_report = self.get_object(request, object_id)
#         start_date = salary_report.start_date
#         end_date = salary_report.end_date
#
#         employees = Employee.objects.filter(
#            joining_date__lte=start_date
#         ).exclude(salaryhistory__isnull=True)
#
#         employee_salary_data = []
#         total_employee_salary = 0
#
#         # Initialize total counters for all fields
#         totals = {
#             'gross_salary': 0,
#             'basic_salary': 0,
#             'house_allowance': 0,
#             'conveyance': 0,
#             'medical_allowance': 0,
#             'project_bonus': 0,
#             'overtime': 0,
#             'festival_bonus': 0,
#             'food_allowance': 0,
#             'leave_bonus': 0,
#             'tds': 0,
#             'salary_loan': 0,
#             'late_fine': 0,
#         }
#
#         for employee in employees:
#             employee_salarys = EmployeeSalary.objects.filter(employee=employee, salary_sheet__date__range=[start_date, end_date])
#
#             total_gross_salary = 0
#             total_basic_salary = 0
#             total_house_allowance = 0
#             total_conveyance = 0
#             total_medical_allowance = 0
#             total_project_bonus = 0
#             total_overtime = 0
#             total_festival_bonus = 0
#             total_food_allowance = 0
#             total_leave_bonus = 0
#             total_tds = 0
#
#             for emp_salary in employee_salarys:
#                 net_salary = emp_salary.net_salary
#                 gross_salary = emp_salary.gross_salary
#                 total_employee_salary += gross_salary
#                 total_gross_salary += gross_salary
#                 total_basic_salary += net_salary * 0.55
#                 total_house_allowance += net_salary * 0.20
#                 total_conveyance += net_salary * 0.15
#                 total_medical_allowance += net_salary * 0.10
#                 total_project_bonus += emp_salary.project_bonus
#                 total_overtime += emp_salary.overtime
#                 total_festival_bonus += emp_salary.festival_bonus
#                 total_food_allowance += emp_salary.food_allowance
#                 total_tds += emp_salary.loan_emi
#                 total_leave_bonus += emp_salary.leave_bonus
#
#             total_salary_emi = Loan.objects.filter(
#                 employee=employee,
#                 loan_type='salary',
#                 created_at__range=[start_date, end_date]
#             ).aggregate(total_emi=Sum('emi'))['total_emi'] or 0
#
#             total_late_fine = LateAttendanceFine.objects.filter(employee=employee, date__range=[start_date, end_date]).aggregate(late_fine=Sum('total_late_attendance_fine'))['late_fine'] or 0
#
#             # Add individual totals to the overall totals
#             totals['gross_salary'] += total_gross_salary
#             totals['basic_salary'] += total_basic_salary
#             totals['house_allowance'] += total_house_allowance
#             totals['conveyance'] += total_conveyance
#             totals['medical_allowance'] += total_medical_allowance
#             totals['project_bonus'] += total_project_bonus
#             totals['overtime'] += total_overtime
#             totals['festival_bonus'] += total_festival_bonus
#             totals['food_allowance'] += total_food_allowance
#             totals['leave_bonus'] += total_leave_bonus
#             totals['tds'] += total_tds
#             totals['salary_loan'] += total_salary_emi
#             totals['late_fine'] += total_late_fine
#
#             employee_salary_data.append({
#                 'index': len(employee_salary_data) + 1,
#                 'name': employee.full_name,
#                 'designation': employee.designation.title,
#                 'gross_salary': total_gross_salary,
#                 'basic_salary': total_basic_salary,
#                 'house_allowance': total_house_allowance,
#                 'conveyance': total_conveyance,
#                 'medical_allowance': total_medical_allowance,
#                 'project_bonus': total_project_bonus,
#                 'overtime': total_overtime,
#                 'festival_bonus': total_festival_bonus,
#                 'leave_bonus': total_leave_bonus,
#                 'food_allowance': total_food_allowance,
#                 'tds': total_tds,
#                 'salary_loan': total_salary_emi,
#                 'late_fine': total_late_fine,
#             })
#
#         extra_context = extra_context or {}
#         extra_context['start_date'] = start_date
#         extra_context['end_date'] = end_date
#         extra_context['total_employee_salary'] = total_employee_salary
#         extra_context['employee_salary_data'] = employee_salary_data
#         extra_context['totals'] = totals  # Pass the totals to the template
#
#         return super().change_view(request, object_id, form_url, extra_context)
#
#     def export_as_excel(self, request, queryset):
#             if queryset.count() != 1:
#                 self.message_user(request, "Please select exactly one Salary Report.", level='error')
#                 return
#
#             salary_report = queryset.first()
#             start_date = salary_report.start_date
#             end_date = salary_report.end_date
#
#             employees = Employee.objects.filter(
#                 active=True, joining_date__lte=start_date
#             ).exclude(salaryhistory__isnull=True)
#
#             employee_salary_data = []
#
#             for employee in employees:
#                 employee_salarys = EmployeeSalary.objects.filter(employee=employee, salary_sheet__date__range=[start_date, end_date])
#
#                 total_gross_salary = 0
#                 total_basic_salary = 0
#                 total_house_allowance = 0
#                 total_conveyance = 0
#                 total_medical_allowance = 0
#                 total_project_bonus = 0
#                 total_overtime = 0
#                 total_festival_bonus = 0
#                 total_food_allowance = 0
#                 total_leave_bonus = 0
#                 total_tds = 0
#
#                 for emp_salary in employee_salarys:
#                     net_salary = emp_salary.net_salary
#                     gross_salary = emp_salary.gross_salary
#
#                     total_gross_salary += gross_salary
#                     total_basic_salary += net_salary * 0.55
#                     total_house_allowance += net_salary * 0.20
#                     total_conveyance += net_salary * 0.15
#                     total_medical_allowance += net_salary * 0.10
#                     total_project_bonus += emp_salary.project_bonus
#                     total_overtime += emp_salary.overtime
#                     total_festival_bonus += emp_salary.festival_bonus
#                     total_food_allowance += emp_salary.food_allowance
#                     total_tds += emp_salary.loan_emi
#                     total_leave_bonus += emp_salary.leave_bonus
#
#                 loans = Loan.objects.filter(employee=employee, loan_type='tds', created_at__range=[start_date, end_date]).values_list('tax_calan_no', flat=True)
#                 loan_list = [loan for loan in loans if loan is not None]
#                 tressury_challn_no = ',<br>'.join(loan_list)
#
#                 total_salary_emi = Loan.objects.filter(
#                     employee=employee,
#                     loan_type='salary',
#                     created_at__range=[start_date, end_date]
#                 ).aggregate(total_emi=Sum('emi'))['total_emi']
#
#                 total_late_fine = LateAttendanceFine.objects.filter(employee=employee, date__range=[start_date, end_date]).aggregate(late_fine=Sum('total_late_attendance_fine'))['late_fine']
#
#                 tin = employee.tax_info or ''
#
#                 employee_salary_data.append({
#                     'index': len(employee_salary_data) + 1,
#                     'name': employee.full_name,
#                     'designation': employee.designation.title,
#                     'tin': tin,
#                     'gross_salary': total_gross_salary,
#                     'basic_salary': total_basic_salary,
#                     'house_allowance': total_house_allowance,
#                     'conveyance': total_conveyance,
#                     'medical_allowance': total_medical_allowance,
#                     'project_bonus': total_project_bonus,
#                     'overtime': total_overtime,
#                     'festival_bonus': total_festival_bonus,
#                     'leave_bonus': total_leave_bonus,
#                     'food_allowance': total_food_allowance,
#                     'tds': total_tds,
#                     'salary_loan': total_salary_emi,
#                     'late_fine': total_late_fine,
#                     'chalan_no': tressury_challn_no
#                 })
#
#             # Generate Excel
#             wb = openpyxl.Workbook()
#             ws = wb.active
#             ws.title = "Salary Report"
#
#             # Headers for Excel sheet
#             headers = [
#                 'SL No', 'Name of Employee', 'Designation', 'TIN', 'Total Salary', 'Basic (55%)',
#                 'House Allowance (20%)', 'Conveyance (15%)', 'Medical Allowance (10%)',
#                 'Project Bonus', 'Overtime', 'Festival Bonus', 'Leave Bonus',
#                 'Food Allowance', 'TDS', 'Salary Loan EMI', 'Late Fine', 'Treasury Challan No'
#             ]
#             ws.append(headers)
#
#             # Populate Excel rows
#             for data in employee_salary_data:
#                 ws.append([
#                     data['index'], data['name'], data['designation'], data['tin'],
#                     data['gross_salary'], data['basic_salary'], data['house_allowance'],
#                     data['conveyance'], data['medical_allowance'], data['project_bonus'],
#                     data['overtime'], data['festival_bonus'], data['leave_bonus'],
#                     data['food_allowance'], data['tds'], data['salary_loan'],
#                     data['late_fine'], data['chalan_no']
#                 ])
#
#             # Save workbook to HTTP response
#             response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#             response['Content-Disposition'] = f'attachment; filename=Salary_Report_{start_date}_to_{end_date}.xlsx'
#             wb.save(response)
#             return response


@admin.register(SalaryReport)
class SalaryReportAdmin(admin.ModelAdmin):
    list_display = ('start_date', 'end_date')
    change_form_template = "admin/salary_report.html"
    actions = ['export_as_excel']

    def change_view(self, request, object_id, form_url='', extra_context=None):
        salary_report = self.get_object(request, object_id)
        start_date = salary_report.start_date
        end_date = salary_report.end_date

        employees = Employee.objects.filter(
            joining_date__lte=start_date
        ).exclude(salaryhistory__isnull=True)

        employee_salary_data = []
        total_employee_salary = 0

        # Initialize total counters for all fields
        totals = {
            'gross_salary': 0,
            'basic_salary': 0,
            'house_allowance': 0,
            'conveyance': 0,
            'medical_allowance': 0,
            'project_bonus': 0,
            'overtime': 0,
            'festival_bonus': 0,
            'food_allowance': 0,
            'leave_bonus': 0,
            'tds': 0,
            'salary_loan': 0,
            'late_fine': 0,
        }

        for employee in employees:
            employee_salarys = EmployeeSalary.objects.filter(employee=employee, salary_sheet__date__range=[start_date, end_date])

            total_gross_salary = 0
            total_basic_salary = 0
            total_house_allowance = 0
            total_conveyance = 0
            total_medical_allowance = 0
            total_project_bonus = 0
            total_overtime = 0
            total_festival_bonus = 0
            total_food_allowance = 0
            total_leave_bonus = 0
            total_tds = 0

            for emp_salary in employee_salarys:
                net_salary = emp_salary.net_salary
                gross_salary = emp_salary.gross_salary
                total_employee_salary += gross_salary
                total_gross_salary += gross_salary
                total_basic_salary += net_salary * 0.55
                total_house_allowance += net_salary * 0.20
                total_conveyance += net_salary * 0.15
                total_medical_allowance += net_salary * 0.10
                total_project_bonus += emp_salary.project_bonus
                total_overtime += emp_salary.overtime
                total_festival_bonus += emp_salary.festival_bonus
                total_food_allowance += emp_salary.food_allowance
                total_tds += emp_salary.loan_emi
                total_leave_bonus += emp_salary.leave_bonus

            total_salary_emi = Loan.objects.filter(
                employee=employee,
                loan_type='salary',
                created_at__range=[start_date, end_date]
            ).aggregate(total_emi=Sum('emi'))['total_emi'] or 0

            total_late_fine = LateAttendanceFine.objects.filter(
                employee=employee,
                date__range=[start_date, end_date]
            ).aggregate(late_fine=Sum('total_late_attendance_fine'))['late_fine'] or 0

            # Only add to employee_salary_data if total_gross_salary is not zero
            if total_gross_salary > 0:
                # Add individual totals to the overall totals
                totals['gross_salary'] += total_gross_salary
                totals['basic_salary'] += total_basic_salary
                totals['house_allowance'] += total_house_allowance
                totals['conveyance'] += total_conveyance
                totals['medical_allowance'] += total_medical_allowance
                totals['project_bonus'] += total_project_bonus
                totals['overtime'] += total_overtime
                totals['festival_bonus'] += total_festival_bonus
                totals['food_allowance'] += total_food_allowance
                totals['leave_bonus'] += total_leave_bonus
                totals['tds'] += total_tds
                totals['salary_loan'] += total_salary_emi
                totals['late_fine'] += total_late_fine

                employee_salary_data.append({
                    'index': len(employee_salary_data) + 1,
                    'name': employee.full_name,
                    'designation': employee.designation.title,
                    'gross_salary': total_gross_salary,
                    'basic_salary': total_basic_salary,
                    'house_allowance': total_house_allowance,
                    'conveyance': total_conveyance,
                    'medical_allowance': total_medical_allowance,
                    'project_bonus': total_project_bonus,
                    'overtime': total_overtime,
                    'festival_bonus': total_festival_bonus,
                    'leave_bonus': total_leave_bonus,
                    'food_allowance': total_food_allowance,
                    'tds': total_tds,
                    'salary_loan': total_salary_emi,
                    'late_fine': total_late_fine,
                })

        extra_context = extra_context or {}
        extra_context['start_date'] = start_date
        extra_context['end_date'] = end_date
        extra_context['total_employee_salary'] = total_employee_salary
        extra_context['employee_salary_data'] = employee_salary_data
        extra_context['totals'] = totals

        return super().change_view(request, object_id, form_url, extra_context)

    def export_as_excel(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(request, "Please select exactly one Salary Report.", level='error')
            return

        salary_report = queryset.first()
        start_date = salary_report.start_date
        end_date = salary_report.end_date

        # employees = Employee.objects.filter(
        #     active=True, joining_date__lte=start_date
        # ).exclude(salaryhistory__isnull=True)


        employees = Employee.objects.filter(
            joining_date__lte=start_date
        ).exclude(salaryhistory__isnull=True)

        employee_salary_data = []

        for employee in employees:
            employee_salarys = EmployeeSalary.objects.filter(
                employee=employee,
                salary_sheet__date__range=[start_date, end_date]
            )

            total_gross_salary = 0
            total_basic_salary = 0
            total_house_allowance = 0
            total_conveyance = 0
            total_medical_allowance = 0
            total_project_bonus = 0
            total_overtime = 0
            total_festival_bonus = 0
            total_food_allowance = 0
            total_leave_bonus = 0
            total_tds = 0

            for emp_salary in employee_salarys:
                net_salary = emp_salary.net_salary
                gross_salary = emp_salary.gross_salary

                total_gross_salary += gross_salary
                total_basic_salary += net_salary * 0.55
                total_house_allowance += net_salary * 0.20
                total_conveyance += net_salary * 0.15
                total_medical_allowance += net_salary * 0.10
                total_project_bonus += emp_salary.project_bonus
                total_overtime += emp_salary.overtime
                total_festival_bonus += emp_salary.festival_bonus
                total_food_allowance += emp_salary.food_allowance
                total_tds += emp_salary.loan_emi
                total_leave_bonus += emp_salary.leave_bonus

            # Only process if total_gross_salary is not zero
            if total_gross_salary > 0:
                loans = Loan.objects.filter(
                    employee=employee,
                    loan_type='tds',
                    created_at__range=[start_date, end_date]
                ).values_list('tax_calan_no', flat=True)

                loan_list = [loan for loan in loans if loan is not None]
                tressury_challn_no = ',<br>'.join(loan_list)

                total_salary_emi = Loan.objects.filter(
                    employee=employee,
                    loan_type='salary',
                    created_at__range=[start_date, end_date]
                ).aggregate(total_emi=Sum('emi'))['total_emi'] or 0

                total_late_fine = LateAttendanceFine.objects.filter(
                    employee=employee,
                    date__range=[start_date, end_date]
                ).aggregate(late_fine=Sum('total_late_attendance_fine'))['late_fine'] or 0

                tin = employee.tax_info or ''

                employee_salary_data.append({
                    'index': len(employee_salary_data) + 1,
                    'name': employee.full_name,
                    'designation': employee.designation.title,
                    'tin': tin,
                    'gross_salary': total_gross_salary,
                    'basic_salary': total_basic_salary,
                    'house_allowance': total_house_allowance,
                    'conveyance': total_conveyance,
                    'medical_allowance': total_medical_allowance,
                    'project_bonus': total_project_bonus,
                    'overtime': total_overtime,
                    'festival_bonus': total_festival_bonus,
                    'leave_bonus': total_leave_bonus,
                    'food_allowance': total_food_allowance,
                    'tds': total_tds,
                    'salary_loan': total_salary_emi,
                    'late_fine': total_late_fine,
                    'chalan_no': tressury_challn_no
                })

        # Generate Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Report"

        # Headers for Excel sheet
        headers = [
            'SL No', 'Name of Employee', 'Designation', 'TIN', 'Total Salary', 'Basic (55%)',
            'House Allowance (20%)', 'Conveyance (15%)', 'Medical Allowance (10%)',
            'Project Bonus', 'Overtime', 'Festival Bonus', 'Leave Bonus',
            'Food Allowance', 'TDS', 'Salary Loan EMI', 'Late Fine', 'Treasury Challan No'
        ]
        ws.append(headers)

        # Populate Excel rows
        for data in employee_salary_data:
            ws.append([
                data['index'], data['name'], data['designation'], data['tin'],
                data['gross_salary'], data['basic_salary'], data['house_allowance'],
                data['conveyance'], data['medical_allowance'], data['project_bonus'],
                data['overtime'], data['festival_bonus'], data['leave_bonus'],
                data['food_allowance'], data['tds'], data['salary_loan'],
                data['late_fine'], data['chalan_no']
            ])

        # Save workbook to HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename=Salary_Report_{start_date}_to_{end_date}.xlsx'
        wb.save(response)
        return response

        