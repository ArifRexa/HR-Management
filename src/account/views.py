from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
from account.models import AccountJournal
from django.utils import timezone
from django.template.loader import get_template
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from weasyprint import HTML
from django.db.models.functions import TruncDate
from django.db.models import Sum, Count
from openpyxl import Workbook
from openpyxl.styles import Alignment

@require_http_methods(["POST", "GET"])
@login_required(login_url="/admin/login/")
def payment_voucher(request, id):
    voucher = get_object_or_404(AccountJournal, id=id)
    
    # get the template
    template = get_template('pdf/payment_voucher.html')
    
    # get the context data
    context = {'voucher': voucher}

    # Render the html template with the context data.
    html_content = template.render(context)

    # Create weasyprint object from the html
    html = HTML(string=html_content)

    # generate pdf
    pdf_file = html.write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = str(timezone.now())
    response['Content-Disposition'] = f'attachment; filename="payment-voucher-{filename}.pdf"'
    
    return response

@require_http_methods(["POST", "GET"])
@login_required(login_url="/admin/login/")
def account_journal(request, id):
    monthly_journal = get_object_or_404(AccountJournal, id=id)
    # create a wrokbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Date Header
    worksheet.merge_cells('A1:A2')
    worksheet.merge_cells('B1:B2')
    worksheet.merge_cells('C1:C2')
    worksheet.merge_cells('F1:F2')
    worksheet.merge_cells('D1:E1')

    a1 = worksheet.cell(row=1, column=1, value='Date')
    a1.alignment = Alignment(horizontal='center', vertical='center')

    b1 = worksheet.cell(row=1, column=2, value='Account Code')
    b1.alignment = Alignment(horizontal='center', vertical='center')

    c1 = worksheet.cell(row=1, column=3, value = 'Head')
    c1.alignment = Alignment(horizontal='center', vertical='center')
    
    d1 = worksheet.cell(row=1, column=4, value='BDT')
    d1.alignment = Alignment(horizontal='center', vertical='center')

    f1 = worksheet.cell(row=1, column=6, value='Detail')
    f1.alignment = Alignment(horizontal='center', vertical='center')

    # Sub header
    d2 = worksheet.cell(row=2, column=4, value='Debit')
    d2.alignment = Alignment(horizontal='center', vertical='center')
    e2 = worksheet.cell(row=2, column=5, value='Credit')
    e2.alignment = Alignment(horizontal='center', vertical='center')

    # data calculation 
    expense_dates = monthly_journal.expenses.annotate(day=TruncDate('date')) \
                                .values('day') \
                                .annotate(count=Count('id')) \
                                .order_by('day')
    
    print(list(expense_dates))
    for row_num, data in enumerate(list(expense_dates), start=3):
        day = data['day']
        count = data['count']
        if count > 0:
            merge = row_num + count
        else:
            merge = row_num
        worksheet.merge_cells(start_row=row_num, start_column=1, end_row=merge - 1, end_column=1)
        worksheet[f'A{row_num}'] = f'{day}'

    for expense_date in expense_dates:
        expenses_data = monthly_journal.expenses.filter(date=expense_date['day']) \
                                        .values('expanse_group__account_code', 'expanse_group__title') \
                                        .order_by('expanse_group__account_code') \
                                        .annotate(expense_amount=Sum('amount'))
        

    # Create a response with the Excel file
    file_name = str(timezone.now())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=account-journal-{file_name}.xlsx'
    workbook.save(response)
    return response