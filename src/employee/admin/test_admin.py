from functools import update_wrapper
import datetime
import math

from django.contrib import admin, messages
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone

from django.db.models import Prefetch, Count, Q

# Needed for optional Features
# from django.db.models import Count, Case, When, Value, BooleanField
from django.shortcuts import redirect
from openpyxl import Workbook

from employee.models import (
    EmployeeAttendance,
    Employee,
    PrayerInfo,
)

from employee.forms.prayer_info import EmployeePrayerInfoForm


from config.settings import employee_ids as management_ids

from django.db.models import OuterRef, Subquery, BooleanField, Value


from employee.models.employee_activity import EmployeeOnline
from project_management.models import DailyProjectUpdate, EmployeeProjectHour


class EmployeeAttendanceOne(EmployeeAttendance):
    class Meta:
        proxy = True


class EmployeeAttendanceThree(EmployeeAttendance):
    class Meta:
        proxy = True


class EmployeeAttendanceTwo(EmployeeAttendance):
    class Meta:
        proxy = True


# base admin class for this all model
from django.core.cache import cache
from django.db.models import F, Sum, Case, When, Value, BooleanField, Exists
from collections import defaultdict
from datetime import timedelta


def sToTime(duration):
    minutes = math.floor((duration / 60) % 60)
    hours = math.floor((duration / (60 * 60)) % 24)

    return f"{hours:01}h: {minutes:01}m"


@admin.register(EmployeeAttendanceOne)
class EmployeeAttendanceOneAdmin(admin.ModelAdmin):
    list_display = ("date", "employee", "entry_time", "exit_time")
    date_hierarchy = "date"
    list_filter = ("employee",)
    search_fields = ("employee__full_name", "date")
    autocomplete_fields = ("employee",)

    def custom_changelist_view(self, request, *args, **kwargs) -> TemplateResponse:
        if not request.user.is_authenticated:
            return redirect("/")

        now = timezone.now()
        DEFAULT_EXIT_HOUR = 12 + 8  # 24 Hour time == 9 pm
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_range = 30
        if request.user.has_perm("employee.can_see_full_month_attendance"):
            day_range = 30

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_range)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]
        last_x_date = (now - datetime.timedelta(day_range)).date()
        last_month = (now.replace(day=1) - datetime.timedelta(days=1)).date()

        # Optimized query with prefetch_related to avoid N+1 problems
        emps = (
            Employee.objects.filter(
                active=True,
                show_in_attendance_list=True,
            )
            .select_related("user")
            .prefetch_related(
                Prefetch(
                    "employeeattendance_set",
                    queryset=EmployeeAttendance.objects.filter(
                        date__gte=last_x_date
                    ).prefetch_related("employeeactivity_set"),
                    to_attr="prefetched_attendances",
                ),
                Prefetch(
                    "dailyprojectupdate_employee",
                    queryset=DailyProjectUpdate.objects.filter(
                        created_at__date__gte=last_x_date, status="approved"
                    ).select_related("manager", "employee", "manager__employeeproject"),
                    to_attr="prefetched_updates",
                ),
                "employeeonline",
                "lateattendancefine_set",
                "employeeprojecthour_set",
                "projecthour_set",
                # Prefetch(
                #     "employeeprojecthour_set",
                #     queryset=EmployeeProjectHour.objects.select_related(
                #         "project_hour"
                #     ).all(),
                #     # to_attr="prefetched_projecthour",
                # ),
            )
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=last_month.year,
                        employeeattendance__date__month=last_month.month,
                    ),
                ),
                # is_online=Exists(
                #     EmployeeOnline.objects.filter(
                #         employee=OuterRef('pk'),
                #         active=True
                #     )
                # )
            ).all()
            .order_by("full_name")
        )
        e = emps.first()
        # Convert queryset to list for manipulation
        emps_list = list(emps)

        # Find and move the current user's employee to the top
        # user_data = None
        # for index, emp in enumerate(emps_list):
        #     if emp.user == request.user:
        #         user_data = emps_list.pop(index)
        #         break

        # if user_data:
        #     emps_list.insert(0, user_data)

        date_datas = {}
        manager_date_and_hours = {}

        # Pre-calculate manager hours
        for emp in emps_list:
            for update in emp.prefetched_updates:
                if update.manager_id != update.employee_id:
                    date = update.created_at.date()
                    manager_id = update.manager_id

                    if manager_id not in manager_date_and_hours:
                        manager_date_and_hours[manager_id] = {}

                    if date not in manager_date_and_hours[manager_id]:
                        manager_date_and_hours[manager_id][date] = update.hours
                    else:
                        manager_date_and_hours[manager_id][date] += update.hours

        # Process employee data efficiently
        for emp in emps_list:
            temp = {date: {} for date in last_x_dates}

            # Process daily project updates
            for update in emp.prefetched_updates:
                date = update.created_at.date()
                if date in temp:
                    if "accepted_hour" not in temp[date]:
                        temp[date]["accepted_hour"] = update.hours
                    else:
                        temp[date]["accepted_hour"] += update.hours

            # Process attendance data
            for attendance in emp.prefetched_attendances:
                date = attendance.date
                if date in temp:
                    activities = list(attendance.employeeactivity_set.all())
                    if activities:
                        activities.sort(key=lambda x: x.start_time)
                        start_time = activities[0].start_time
                        end_time = activities[-1].end_time
                        is_updated_by_bot = activities[-1].is_updated_by_bot
                        break_time = 0
                        inside_time = 0

                        # Calculate break time
                        for i in range(len(activities) - 1):
                            et = activities[i].end_time
                            if et and et.date() == activities[i + 1].start_time.date():
                                break_time += (
                                    activities[i + 1].start_time.timestamp()
                                    - et.timestamp()
                                )

                        # Calculate inside time
                        for i in range(len(activities)):
                            st, et = activities[i].start_time, activities[i].end_time
                            if not et:
                                et = timezone.now()
                            inside_time += et.timestamp() - st.timestamp()

                        break_time_s = sToTime(break_time)
                        inside_time_s = sToTime(inside_time)
                        employee_is_lead = emp.lead or emp.manager

                        # Check for late attendance
                        is_late = False
                        if start_time:
                            start_time_timeobj = start_time.time()
                            late_time_change_date = datetime.datetime.strptime(
                                "2025-02-11", "%Y-%m-%d"
                            ).date()
                            today = attendance.date

                            if today >= late_time_change_date:
                                is_late = (
                                    employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour == 11
                                            and start_time_timeobj.minute > 10
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                ) or (
                                    not employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour >= 11
                                            and start_time_timeobj.minute > 10
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                )
                            else:
                                is_late = (
                                    employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour == 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                ) or (
                                    not employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour >= 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                )

                        # Update attendance data
                        temp[date].update(
                            {
                                "entry_time": start_time.time() if start_time else "-",
                                "exit_time": end_time.time() if end_time else "-",
                                "is_updated_by_bot": is_updated_by_bot,
                                "break_time": break_time_s,
                                "break_time_hour": math.floor(
                                    (break_time / (60 * 60)) % 24
                                ),
                                "break_time_minute": math.floor(break_time / 60),
                                "inside_time": inside_time_s,
                                "inside_time_hour": math.floor(
                                    (inside_time / (60 * 60)) % 24
                                ),
                                "inside_time_minute": math.floor(inside_time / 60),
                                "total_time": sToTime(inside_time + break_time),
                                "total_time_hour": math.floor(
                                    (inside_time + break_time) / (60 * 60) % 24
                                ),
                                "employee_is_lead": employee_is_lead,
                                "is_late": is_late,
                            }
                        )

            # Add manager hours if available
            if emp.id in manager_date_and_hours:
                for date in last_x_dates:
                    accepted_hour = temp[date].get("accepted_hour", 0)
                    manager_hour = manager_date_and_hours[emp.id].get(date, 0)
                    temp[date]["accepted_hour"] = accepted_hour
                    temp[date]["manager_hour"] = manager_hour

            date_datas[emp] = temp

        # Handle sorting if requested
        o = request.GET.get("o", None)
        if o:
            if o == "entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), {})
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                )
                o = "-entry"
            elif o == "-entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), {})
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                    reverse=True,
                )
                o = "entry"

            date_datas = dict(date_datas_sorted)

        # Calculate online status form visibility
        online_status_form = str(request.user.employee.id) not in management_ids

        # Prepare context and render response
        context = dict(
            self.admin_site.each_context(request),
            dates=last_x_dates,
            last_month=last_month,
            date_datas=date_datas,
            o=o,  # order key
            online_status_form=online_status_form,
        )

        return TemplateResponse(
            request, "admin/employee/employee_attendance.html", context
        )

    def _check_if_late(self, start_time, is_lead):
        """Helper method to check if employee is late"""
        if not start_time:
            return False
        late_change_date = datetime.date(2025, 2, 11)
        hour, minute = start_time.hour, start_time.minute
        late_limit = (11, 10) if start_time.date() >= late_change_date else (11, 30)
        return (hour > late_limit[0]) or (
            hour == late_limit[0] and minute > late_limit[1]
        )

    def waqt_select(self, request, *args, **kwargs) -> redirect:
        if not request.user.is_authenticated:
            return redirect("/")
        if request.method == "POST":
            prayerobj = PrayerInfo.objects.filter(
                employee=request.user.employee, created_at__date=timezone.now().date()
            ).last()
            form = EmployeePrayerInfoForm(request.POST, instance=prayerobj)
            if form.is_valid():
                prayer_info = form.save(commit=False)
                prayer_info.employee = request.user.employee
                prayer_info.save()
                messages.success(request, "Submitted Successfully")

        return redirect("admin:employee_attendance")

    def get_urls(self):
        def wrap(view):
            def wrapper(*args, **kwargs):
                return self.admin_site.admin_view(view)(*args, **kwargs)

            wrapper.model_admin = self
            return update_wrapper(wrapper, view)

        info = self.model._meta.app_label, self.model._meta.model_name

        urls = super().get_urls()

        employee_online_urls = [
            path("admin/", wrap(self.changelist_view), name="%s_%s_changelist" % info),
            path("", self.custom_changelist_view, name="employee_attendance"),
            path("waqtselect/", self.waqt_select, name="waqt_select"),
            path(
                "generate-excel/",
                self.generate_monthly_attendance_report_xlsx,
                name="generate_monthly_attendance_report_xlsx",
            ),
        ]
        return employee_online_urls + urls

    # def has_module_permission(self, request):
    #     return False

    def generate_monthly_attendance_report_xlsx(self, request, *args, **kwargs):
        from openpyxl.writer.excel import save_virtual_workbook
        from openpyxl.styles import Alignment

        data = request.POST
        to_date = datetime.datetime.strptime(data.get("to_date"), "%Y-%m-%d")
        from_date = datetime.datetime.strptime(data.get("from_date"), "%Y-%m-%d")
        now = to_date
        DEFAULT_EXIT_HOUR = 12 + 8  # 24 Hour time == 9 pm
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_count = to_date - from_date
        day_range = day_count.days + 1

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_range)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]
        last_x_date = (now - datetime.timedelta(10)).date()
        last_month = (now.replace(day=1) - datetime.timedelta(days=1)).date()

        emps = (
            Employee.objects.filter(
                active=True,
                show_in_attendance_list=True,
            )
            .order_by("full_name")
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=last_month.year,
                        employeeattendance__date__month=last_month.month,
                    ),
                )
            )
        )

        emps = sorted(emps, key=lambda item: (item.is_online))
        user_data = None
        for index, emp in enumerate(emps):
            if emp.user == request.user:
                user_data = emps.pop(index)
                break
        if user_data:
            emps.insert(0, user_data)

        date_datas = {}

        manager_date_and_hours = {}

        for emp in emps:
            temp = {}
            attendances = emp.employeeattendance_set.all()

            empdailyhours = emp.dailyprojectupdate_employee.filter(
                created_at__date__gte=last_x_date,
                status="approved",
            )

            for date in last_x_dates:
                temp[date] = dict()

                for edh in reversed(empdailyhours):
                    if edh.created_at.date() == date:
                        if edh.manager != edh.employee:
                            manager_id = edh.manager.id

                            if manager_id not in manager_date_and_hours:
                                manager_date_and_hours[manager_id] = {}

                            if date not in manager_date_and_hours[manager_id]:
                                manager_date_and_hours[manager_id][date] = edh.hours
                            else:
                                manager_date_and_hours[manager_id][date] += edh.hours

                        if date not in temp:
                            temp[date] = {"accepted_hour": edh.hours}
                        else:
                            temp[date]["accepted_hour"] = (
                                temp[date].get("accepted_hour", 0) + edh.hours
                            )

                for attendance in attendances:
                    if attendance.date == date:
                        activities = attendance.employeeactivity_set.all()
                        if activities.exists():
                            activities = list(activities)
                            al = len(activities)
                            start_time = activities[0].start_time
                            end_time = activities[-1].end_time
                            is_updated_by_bot = activities[-1].is_updated_by_bot
                            break_time = 0
                            inside_time = 0

                            for i in range(al - 1):
                                et = activities[i].end_time
                                if (
                                    et
                                    and et.date() == activities[i + 1].start_time.date()
                                ):
                                    break_time += (
                                        activities[i + 1].start_time.timestamp()
                                        - et.timestamp()
                                    )
                            for i in range(al):
                                st, et = (
                                    activities[i].start_time,
                                    activities[i].end_time,
                                )
                                if not et:
                                    et = timezone.now()
                                inside_time += et.timestamp() - st.timestamp()

                            break_time_s = sToTime(break_time)
                            inside_time_s = sToTime(inside_time)
                            employee_is_lead = emp.lead or emp.manager
                            start_time_timeobj = start_time.time()
                            if start_time:
                                is_late = (
                                    employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour == 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                ) or (
                                    not employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour >= 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                )
                            temp[date].update(
                                {
                                    "entry_time": (
                                        start_time.time() if start_time else "-"
                                    ),
                                    "exit_time": end_time.time() if end_time else "-",
                                    "is_updated_by_bot": is_updated_by_bot,
                                    "break_time": break_time_s,
                                    "break_time_hour": math.floor(
                                        (break_time / (60 * 60)) % 24
                                    ),
                                    "break_time_minute": math.floor(break_time / 60),
                                    "inside_time": inside_time_s,
                                    "inside_time_hour": math.floor(
                                        (inside_time / (60 * 60)) % 24
                                    ),
                                    "inside_time_minute": math.floor(inside_time / 60),
                                    "total_time": sToTime(inside_time + break_time),
                                    "total_time_hour": math.floor(
                                        (inside_time + break_time) / (60 * 60) % 24
                                    ),
                                    "employee_is_lead": emp.lead or emp.manager,
                                    "is_late": is_late,
                                }
                            )
                        break
            date_datas.update({emp: temp})

        # for emp in emps:
        #         if manager_date_and_hours.get(emp.id):
        #                 print(date_datas[emp])
        #                 for date in last_x_dates:
        #                     if date_datas[emp][date].get('accepted_hour'):
        #                         date_datas[emp][date]['accepted_hour'] +=  manager_date_and_hours.get(emp.id).get(0, date)
        #                     else:
        #                         date_datas[emp][date]['accepted_hour'] = manager_date_and_hours.get(emp.id).get(0, date)
        for emp in emps:
            if manager_date_and_hours.get(emp.id):
                # print(date_datas[emp])
                for date in last_x_dates:
                    accepted_hour = date_datas[emp][date].get("accepted_hour", 0)

                    manager_hour = manager_date_and_hours.get(emp.id, {}).get(date, 0)

                    date_datas[emp][date]["accepted_hour"] = accepted_hour
                    date_datas[emp][date]["manager_hour"] = manager_hour

        online_status_form = False
        if not str(request.user.employee.id) in management_ids:
            online_status_form = True

        o = request.GET.get("o", None)

        if o:
            if o == "entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                )
                o = "-entry"
            elif o == "-entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                    reverse=True,
                )
                o = "entry"

            date_datas = dict(date_datas_sorted)
        context = dict(
            self.admin_site.each_context(request),
            dates=last_x_dates,
            last_month=last_month,
            date_datas=date_datas,
            o=o,  # order key
            online_status_form=online_status_form,
        )
        wb = Workbook()
        attendance_sheet = wb.create_sheet(title="Employee Attendance")
        attendance_sheet.append(
            [
                "Employee",
                "Date",
                "Entry Time",
                "Exit Time",
                "Break Time",
                "Inside Hours",
                "Total Hours",
            ]
        )
        row_num = 2
        for employee, all_data in date_datas.items():
            first_row = row_num
            for date, data in all_data.items():
                break_time = f"{data.get('break_time_hour', 0)}h: {data.get('break_time_minute', 0)}m"
                inside_time = f"{data.get('inside_time_hour', 0)}h: {data.get('inside_time_minute', 0)}m"
                attendance_sheet.append(
                    [
                        employee.full_name if row_num == first_row else "",
                        date.strftime("%d/%m/%Y"),
                        data.get("entry_time", None),
                        data.get("exit_time", None),
                        data.get("break_time", 0),
                        data.get("inside_time", 0),
                        data.get("total_time", 0),
                    ]
                )
                row_num += 1

            attendance_sheet.merge_cells(f"A{first_row}:A{row_num-1}")

            # Center the text in the merged cell
            attendance_sheet[f"A{first_row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        columns = ["A", "B", "C", "D", "E", "F", "G"]
        for col in columns:
            attendance_sheet.column_dimensions[col].width = 20
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = (
            "attachment; filename=Employee-Attendance.xlsx"
        )
        return response
