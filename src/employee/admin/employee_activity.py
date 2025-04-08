import datetime
import math
from functools import update_wrapper

from django.contrib import admin, messages
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse

# Needed for optional Features
# from django.db.models import Count, Case, When, Value, BooleanField
from django.shortcuts import redirect
from django.template.loader import get_template
from django.template.response import TemplateResponse
from django.urls import path
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.html import format_html
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook

from config.settings import employee_ids as management_ids
from employee.forms.prayer_info import EmployeePrayerInfoForm
from employee.models import (
    Employee,
    EmployeeActivity,
    EmployeeAttendance,
    EmployeeOnline,
    EmployeeSkill,
    PrayerInfo,
)
from employee.models.employee_activity import EmployeeProject, TrialEmployeeAttendance


def sToTime(duration):
    minutes = math.floor((duration / 60) % 60)
    hours = math.floor((duration / (60 * 60)) % 24)

    return f"{hours:01}h: {minutes:01}m"


@admin.register(EmployeeOnline)
class EmployeeOnlineAdmin(admin.ModelAdmin):
    list_display = ("employee", "get_status", "active")
    list_editable = ("active",)
    list_filter = ("active",)
    search_fields = ("employee__full_name",)
    autocomplete_fields = ("employee",)

    def get_queryset(self, request):
        EmployeeOnline.objects.filter()
        query_set = super(EmployeeOnlineAdmin, self).get_queryset(request)
        if not request.user.is_superuser and not request.user.has_perm(
            "employee.can_see_all_break"
        ):
            return query_set.filter(employee=request.user.employee.id)
        return query_set.filter(employee__active=True).exclude(
            employee_id__in=[7, 30, 76, 49]
        )

    @admin.display(description="Status")
    def get_status(self, obj):
        html_template = get_template("admin/employee_online/list/col_status.html")
        html_content = html_template.render({"employee_online": obj})
        return format_html(html_content)

    def get_urls(self):
        urls = super(EmployeeOnlineAdmin, self).get_urls()

        employee_online_urls = [
            path(
                "employee-online-graph/",
                self.admin_site.admin_view(self.graph),
                name="employee.online.graph",
            ),
            # path('employee-online-graph/<str:date>/', self.admin_site.admin_view(self.graph)),
            path("save-employee-online-status/", self.save_status),
        ]

        return employee_online_urls + urls

    @csrf_exempt
    def save_status(self, request):
        if request.method == "POST":
            data = request.POST
            print(data.get("status"))
        return JsonResponse({"status": 200})

    def graph(self, request, *args, **kwargs):
        target_date = timezone.now().date()

        if "date" in request.GET and parse_date(request.GET.get("date")):
            target_date = parse_date(request.GET.get("date"))

        employee_attendance = EmployeeAttendance.objects.filter(date=target_date).all()
        graph_data = []

        for attendance in employee_attendance:
            element = []
            breaks_timestamp = []
            for employee_break in attendance.employeeactivity_set.all():
                breaks_timestamp.append(
                    int(employee_break.start_time.timestamp() * 1000)
                )
                if employee_break.end_time:
                    breaks_timestamp.append(
                        int(employee_break.end_time.timestamp() * 1000)
                    )
                else:
                    breaks_timestamp.append((int(timezone.now().timestamp() * 1000)))

            element.append(attendance.employee.full_name)
            element.append(breaks_timestamp)

            graph_data.append(element)

        context = dict(self.admin_site.each_context(request), graph_data=graph_data)
        return TemplateResponse(
            request, "admin/employee_online/graph.html", context=context
        )

    def has_module_permission(self, request):
        return False


@admin.register(EmployeeAttendance)
class EmployeeAttendanceAdmin(admin.ModelAdmin):
    list_display = ("date", "employee", "entry_time", "exit_time")
    date_hierarchy = "date"
    list_filter = ("employee",)
    search_fields = ("employee__full_name", "date")
    autocomplete_fields = ("employee",)

    def custom_changelist_view(self, request, *args, **kwargs) -> TemplateResponse:
        if not request.user.has_perm("employee.view_employeeattendance"):
            return super().changelist_view(request, *args, **kwargs)

        if not request.user.is_authenticated:
            return redirect("/")

        now = timezone.now()
        DEFAULT_EXIT_HOUR = 12 + 8  # 24 Hour time == 9 pm
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_range = 30
        if request.user.has_perm("employee.can_see_full_month_attendance"):
            day_range = 30

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_range)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]
        last_x_date: datetime._Date = (now - datetime.timedelta(day_range)).date()
        last_month = (now.replace(day=1) - datetime.timedelta(days=1)).date()

        # # Filter employees based on user permissions
        # if request.user.is_superuser or request.user.has_perm('employee.view_all_attendance'):
        #     emps = Employee.objects.filter(
        #         active=True,
        #         show_in_attendance_list=True,
        #     ).order_by("full_name")
        # else:
        #     # Only show the current user's employee record
        #     emps = Employee.objects.filter(
        #         active=True,
        #         show_in_attendance_list=True,
        #         user=request.user
        #     ).order_by("full_name")
        #
        # emps = emps.annotate(
        #     last_month_attendance=Count(
        #         "employeeattendance",
        #         filter=Q(
        #             employeeattendance__date__year=last_month.year,
        #             employeeattendance__date__month=last_month.month,
        #         ),
        #     )
        # )

        emps = (
            Employee.objects.filter(
                active=True,
                show_in_attendance_list=True,
            )
            .order_by("full_name")
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=last_month.year,
                        employeeattendance__date__month=last_month.month,
                    ),
                )
            )
        )

        emps = sorted(emps, key=lambda item: (item.is_online))
        user_data = None
        for index, emp in enumerate(emps):
            if emp.user == request.user:
                user_data = emps.pop(index)
                break
        if user_data:
            emps.insert(0, user_data)

        date_datas = {}

        manager_date_and_hours = {}

        for emp in emps:
            temp = {}
            attendances = emp.employeeattendance_set.all()

            empdailyhours = emp.dailyprojectupdate_employee.filter(
                created_at__date__gte=last_x_date,
                status="approved",
            )

            for date in last_x_dates:
                temp[date] = dict()

                for edh in reversed(empdailyhours):
                    if edh.created_at.date() == date:
                        if edh.manager != edh.employee:
                            manager_id = edh.manager.id

                            if manager_id not in manager_date_and_hours:
                                manager_date_and_hours[manager_id] = {}

                            if date not in manager_date_and_hours[manager_id]:
                                manager_date_and_hours[manager_id][date] = edh.hours
                            else:
                                manager_date_and_hours[manager_id][date] += edh.hours

                        if date not in temp:
                            temp[date] = {"accepted_hour": edh.hours}
                        else:
                            temp[date]["accepted_hour"] = (
                                temp[date].get("accepted_hour", 0) + edh.hours
                            )
                avg_total = 0
                for attendance in attendances:
                    if attendance.date == date:
                        activities = attendance.employeeactivity_set.all()
                        if activities.exists():
                            activities = list(activities)
                            al = len(activities)
                            start_time = activities[0].start_time
                            end_time = activities[-1].end_time
                            is_updated_by_bot = activities[-1].is_updated_by_bot
                            break_time = 0
                            inside_time = 0

                            for i in range(al - 1):
                                et = activities[i].end_time
                                if (
                                    et
                                    and et.date() == activities[i + 1].start_time.date()
                                ):
                                    break_time += (
                                        activities[i + 1].start_time.timestamp()
                                        - et.timestamp()
                                    )
                            for i in range(al):
                                st, et = (
                                    activities[i].start_time,
                                    activities[i].end_time,
                                )
                                if not et:
                                    et = timezone.now()
                                inside_time += et.timestamp() - st.timestamp()

                            break_time_s = sToTime(break_time)
                            inside_time_s = sToTime(inside_time)
                            employee_is_lead = emp.lead or emp.manager
                            start_time_timeobj = start_time.time()
                            is_late = False
                            late_time_change_date = datetime.datetime.strptime(
                                "2025-02-11", "%Y-%m-%d"
                            ).date()  # it may change if late time change
                            today = attendance.date
                            if start_time:
                                if today >= late_time_change_date:
                                    is_late = (
                                        employee_is_lead
                                        and (
                                            (
                                                start_time_timeobj.hour == 11
                                                and start_time_timeobj.minute > 10
                                            )
                                            or start_time_timeobj.hour >= 12
                                        )
                                    ) or (
                                        not employee_is_lead
                                        and (
                                            (
                                                start_time_timeobj.hour >= 11
                                                and start_time_timeobj.minute > 10
                                            )
                                            or start_time_timeobj.hour >= 12
                                        )
                                    )
                                else:
                                    is_late = (
                                        employee_is_lead
                                        and (
                                            (
                                                start_time_timeobj.hour == 11
                                                and start_time_timeobj.minute > 30
                                            )
                                            or start_time_timeobj.hour >= 12
                                        )
                                    ) or (
                                        not employee_is_lead
                                        and (
                                            (
                                                start_time_timeobj.hour >= 11
                                                and start_time_timeobj.minute > 30
                                            )
                                            or start_time_timeobj.hour >= 12
                                        )
                                    )
                            temp[date].update(
                                {
                                    "entry_time": (
                                        start_time.time() if start_time else "-"
                                    ),
                                    "exit_time": end_time.time() if end_time else "-",
                                    "is_updated_by_bot": is_updated_by_bot,
                                    "break_time": break_time_s,
                                    "break_time_hour": math.floor(
                                        (break_time / (60 * 60)) % 24
                                    ),
                                    "break_time_minute": math.floor(break_time / 60),
                                    "inside_time": inside_time_s,
                                    "inside_time_hour": math.floor(
                                        (inside_time / (60 * 60)) % 24
                                    ),
                                    "inside_time_minute": math.floor(inside_time / 60),
                                    "total_time": sToTime(inside_time + break_time),
                                    "total_time_hour": math.floor(
                                        (inside_time + break_time) / (60 * 60) % 24
                                    ),
                                    "employee_is_lead": emp.lead or emp.manager,
                                    "is_late": is_late,
                                }
                            )
                        break
            date_datas.update({emp: temp})

        # for emp in emps:
        #         if manager_date_and_hours.get(emp.id):
        #                 print(date_datas[emp])
        #                 for date in last_x_dates:
        #                     if date_datas[emp][date].get('accepted_hour'):
        #                         date_datas[emp][date]['accepted_hour'] +=  manager_date_and_hours.get(emp.id).get(0, date)
        #                     else:
        #                         date_datas[emp][date]['accepted_hour'] = manager_date_and_hours.get(emp.id).get(0, date)
        for emp in emps:
            if manager_date_and_hours.get(emp.id):
                # print(date_datas[emp])
                for date in last_x_dates:
                    accepted_hour = date_datas[emp][date].get("accepted_hour", 0)

                    manager_hour = manager_date_and_hours.get(emp.id, {}).get(date, 0)

                    date_datas[emp][date]["accepted_hour"] = accepted_hour
                    date_datas[emp][date]["manager_hour"] = manager_hour

        online_status_form = False
        if str(request.user.employee.id) not in management_ids:
            online_status_form = True

        o = request.GET.get("o", None)

        if o:
            if o == "entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                )
                o = "-entry"
            elif o == "-entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                    reverse=True,
                )
                o = "entry"

            date_datas = dict(date_datas_sorted)
            print("*" * 100)
            print(last_x_dates)

        employee_avg_hours_dict = {}
        # for emp, dates_data in date_datas.items():
        #     total_inside_hours = Decimal('0.0')
        #     days_with_data = 0
        #
        #     for date, data in dates_data.items():
        #         if isinstance(data, dict) and 'inside_time_hour' in data:
        #             hours = Decimal(str(data['inside_time_hour']))
        #             minutes = Decimal(str(data.get('inside_time_minute', 0))) / Decimal('60')
        #             total_inside_hours += hours + minutes
        #             days_with_data += 1
        #
        #     if days_with_data > 0:
        #         avg_hours = (total_inside_hours / Decimal(str(days_with_data))).quantize(Decimal('0.01'))
        #         # Store the average directly in date_datas
        #         date_datas[emp]['avg_hours'] = avg_hours
        context = dict(
            self.admin_site.each_context(request),
            dates=last_x_dates,
            last_month=last_month,
            date_datas=date_datas,
            o=o,  # order key
            online_status_form=online_status_form,
        )
        return TemplateResponse(
            request, "admin/employee/employee_attendance.html", context
        )

    def waqt_select(self, request, *args, **kwargs) -> redirect:
        if not request.user.is_authenticated:
            return redirect("/")
        if request.method == "POST":
            prayerobj = PrayerInfo.objects.filter(
                employee=request.user.employee, created_at__date=timezone.now().date()
            ).last()
            form = EmployeePrayerInfoForm(request.POST, instance=prayerobj)
            if form.is_valid():
                prayer_info = form.save(commit=False)
                prayer_info.employee = request.user.employee
                prayer_info.save()
                messages.success(request, "Submitted Successfully")

        return redirect("admin:employee_attendance")

    def get_urls(self):
        def wrap(view):
            def wrapper(*args, **kwargs):
                return self.admin_site.admin_view(view)(*args, **kwargs)

            wrapper.model_admin = self
            return update_wrapper(wrapper, view)

        info = self.model._meta.app_label, self.model._meta.model_name

        urls = super(EmployeeAttendanceAdmin, self).get_urls()

        employee_online_urls = [
            path("admin/", wrap(self.changelist_view), name="%s_%s_changelist" % info),
            path("", self.custom_changelist_view, name="employee_attendance"),
            path("waqtselect/", self.waqt_select, name="waqt_select"),
            path(
                "generate-excel/",
                self.generate_monthly_attendance_report_xlsx,
                name="generate_monthly_attendance_report_xlsx",
            ),
        ]
        return employee_online_urls + urls

    # def has_module_permission(self, request):
    #     return False

    def generate_monthly_attendance_report_xlsx(self, request, *args, **kwargs):
        from openpyxl.styles import Alignment
        from openpyxl.writer.excel import save_virtual_workbook

        data = request.POST
        to_date = datetime.datetime.strptime(data.get("to_date"), "%Y-%m-%d")
        from_date = datetime.datetime.strptime(data.get("from_date"), "%Y-%m-%d")
        now = to_date
        DEFAULT_EXIT_HOUR = 12 + 8  # 24 Hour time == 9 pm
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_count = to_date - from_date
        day_range = day_count.days + 1

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_range)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]
        last_x_date = (now - datetime.timedelta(10)).date()
        last_month = (now.replace(day=1) - datetime.timedelta(days=1)).date()

        emps = (
            Employee.objects.filter(
                active=True,
                show_in_attendance_list=True,
            )
            .order_by("full_name")
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=last_month.year,
                        employeeattendance__date__month=last_month.month,
                    ),
                )
            )
        )

        emps = sorted(emps, key=lambda item: (item.is_online))
        user_data = None
        for index, emp in enumerate(emps):
            if emp.user == request.user:
                user_data = emps.pop(index)
                break
        if user_data:
            emps.insert(0, user_data)

        date_datas = {}

        manager_date_and_hours = {}

        for emp in emps:
            temp = {}
            attendances = emp.employeeattendance_set.all()

            empdailyhours = emp.dailyprojectupdate_employee.filter(
                created_at__date__gte=last_x_date,
                status="approved",
            )

            for date in last_x_dates:
                temp[date] = dict()

                for edh in reversed(empdailyhours):
                    if edh.created_at.date() == date:
                        if edh.manager != edh.employee:
                            manager_id = edh.manager.id

                            if manager_id not in manager_date_and_hours:
                                manager_date_and_hours[manager_id] = {}

                            if date not in manager_date_and_hours[manager_id]:
                                manager_date_and_hours[manager_id][date] = edh.hours
                            else:
                                manager_date_and_hours[manager_id][date] += edh.hours

                        if date not in temp:
                            temp[date] = {"accepted_hour": edh.hours}
                        else:
                            temp[date]["accepted_hour"] = (
                                temp[date].get("accepted_hour", 0) + edh.hours
                            )

                for attendance in attendances:
                    if attendance.date == date:
                        activities = attendance.employeeactivity_set.all()
                        if activities.exists():
                            activities = list(activities)
                            al = len(activities)
                            start_time = activities[0].start_time
                            end_time = activities[-1].end_time
                            is_updated_by_bot = activities[-1].is_updated_by_bot
                            break_time = 0
                            inside_time = 0

                            for i in range(al - 1):
                                et = activities[i].end_time
                                if (
                                    et
                                    and et.date() == activities[i + 1].start_time.date()
                                ):
                                    break_time += (
                                        activities[i + 1].start_time.timestamp()
                                        - et.timestamp()
                                    )
                            for i in range(al):
                                st, et = (
                                    activities[i].start_time,
                                    activities[i].end_time,
                                )
                                if not et:
                                    et = timezone.now()
                                inside_time += et.timestamp() - st.timestamp()

                            break_time_s = sToTime(break_time)
                            inside_time_s = sToTime(inside_time)
                            employee_is_lead = emp.lead or emp.manager
                            start_time_timeobj = start_time.time()
                            if start_time:
                                is_late = (
                                    employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour == 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                ) or (
                                    not employee_is_lead
                                    and (
                                        (
                                            start_time_timeobj.hour >= 11
                                            and start_time_timeobj.minute > 30
                                        )
                                        or start_time_timeobj.hour >= 12
                                    )
                                )
                            temp[date].update(
                                {
                                    "entry_time": (
                                        start_time.time() if start_time else "-"
                                    ),
                                    "exit_time": end_time.time() if end_time else "-",
                                    "is_updated_by_bot": is_updated_by_bot,
                                    "break_time": break_time_s,
                                    "break_time_hour": math.floor(
                                        (break_time / (60 * 60)) % 24
                                    ),
                                    "break_time_minute": math.floor(break_time / 60),
                                    "inside_time": inside_time_s,
                                    "inside_time_hour": math.floor(
                                        (inside_time / (60 * 60)) % 24
                                    ),
                                    "inside_time_minute": math.floor(inside_time / 60),
                                    "total_time": sToTime(inside_time + break_time),
                                    "total_time_hour": math.floor(
                                        (inside_time + break_time) / (60 * 60) % 24
                                    ),
                                    "employee_is_lead": emp.lead or emp.manager,
                                    "is_late": is_late,
                                }
                            )
                        break
            date_datas.update({emp: temp})

        # for emp in emps:
        #         if manager_date_and_hours.get(emp.id):
        #                 print(date_datas[emp])
        #                 for date in last_x_dates:
        #                     if date_datas[emp][date].get('accepted_hour'):
        #                         date_datas[emp][date]['accepted_hour'] +=  manager_date_and_hours.get(emp.id).get(0, date)
        #                     else:
        #                         date_datas[emp][date]['accepted_hour'] = manager_date_and_hours.get(emp.id).get(0, date)
        for emp in emps:
            if manager_date_and_hours.get(emp.id):
                # print(date_datas[emp])
                for date in last_x_dates:
                    accepted_hour = date_datas[emp][date].get("accepted_hour", 0)

                    manager_hour = manager_date_and_hours.get(emp.id, {}).get(date, 0)

                    date_datas[emp][date]["accepted_hour"] = accepted_hour
                    date_datas[emp][date]["manager_hour"] = manager_hour

        online_status_form = False
        if str(request.user.employee.id) not in management_ids:
            online_status_form = True

        o = request.GET.get("o", None)

        if o:
            if o == "entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                )
                o = "-entry"
            elif o == "-entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                    reverse=True,
                )
                o = "entry"

            date_datas = dict(date_datas_sorted)
        context = dict(
            self.admin_site.each_context(request),
            dates=last_x_dates,
            last_month=last_month,
            date_datas=date_datas,
            o=o,  # order key
            online_status_form=online_status_form,
        )
        wb = Workbook()
        attendance_sheet = wb.create_sheet(title="Employee Attendance")
        attendance_sheet.append(
            [
                "Employee",
                "Date",
                "Entry Time",
                "Exit Time",
                "Break Time",
                "Inside Hours",
                "Total Hours",
            ]
        )
        row_num = 2
        for employee, all_data in date_datas.items():
            first_row = row_num
            for date, data in all_data.items():
                break_time = f"{data.get('break_time_hour', 0)}h: {data.get('break_time_minute', 0)}m"
                inside_time = f"{data.get('inside_time_hour', 0)}h: {data.get('inside_time_minute', 0)}m"
                attendance_sheet.append(
                    [
                        employee.full_name if row_num == first_row else "",
                        date.strftime("%d/%m/%Y"),
                        data.get("entry_time", None),
                        data.get("exit_time", None),
                        data.get("break_time", 0),
                        data.get("inside_time", 0),
                        data.get("total_time", 0),
                    ]
                )
                row_num += 1

            attendance_sheet.merge_cells(f"A{first_row}:A{row_num-1}")

            # Center the text in the merged cell
            attendance_sheet[f"A{first_row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        columns = ["A", "B", "C", "D", "E", "F", "G"]
        for col in columns:
            attendance_sheet.column_dimensions[col].width = 20
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = (
            "attachment; filename=Employee-Attendance.xlsx"
        )
        return response


"""
A trial employee attendance model is created to track the attendance of trial employees.
"""


@admin.register(TrialEmployeeAttendance)
class TrialEmployeeAttendanceAdmin(admin.ModelAdmin):
    list_display = ("date", "employee", "entry_time", "exit_time")
    date_hierarchy = "date"
    list_filter = ("employee",)
    search_fields = ("employee__full_name", "date")
    autocomplete_fields = ("employee",)

    def custom_changelist_view(self, request, *args, **kwargs) -> TemplateResponse:
        if not request.user.has_perm("employee.view_employeeattendance"):
            return super().changelist_view(request, *args, **kwargs)

        if not request.user.is_authenticated:
            return redirect("/")

        now = timezone.now()
        DEFAULT_EXIT_HOUR = 20  # 8 PM
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_range = (
            30
            if request.user.has_perm("employee.can_see_full_month_attendance")
            else 30
        )

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_range)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]

        emps = (
            Employee.objects.filter(active=True, show_in_attendance_list=True)
            .order_by("full_name")
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=now.year,
                        employeeattendance__date__month=now.month - 1,
                    ),
                )
            )
        )

        # Move the current user to the top of the list
        user_data = next((emp for emp in emps if emp.user == request.user), None)
        if user_data:
            emps = [user_data] + [emp for emp in emps if emp != user_data]

        date_datas = self.get_employee_attendance_data(emps, last_x_dates)

        context = {
            **self.admin_site.each_context(request),
            "dates": last_x_dates,
            "date_datas": date_datas,
            "online_status_form": str(request.user.employee.id) not in management_ids,
        }
        return TemplateResponse(
            request, "admin/employee/employee_attendance.html", context
        )

    def get_employee_attendance_data(self, emps, last_x_dates):
        date_datas = {}
        manager_date_and_hours = {}

        for emp in emps:
            temp = {date: {} for date in last_x_dates}
            attendances = emp.employeeattendance_set.all()
            empdailyhours = emp.dailyprojectupdate_employee.filter(
                created_at__date__gte=last_x_dates[-1],
                status="approved",
            )

            for date in last_x_dates:
                # Process daily hours
                for edh in reversed(empdailyhours):
                    if edh.created_at.date() == date:
                        manager_id = edh.manager.id
                        manager_date_and_hours.setdefault(manager_id, {}).setdefault(
                            date, 0
                        )
                        manager_date_and_hours[manager_id][date] += edh.hours

                        temp[date]["accepted_hour"] = (
                            temp[date].get("accepted_hour", 0) + edh.hours
                        )
                        break

                # Process attendance
                for attendance in attendances:
                    if attendance.date == date:
                        activities = attendance.employeeactivity_set.all()
                        if activities.exists():
                            activities = list(activities)
                            if len(activities) > 0:  # Check if the list is not empty
                                start_time = activities[0].start_time
                                end_time = activities[-1].end_time
                                is_updated_by_bot = activities[-1].is_updated_by_bot
                                time_data = self.calculate_times(
                                    activities
                                )  # Get the full dictionary
                                break_time = time_data["break_time_str"]
                                inside_time = time_data["inside_time_str"]
                                temp[date].update(
                                    {
                                        "entry_time": (
                                            start_time.time() if start_time else "-"
                                        ),
                                        "exit_time": (
                                            end_time.time() if end_time else "-"
                                        ),
                                        "is_updated_by_bot": is_updated_by_bot,
                                        "break_time": break_time,
                                        "inside_time": inside_time,
                                        "total_time": sToTime(inside_time + break_time),
                                    }
                                )
                        break

            date_datas[emp] = temp

        # Update accepted hours with manager hours
        for emp in emps:
            if emp.id in manager_date_and_hours:
                for date in last_x_dates:
                    accepted_hour = date_datas[emp][date].get("accepted_hour", 0)
                    manager_hour = manager_date_and_hours[emp.id].get(date, 0)
                    date_datas[emp][date]["accepted_hour"] = accepted_hour
                    date_datas[emp][date]["manager_hour"] = manager_hour

        return date_datas

    def calculate_times(self, activities):
        break_time = 0
        inside_time = 0

        # Existing time calculation logic
        for i in range(len(activities) - 1):
            et = activities[i].end_time
            if et and et.date() == activities[i + 1].start_time.date():
                break_time += activities[i + 1].start_time.timestamp() - et.timestamp()

        for activity in activities:
            st, et = activity.start_time, activity.end_time
            if not et:
                et = timezone.now()
            inside_time += et.timestamp() - st.timestamp()

        # Return both formatted and numerical values
        return {
            "break_time_str": sToTime(break_time),
            "inside_time_str": sToTime(inside_time),
            "break_time_hours": break_time / 3600,
            "break_time_minutes": (break_time % 3600) / 60,
            "inside_time_hours": inside_time / 3600,
            "inside_time_minutes": (inside_time % 3600) / 60,
            "total_time": (break_time + inside_time) / 3600,
        }

    def waqt_select(self, request, *args, **kwargs) -> redirect:
        if not request.user.is_authenticated:
            return redirect("/")
        if request.method == "POST":
            prayerobj = PrayerInfo.objects.filter(
                employee=request.user.employee, created_at__date=timezone.now().date()
            ).last()
            form = EmployeePrayerInfoForm(request.POST, instance=prayerobj)
            if form.is_valid():
                prayer_info = form.save(commit=False)
                prayer_info.employee = request.user.employee
                prayer_info.save()
                messages.success(request, "Submitted Successfully")

        return redirect("admin:employee_attendance")

    def get_urls(self):
        def wrap(view):
            def wrapper(*args, **kwargs):
                return self.admin_site.admin_view(view)(*args, **kwargs)

            wrapper.model_admin = self
            return update_wrapper(wrapper, view)

        info = self.model._meta.app_label, self.model._meta.model_name

        urls = super(TrialEmployeeAttendanceAdmin, self).get_urls()

        employee_online_urls = [
            path("admin/", wrap(self.changelist_view), name="%s_%s_changelist" % info),
            path("", self.custom_changelist_view, name="employee_attendance"),
            path("waqtselect/", self.waqt_select, name="waqt_select"),
            path(
                "generate-excel/",
                self.generate_monthly_attendance_report_xlsx,
                name="generate_monthly_attendance_report_xlsx",
            ),
        ]
        return employee_online_urls + urls

    def generate_monthly_attendance_report_xlsx(self, request, *args, **kwargs):
        from openpyxl.styles import Alignment
        from openpyxl.writer.excel import save_virtual_workbook

        data = request.POST
        to_date = datetime.datetime.strptime(data.get("to_date"), "%Y-%m-%d")
        from_date = datetime.datetime.strptime(data.get("from_date"), "%Y-%m-%d")
        now = to_date
        DEFAULT_EXIT_HOUR = 20  # 8 PM
        DEFAULT_EXIT_TIME = now.replace(hour=DEFAULT_EXIT_HOUR, minute=0, second=0)
        day_count = (to_date - from_date).days + 1

        last_x_dates = [
            (now - datetime.timedelta(i)).date()
            for i in range(day_count)
            if (now - datetime.timedelta(i)).date().strftime("%a") not in ["Sat", "Sun"]
        ]
        last_x_date = (now - datetime.timedelta(10)).date()
        last_month = (now.replace(day=1) - datetime.timedelta(days=1)).date()

        emps = (
            Employee.objects.filter(
                active=True,
                show_in_attendance_list=True,
            )
            .order_by("full_name")
            .annotate(
                last_month_attendance=Count(
                    "employeeattendance",
                    filter=Q(
                        employeeattendance__date__year=last_month.year,
                        employeeattendance__date__month=last_month.month,
                    ),
                )
            )
        )

        emps = sorted(emps, key=lambda item: (item.is_online))
        user_data = None
        for index, emp in enumerate(emps):
            if emp.user == request.user:
                user_data = emps.pop(index)
                break
        if user_data:
            emps.insert(0, user_data)

        date_datas = {}
        manager_date_and_hours = {}

        for emp in emps:
            temp = {}
            attendances = emp.employeeattendance_set.all()

            empdailyhours = emp.dailyprojectupdate_employee.filter(
                created_at__date__gte=last_x_date,
                status="approved",
            )

            for date in last_x_dates:
                temp[date] = dict()

                for edh in reversed(empdailyhours):
                    if edh.created_at.date() == date:
                        if edh.manager != edh.employee:
                            manager_id = edh.manager.id

                            if manager_id not in manager_date_and_hours:
                                manager_date_and_hours[manager_id] = {}

                            if date not in manager_date_and_hours[manager_id]:
                                manager_date_and_hours[manager_id][date] = edh.hours
                            else:
                                manager_date_and_hours[manager_id][date] += edh.hours

                        if date not in temp:
                            temp[date] = {"accepted_hour": edh.hours}
                        else:
                            temp[date]["accepted_hour"] = (
                                temp[date].get("accepted_hour", 0) + edh.hours
                            )

                for attendance in attendances:
                    if attendance.date == date:
                        activities = attendance.employeeactivity_set.all()
                        if activities.exists():
                            activities = list(activities)
                            if len(activities) > 0:  # Check if the list is not empty
                                start_time = activities[0].start_time
                                end_time = activities[-1].end_time
                                is_updated_by_bot = activities[-1].is_updated_by_bot
                                time_data = self.calculate_times(activities)
                                break_time = time_data["break_time_str"]
                                inside_time = time_data["inside_time_str"]
                                temp[date].update(
                                    {
                                        **time_data,
                                        "entry_time": (
                                            start_time.time() if start_time else "-"
                                        ),
                                        "exit_time": (
                                            end_time.time() if end_time else "-"
                                        ),
                                        "is_updated_by_bot": is_updated_by_bot,
                                    }
                                )

                                temp[date].update(
                                    {
                                        "entry_time": (
                                            start_time.time() if start_time else "-"
                                        ),
                                        "exit_time": (
                                            end_time.time() if end_time else "-"
                                        ),
                                        "is_updated_by_bot": is_updated_by_bot,
                                        "break_time": break_time,
                                        "inside_time": inside_time,
                                        "total_time": sToTime(inside_time + break_time),
                                    }
                                )
                        break
            date_datas.update({emp: temp})

        for emp in emps:
            if manager_date_and_hours.get(emp.id):
                for date in last_x_dates:
                    accepted_hour = date_datas[emp][date].get("accepted_hour", 0)
                    manager_hour = manager_date_and_hours.get(emp.id, {}).get(date, 0)
                    date_datas[emp][date]["accepted_hour"] = accepted_hour
                    date_datas[emp][date]["manager_hour"] = manager_hour

        online_status_form = False
        if str(request.user.employee.id) not in management_ids:
            online_status_form = True

        o = request.GET.get("o", None)

        if o:
            if o == "entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                )
                o = "-entry"
            elif o == "-entry":
                date_datas_sorted = sorted(
                    date_datas.items(),
                    key=lambda x: x[-1]
                    .get(datetime.datetime.now().date(), datetime.datetime.now().date())
                    .get("entry_time", DEFAULT_EXIT_TIME.time()),
                    reverse=True,
                )
                o = "entry"

            date_datas = dict(date_datas_sorted)

        context = dict(
            self.admin_site.each_context(request),
            dates=last_x_dates,
            last_month=last_month,
            date_datas=date_datas,
            o=o,  # order key
            online_status_form=online_status_form,
        )

        wb = Workbook()
        attendance_sheet = wb.create_sheet(title="Employee Attendance")
        attendance_sheet.append(
            [
                employee.full_name if row_num == first_row else "",
                date.strftime("%d/%m/%Y"),
                data.get("entry_time", None),
                data.get("exit_time", None),
                data.get("break_time_hour", 0)
                + data.get("break_time_minute", 0) / 60,  # Total hours as float
                data.get("inside_time_hour", 0)
                + data.get("inside_time_minute", 0) / 60,  # Total hours as float
                (data.get("inside_time_hour", 0) + data.get("break_time_hour", 0))
                + (data.get("inside_time_minute", 0) + data.get("break_time_minute", 0))
                / 60,
            ]
        )
        row_num = 2
        for employee, all_data in date_datas.items():
            first_row = row_num
            for date, data in all_data.items():
                break_time = f"{data.get('break_time_hour', 0)}h: {data.get('break_time_minute', 0)}m"
                inside_time = f"{data.get('inside_time_hour', 0)}h: {data.get('inside_time_minute', 0)}m"
                attendance_sheet.append(
                    [
                        employee.full_name if row_num == first_row else "",
                        date.strftime("%d/%m/%Y"),
                        data.get("entry_time", None),
                        data.get("exit_time", None),
                        data.get("break_time", 0),
                        data.get("inside_time", 0),
                        data.get("total_time", 0),
                    ]
                )
                row_num += 1

            attendance_sheet.merge_cells(f"A{first_row}:A{row_num-1}")

            # Center the text in the merged cell
            attendance_sheet[f"A{first_row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

        columns = ["A", "B", "C", "D", "E", "F", "G"]
        for col in columns:
            attendance_sheet.column_dimensions[col].width = 20
        wb.remove(wb["Sheet"])
        response = HttpResponse(
            content=save_virtual_workbook(wb), content_type="application/ms-excel"
        )
        response["Content-Disposition"] = (
            "attachment; filename=Employee-Attendance.xlsx"
        )
        return response


@admin.register(EmployeeActivity)
class EmployeeBreakAdmin(admin.ModelAdmin):
    list_display = ("employee_attendance", "get_start_time", "get_end_time")
    date_hierarchy = "employee_attendance__date"
    autocomplete_fields = ("employee_attendance",)
    list_filter = ("employee_attendance__employee",)
    search_fields = ("employee_attendance__employee__full_name",)
    # readonly_fields = ('start_by', 'end_by',)

    # @admin.display(description='Created By', ordering="created_by")
    # def get_created_by(self, obj):
    #     return obj.created_by.employee.full_name

    @admin.display(description="Start Time", ordering="start_time")
    def get_start_time(self, obj):
        start_time = ""

        if obj.start_time:
            start_time += obj.start_time.strftime("%b %d, %Y, %I:%M %p")

            if (
                obj.created_by
                and obj.created_by != obj.employee_attendance.employee.user
            ):
                start_time += (
                    '<span style="color: red; font-weight: bold;"> ('
                    + obj.created_by.employee.full_name
                    + ")</span>"
                )

        return format_html(start_time)

    @admin.display(description="End Time", ordering="end_time")
    def get_end_time(self, obj):
        end_time = ""

        if obj.end_time:
            end_time += obj.end_time.strftime("%b %d, %Y, %I:%M %p")

            if (
                obj.updated_by
                and obj.updated_by != obj.employee_attendance.employee.user
            ):
                end_time += (
                    '<span style="color: red; font-weight: bold;"> ('
                    + obj.updated_by.employee.full_name
                    + ")</span>"
                )

        return format_html(end_time)

    # To hide from main menu
    def has_module_permission(self, request):
        return False


@admin.register(EmployeeProject)
class EmployeeProjectAdmin(admin.ModelAdmin):
    list_display = ("employee", "get_projects")
    autocomplete_fields = ("employee", "project")
    list_filter = ("project",)
    search_fields = ("employee__full_name", "project__title")

    # Can be used when hide employees without projects
    # If used, no project can be assigned by admin due to onetoone field

    # def get_queryset(self, request):
    #     qs = super().get_queryset(request)
    #     return qs.annotate(
    #         project_count=Count("project"),
    #         project_exists=Case(
    #             When(project_count=0, then=Value(False)),
    #             default=Value(True),
    #             output_field=BooleanField()
    #         )
    #     ).filter(project_exists=True)

    @admin.display(description="Projects")
    def get_projects(self, obj):
        projects = " | ".join(
            obj.project.filter(active=True).values_list("title", flat=True)
        )
        if projects == "":
            projects = "-"
        return projects

    def has_module_permission(self, request):
        return False


@admin.register(PrayerInfo)
class EmployeePrayerInfoAdmin(admin.ModelAdmin):
    list_display = ("get_date", "employee", "num_of_waqt_done")
    autocomplete_fields = ("employee",)
    list_filter = ("employee",)
    search_fields = ("employee__full_name",)

    @admin.display(description="Date", ordering="created_at")
    def get_date(self, obj, *args, **kwargs):
        return obj.created_at.strftime("%b %d, %Y")

    def has_module_permission(self, request):
        return False


@admin.register(EmployeeSkill)
class EmployeeSkillAdmin(admin.ModelAdmin):
    list_display = ("employee", "skill", "percentage")
    search_fields = ("employee__name", "skill__title")
    list_filter = ("skill",)
    ordering = ("employee", "skill")

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.select_related("employee", "skill")
        return queryset

    def skill_title(self, obj):
        return obj.skill.title

    skill_title.admin_order_field = "skill__title"
    skill_title.short_description = "Skill Title"

    def employee_name(self, obj):
        return obj.employee.name

    employee_name.admin_order_field = "employee__name"
    employee_name.short_description = "Employee Name"
